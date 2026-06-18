"""
EXCEL PART SEARCH WEB APP dengan AUTO-LOADING + LOGIN SYSTEM + THRESHOLD + BATCH DOWNLOAD + EDIT POPULASI (GitHub Sync)
=============================================================
"""

import streamlit as st
import streamlit.components.v1 as _stc
import pandas as pd
import os
from pathlib import Path
from datetime import datetime
import warnings
from concurrent.futures import ThreadPoolExecutor, as_completed
import hashlib
import pickle
import hmac
import re
import io
import json
import requests


# ── Supabase (Auth + Permissions) ───────────────────────────────────
try:
    from supabase import (
        SUPABASE_ENABLED,
        load_users_from_supabase,
        authenticate_from_supabase,
        render_user_management_tab,
        SupabasePermissions,
        PERM_MENU, PERM_COLUMN, PERM_HARGA,
        SUPABASE_PERMS_ENABLED,
    )
except ImportError:
    SUPABASE_ENABLED           = False
    SUPABASE_PERMS_ENABLED     = False
    load_users_from_supabase   = None
    authenticate_from_supabase = None
    render_user_management_tab = None
    SupabasePermissions        = None

# ── User Monitoring (logging + dashboard) ────────────────────────────
try:
    from user_monitoring import (
        log_activity,
        touch_active,
        render_user_monitoring_tab,
    )
    USER_MONITORING_ENABLED = True
except Exception:
    USER_MONITORING_ENABLED = False
    def log_activity(*_a, **_kw): pass
    def touch_active(*_a, **_kw): pass
    def render_user_monitoring_tab():
        try:
            st.warning("⚠️ Modul `user_monitoring.py` tidak ditemukan.")
        except Exception:
            pass

# ── Admin Foto Part ──────────────────────────────────────────────────
try:
    from admin_foto_part import render_foto_part_tab, get_supabase_photo_urls
    FOTO_PART_ENABLED = True
except ImportError:
    FOTO_PART_ENABLED = False
    def render_foto_part_tab():
        st.warning("⚠️ `admin_foto_part.py` tidak ditemukan.")
    def get_supabase_photo_urls(pn: str) -> list:
        return []

# ── Image Search (Cari by Foto) ──────────────────────────────────────
try:
    from image_search import render_search_image_tab
    from admin_image_index import render_image_index_tab
    IMAGE_SEARCH_ENABLED = True
except ImportError as _ise:
    IMAGE_SEARCH_ENABLED = False
    _IMAGE_SEARCH_ERR = str(_ise)
    def render_search_image_tab():
        st.warning(f"⚠️ Fitur Cari by Foto tidak tersedia: {_IMAGE_SEARCH_ERR}")
    def render_image_index_tab():
        st.warning(f"⚠️ Tab Image Index tidak tersedia: {_IMAGE_SEARCH_ERR}")

# ── Admin Data Uploader (Harga + Populasi) ───────────────────────────
try:
    from admin_data_uploader import render_data_uploader_tab
    DATA_UPLOADER_ENABLED = True
except ImportError:
    DATA_UPLOADER_ENABLED = False
    def render_data_uploader_tab():
        st.warning("⚠️ `admin_data_uploader.py` tidak ditemukan.")

# ── Stok Opname ──────────────────────────────────────────────────────
try:
    import stok_opname as _so
    STOK_OPNAME_ENABLED = True
except ImportError:
    STOK_OPNAME_ENABLED = False
    _so = None

# ── Chat AI (DeepSeek) ───────────────────────────────────────────────
try:
    from chat_ai import render_chat_ai_tab
    CHAT_AI_ENABLED = True
except ImportError:
    CHAT_AI_ENABLED = False
    def render_chat_ai_tab(**_kw):
        st.warning("⚠️ `chat_ai.py` tidak ditemukan.")

# ── Konfigurasi Gudang (akun → cabang + stok terdekat) ───────────────
try:
    from gudang_config import (
        gudang_for_user, fallback_order, gudang_label,
    )
    GUDANG_CONFIG_ENABLED = True
except ImportError:
    GUDANG_CONFIG_ENABLED = False
    gudang_for_user = lambda u, r: None          # noqa: E731 — default: lihat semua
    fallback_order  = lambda own, allg: []        # noqa: E731
    gudang_label    = lambda g: g                 # noqa: E731


warnings.filterwarnings('ignore')

# ── Admin Menu Control (inline — tidak perlu file terpisah) ─────────────────

# Semua tab yang tersedia
ALL_MENU_TABS: dict = {
    "tab_search_pn":    "🔢 Search Part Number",
    "tab_search_name":  "📝 Search Part Name",
    "tab_search_image": "🖼️ Cari by Foto",
    "tab_compare":      "🔍 Bandingkan 2 Part",
    "tab_batch":        "📥 Batch Download",
    "tab_populasi":     "🚛 Populasi Unit",
    "tab_harga":        "💰 Harga",
    "tab_opname":       "📋 Stok Opname",
    "tab_chat_ai":      "🤖 Chat AI",
}

# Tab yang SELALU aktif (tidak bisa dinonaktifkan admin)
ALWAYS_ALLOWED: set = {"tab_search_pn"}

# File penyimpanan konfigurasi akses menu
MENU_CONFIG_FILE = Path("login/menu_permissions.json")

# ── Konfigurasi Akses Kolom ─────────────────────────────────────────────────

# Kolom yang dapat dikontrol aksesnya oleh admin
ALL_COLUMN_ACCESS: dict = {
    "col_stok":  "📦 Kolom Stok",
    "col_harga": "💲 Kolom Harga",
}

# File penyimpanan konfigurasi akses kolom
COLUMN_CONFIG_FILE = Path("login/column_permissions.json")

# ── Konfigurasi Akses Sub-Tab Harga ─────────────────────────────────────────

# Sub-tab yang ada di dalam Tab Harga
ALL_HARGA_SUBTABS: dict = {
    "subtab_list_harga":  "📋 List Harga",
    "subtab_cari_harga":  "🔍 Cari Harga",
    "subtab_batch_harga": "📥 Batch Cari Harga",
}

# File penyimpanan konfigurasi akses sub-tab harga
HARGA_SUBTAB_CONFIG_FILE = Path("login/harga_subtab_permissions.json")


class HargaSubTabManager:
    _CACHE_KEY = "_harga_subtab_permissions_cache"

    @classmethod
    def load_permissions(cls, force: bool = False) -> dict:
        if not force and cls._CACHE_KEY in st.session_state:
            return st.session_state[cls._CACHE_KEY]

        default_keys = list(ALL_HARGA_SUBTABS.keys())

        if SUPABASE_PERMS_ENABLED:
            data = SupabasePermissions.load(PERM_HARGA, default_keys, force=force)
        else:
            data = {}
            if HARGA_SUBTAB_CONFIG_FILE.exists():
                try:
                    with open(HARGA_SUBTAB_CONFIG_FILE, "r", encoding="utf-8") as f:
                        data = json.load(f)
                except Exception:
                    data = {}

        if "__default__" not in data:
            data["__default__"] = default_keys

        st.session_state[cls._CACHE_KEY] = data
        return data

    @classmethod
    def save_permissions(cls, permissions: dict) -> tuple:
        try:
            if SUPABASE_PERMS_ENABLED:
                ok = SupabasePermissions.save_all(PERM_HARGA, permissions)
                if not ok:
                    return False, "Gagal simpan ke Supabase"
            else:
                HARGA_SUBTAB_CONFIG_FILE.parent.mkdir(parents=True, exist_ok=True)
                with open(HARGA_SUBTAB_CONFIG_FILE, "w", encoding="utf-8") as f:
                    json.dump(permissions, f, indent=2, ensure_ascii=False)


            st.session_state[cls._CACHE_KEY] = permissions
            return True, None
        except Exception as e:
            return False, str(e)

    @classmethod
    def get_user_subtabs(cls, username: str) -> list:
        perms = cls.load_permissions()
        uname = username.strip().lower()
        return perms.get(uname, perms.get("__default__", list(ALL_HARGA_SUBTABS.keys())))

    @classmethod
    def set_user_subtabs(cls, username: str, subtab_keys: list) -> tuple:
        perms = cls.load_permissions()
        uname = username.strip().lower()
        final = [sk for sk in subtab_keys if sk in ALL_HARGA_SUBTABS]
        if SUPABASE_PERMS_ENABLED:
            ok = SupabasePermissions.save(PERM_HARGA, uname, final)
            if ok:
                perms[uname] = final
                st.session_state[cls._CACHE_KEY] = perms
            return (True, None) if ok else (False, "Gagal simpan ke Supabase")
        perms[uname] = final
        return cls.save_permissions(perms)

    @classmethod
    def set_default_subtabs(cls, subtab_keys: list) -> tuple:
        perms = cls.load_permissions()
        final = [sk for sk in subtab_keys if sk in ALL_HARGA_SUBTABS]
        if SUPABASE_PERMS_ENABLED:
            ok = SupabasePermissions.save(PERM_HARGA, "__default__", final)
            if ok:
                perms["__default__"] = final
                st.session_state[cls._CACHE_KEY] = perms
            return (True, None) if ok else (False, "Gagal simpan ke Supabase")
        perms["__default__"] = final
        return cls.save_permissions(perms)

    @classmethod
    def remove_user_config(cls, username: str) -> tuple:
        perms = cls.load_permissions()
        uname = username.strip().lower()
        if uname in perms:
            if SUPABASE_PERMS_ENABLED:
                ok = SupabasePermissions.remove(PERM_HARGA, uname)
                if ok:
                    perms.pop(uname, None)
                    st.session_state[cls._CACHE_KEY] = perms
                return (True, None) if ok else (False, "Gagal hapus dari Supabase")
            del perms[uname]
            return cls.save_permissions(perms)
        return True, None


def get_allowed_harga_subtabs(username: str, role: str) -> set:
    """Admin selalu mendapat semua sub-tab harga. User lain sesuai konfigurasi."""
    if role == "admin":
        return set(ALL_HARGA_SUBTABS.keys())
    return set(HargaSubTabManager.get_user_subtabs(username))


class ColumnAccessManager:
    """
    Mengelola izin tampil kolom Stok dan Harga per username.
    Disimpan di login/column_permissions.json.
    Format: {"username": ["col_stok", "col_harga"], "__default__": ["col_stok", "col_harga"]}
    """
    _CACHE_KEY = "_column_permissions_cache"

    @classmethod
    def load_permissions(cls, force: bool = False) -> dict:
        if not force and cls._CACHE_KEY in st.session_state:
            return st.session_state[cls._CACHE_KEY]

        default_keys = list(ALL_COLUMN_ACCESS.keys())

        if SUPABASE_PERMS_ENABLED:
            data = SupabasePermissions.load(PERM_COLUMN, default_keys, force=force)
        else:
            data = {}
            if COLUMN_CONFIG_FILE.exists():
                try:
                    with open(COLUMN_CONFIG_FILE, "r", encoding="utf-8") as f:
                        data = json.load(f)
                except Exception:
                    data = {}

        if "__default__" not in data:
            data["__default__"] = default_keys

        st.session_state[cls._CACHE_KEY] = data
        return data

    @classmethod
    def save_permissions(cls, permissions: dict) -> tuple:
        try:
            if SUPABASE_PERMS_ENABLED:
                ok = SupabasePermissions.save_all(PERM_COLUMN, permissions)
                if not ok:
                    return False, "Gagal simpan ke Supabase"
            else:
                COLUMN_CONFIG_FILE.parent.mkdir(parents=True, exist_ok=True)
                with open(COLUMN_CONFIG_FILE, "w", encoding="utf-8") as f:
                    json.dump(permissions, f, indent=2, ensure_ascii=False)


            st.session_state[cls._CACHE_KEY] = permissions
            return True, None
        except Exception as e:
            return False, str(e)

    @classmethod
    def get_user_columns(cls, username: str) -> list:
        perms = cls.load_permissions()
        uname = username.strip().lower()
        return perms.get(uname, perms.get("__default__", list(ALL_COLUMN_ACCESS.keys())))

    @classmethod
    def set_user_columns(cls, username: str, col_keys: list) -> tuple:
        perms = cls.load_permissions()
        uname = username.strip().lower()
        final = [ck for ck in col_keys if ck in ALL_COLUMN_ACCESS]
        if SUPABASE_PERMS_ENABLED:
            ok = SupabasePermissions.save(PERM_COLUMN, uname, final)
            if ok:
                perms[uname] = final
                st.session_state[cls._CACHE_KEY] = perms
            return (True, None) if ok else (False, "Gagal simpan ke Supabase")
        perms[uname] = final
        return cls.save_permissions(perms)

    @classmethod
    def set_default_columns(cls, col_keys: list) -> tuple:
        perms = cls.load_permissions()
        final = [ck for ck in col_keys if ck in ALL_COLUMN_ACCESS]
        if SUPABASE_PERMS_ENABLED:
            ok = SupabasePermissions.save(PERM_COLUMN, "__default__", final)
            if ok:
                perms["__default__"] = final
                st.session_state[cls._CACHE_KEY] = perms
            return (True, None) if ok else (False, "Gagal simpan ke Supabase")
        perms["__default__"] = final
        return cls.save_permissions(perms)

    @classmethod
    def remove_user_config(cls, username: str) -> tuple:
        perms = cls.load_permissions()
        uname = username.strip().lower()
        if uname in perms:
            if SUPABASE_PERMS_ENABLED:
                ok = SupabasePermissions.remove(PERM_COLUMN, uname)
                if ok:
                    perms.pop(uname, None)
                    st.session_state[cls._CACHE_KEY] = perms
                return (True, None) if ok else (False, "Gagal hapus dari Supabase")
            del perms[uname]
            return cls.save_permissions(perms)
        return True, None


def get_allowed_columns(username: str, role: str) -> set:
    """
    Kembalikan set col_key yang boleh ditampilkan untuk user ini.
    Admin selalu mendapatkan semua kolom.
    """
    if role == "admin":
        return set(ALL_COLUMN_ACCESS.keys())
    return set(ColumnAccessManager.get_user_columns(username))


class MenuAccessManager:
    """
    Mengelola izin akses menu per username.
    Disimpan di login/menu_permissions.json.
    Format: {"username": ["tab_key", ...], "__default__": ["tab_key", ...]}
    """
    _CACHE_KEY = "_menu_permissions_cache"

    @classmethod
    def load_permissions(cls, force: bool = False) -> dict:
        if not force and cls._CACHE_KEY in st.session_state:
            return st.session_state[cls._CACHE_KEY]

        default_keys = list(ALL_MENU_TABS.keys())

        if SUPABASE_PERMS_ENABLED:
            data = SupabasePermissions.load(PERM_MENU, default_keys, force=force)
        else:
            data = {}
            if MENU_CONFIG_FILE.exists():
                try:
                    with open(MENU_CONFIG_FILE, "r", encoding="utf-8") as f:
                        data = json.load(f)
                except Exception:
                    data = {}

        if "__default__" not in data:
            data["__default__"] = default_keys

        st.session_state[cls._CACHE_KEY] = data
        return data

    @classmethod
    def save_permissions(cls, permissions: dict) -> tuple:
        try:
            if SUPABASE_PERMS_ENABLED:
                ok = SupabasePermissions.save_all(PERM_MENU, permissions)
                if not ok:
                    st.toast("⚠️ Gagal simpan ke Supabase.", icon="⚠️")
                    return False, "Gagal simpan ke Supabase"
            else:
                MENU_CONFIG_FILE.parent.mkdir(parents=True, exist_ok=True)
                with open(MENU_CONFIG_FILE, "w", encoding="utf-8") as f:
                    json.dump(permissions, f, indent=2, ensure_ascii=False)


            st.session_state[cls._CACHE_KEY] = permissions
            return True, None
        except Exception as e:
            return False, str(e)

    @classmethod
    def get_user_tabs(cls, username: str) -> list:
        perms = cls.load_permissions()
        uname = username.strip().lower()
        allowed = perms.get(uname, perms.get("__default__", list(ALL_MENU_TABS.keys())))
        result = list(ALWAYS_ALLOWED)
        for tab in allowed:
            if tab in ALL_MENU_TABS and tab not in result:
                result.append(tab)
        return result

    @classmethod
    def set_user_tabs(cls, username: str, tab_keys: list) -> tuple:
        perms = cls.load_permissions()
        uname = username.strip().lower()
        final = list(ALWAYS_ALLOWED)
        for tk in tab_keys:
            if tk in ALL_MENU_TABS and tk not in final:
                final.append(tk)
        if SUPABASE_PERMS_ENABLED:
            ok = SupabasePermissions.save(PERM_MENU, uname, final)
            if ok:
                perms[uname] = final
                st.session_state[cls._CACHE_KEY] = perms
            return (True, None) if ok else (False, "Gagal simpan ke Supabase")
        perms[uname] = final
        return cls.save_permissions(perms)

    @classmethod
    def set_default_tabs(cls, tab_keys: list) -> tuple:
        perms = cls.load_permissions()
        final = list(ALWAYS_ALLOWED)
        for tk in tab_keys:
            if tk in ALL_MENU_TABS and tk not in final:
                final.append(tk)
        if SUPABASE_PERMS_ENABLED:
            ok = SupabasePermissions.save(PERM_MENU, "__default__", final)
            if ok:
                perms["__default__"] = final
                st.session_state[cls._CACHE_KEY] = perms
            return (True, None) if ok else (False, "Gagal simpan ke Supabase")
        perms["__default__"] = final
        return cls.save_permissions(perms)

    @classmethod
    def remove_user_config(cls, username: str) -> tuple:
        perms = cls.load_permissions()
        uname = username.strip().lower()
        if uname in perms:
            if SUPABASE_PERMS_ENABLED:
                ok = SupabasePermissions.remove(PERM_MENU, uname)
                if ok:
                    perms.pop(uname, None)
                    st.session_state[cls._CACHE_KEY] = perms
                return (True, None) if ok else (False, "Gagal hapus dari Supabase")
            del perms[uname]
            return cls.save_permissions(perms)
        return True, None

    @classmethod
    def get_all_configured_users(cls) -> list:
        perms = cls.load_permissions()
        return [k for k in perms.keys() if k != "__default__"]


def get_allowed_tabs(username: str, role: str) -> list:
    """Admin selalu dapat semua tab. User lain sesuai konfigurasi."""
    if role == "admin":
        return list(ALL_MENU_TABS.keys()) + [
            "tab_menu_control", "tab_data_upload", "tab_foto_part", "tab_image_index",
            "tab_user_mgmt", "tab_user_monitoring",
        ]
    return MenuAccessManager.get_user_tabs(username)


def _mac_quick_actions(section_key: str, all_keys: list, locked: set,
                       current: list, default_keys: list) -> list:
    """Tombol Pilih Semua / Kosongkan / Pakai Default untuk 1 section.
    Return list state baru (dari session_state per checkbox)."""
    state_key = f"_mac_state_{section_key}"
    if state_key not in st.session_state:
        st.session_state[state_key] = list(current)

    btn_cols = st.columns(3)
    with btn_cols[0]:
        if st.button("✅ Pilih Semua", key=f"qa_all_{section_key}", use_container_width=True):
            st.session_state[state_key] = list(all_keys)
            st.rerun()
    with btn_cols[1]:
        if st.button("❌ Kosongkan", key=f"qa_none_{section_key}", use_container_width=True):
            # Tetap pertahankan yang locked (wajib)
            st.session_state[state_key] = list(locked)
            st.rerun()
    with btn_cols[2]:
        if st.button("↩️ Pakai Default", key=f"qa_def_{section_key}", use_container_width=True):
            st.session_state[state_key] = list(default_keys)
            st.rerun()

    return st.session_state[state_key]


def _mac_render_section(title: str, items: list, current: list, default_keys: list,
                        locked: set, section_key: str) -> list:
    """Render 1 section (Menu/Kolom/HargaSubtab) dengan quick actions + checkbox.
    Return list final selection."""
    st.markdown(f"#### {title}")

    n_total   = len(items)
    n_picked  = len([k for k, _ in items if k in current])
    st.caption(f"Aktif: **{n_picked}/{n_total}**")

    state = _mac_quick_actions(section_key, [k for k, _ in items], locked, current, default_keys)

    new_sel: list = []
    cols = st.columns(2)
    for i, (key, label) in enumerate(items):
        with cols[i % 2]:
            forced = key in locked
            val = st.checkbox(
                label=(label + " *(wajib)*") if forced else label,
                value=True if forced else (key in state),
                key=f"{section_key}_cb_{key}",
                disabled=forced,
                help="Selalu aktif, tidak bisa dimatikan." if forced else "",
            )
            if forced or val:
                new_sel.append(key)

    # Update state agar tombol quick action konsisten dengan klik manual
    st.session_state[f"_mac_state_{section_key}"] = list(new_sel)
    return new_sel


def render_admin_menu_control_tab():
    """UI tab 🛡️ Menu Control — hanya tampil untuk admin."""
    st.markdown("### 🛡️ Kontrol Akses per User")
    st.caption(
        "Atur akses **menu**, **kolom**, dan **sub-tab harga** per user, "
        "atau atur **default** untuk semua user yang belum punya konfigurasi khusus. "
        "Tab **Search Part Number** selalu aktif (tidak bisa dimatikan)."
    )

    # ── Tombol force-refresh dari Supabase ───────────────────────────
    rc1, rc2 = st.columns([1, 5])
    with rc1:
        if st.button("🔄 Refresh", use_container_width=True,
                     help="Reload permissions terbaru dari Supabase"):
            for k in ("_menu_permissions_cache",
                      "_column_permissions_cache",
                      "_harga_subtab_permissions_cache"):
                st.session_state.pop(k, None)
            for k in list(st.session_state.keys()):
                if k.startswith("_mac_state_"):
                    st.session_state.pop(k, None)
            try:
                if SUPABASE_PERMS_ENABLED:
                    SupabasePermissions.invalidate(PERM_MENU)
                    SupabasePermissions.invalidate(PERM_COLUMN)
                    SupabasePermissions.invalidate(PERM_HARGA)
            except Exception:
                pass
            st.rerun()

    df_users: pd.DataFrame = st.session_state.get("login_users_df", pd.DataFrame())
    if df_users.empty:
        st.warning("⚠️ Data user belum dimuat. Klik **Reload Users** di sidebar terlebih dahulu.")
        return

    non_admin_users = df_users[df_users["role"] != "admin"]["username"].tolist()
    if not non_admin_users:
        st.info("Tidak ada user non-admin yang terdaftar.")
        return

    perms      = MenuAccessManager.load_permissions()
    col_perms  = ColumnAccessManager.load_permissions()
    hs_perms   = HargaSubTabManager.load_permissions()
    tab_items  = list(ALL_MENU_TABS.items())
    col_items  = list(ALL_COLUMN_ACCESS.items())
    hs_items   = list(ALL_HARGA_SUBTABS.items())

    default_tab_keys = list(ALL_MENU_TABS.keys())
    default_col_keys = list(ALL_COLUMN_ACCESS.keys())
    default_hs_keys  = list(ALL_HARGA_SUBTABS.keys())

    # ── Mode: edit per-user atau edit default ─────────────────────────
    mode = st.radio(
        "Mode edit:",
        ["👤 Per User", "🌐 Default (untuk semua user baru)"],
        horizontal=True,
        key="mac_mode",
    )

    is_default_mode = mode.startswith("🌐")

    if is_default_mode:
        target = "__default__"
        target_label = "DEFAULT (semua user baru)"
    else:
        selected_user = st.selectbox(
            "Pilih Username:", options=non_admin_users, key="mac_sel_user",
        )
        if not selected_user:
            return
        target = selected_user.strip().lower()
        target_label = target

        # Status badge: punya config khusus atau pakai default?
        has_specific = target in perms or target in col_perms or target in hs_perms
        if has_specific:
            st.markdown(
                f"<div style='background:#DBEAFE;border-left:3px solid #2563EB;"
                f"padding:6px 10px;border-radius:4px;margin:4px 0;font-size:.82rem;'>"
                f"📌 User <b>{target}</b> memiliki konfigurasi khusus.</div>",
                unsafe_allow_html=True,
            )
        else:
            st.markdown(
                f"<div style='background:#FEF3C7;border-left:3px solid #D97706;"
                f"padding:6px 10px;border-radius:4px;margin:4px 0;font-size:.82rem;'>"
                f"ℹ️ User <b>{target}</b> belum punya konfigurasi khusus — saat ini "
                f"mengikuti pengaturan <b>Default</b>.</div>",
                unsafe_allow_html=True,
            )

    # ── Reset state ketika ganti target/mode ──────────────────────────
    last_target_key = "_mac_last_target"
    if st.session_state.get(last_target_key) != target:
        # Hapus semua state per-checkbox dari section sebelumnya supaya nilai
        # baru terpasang dari "current"
        for k in list(st.session_state.keys()):
            if k.startswith("_mac_state_"):
                st.session_state.pop(k, None)
        st.session_state[last_target_key] = target

    # ── Akses Menu ────────────────────────────────────────────────────
    st.markdown("---")
    cur_tabs = perms.get(target, perms.get("__default__", default_tab_keys))
    new_selection = _mac_render_section(
        "📋 Akses Menu", tab_items, cur_tabs, default_tab_keys,
        locked=ALWAYS_ALLOWED, section_key=f"menu_{target}",
    )

    # ── Akses Kolom ───────────────────────────────────────────────────
    st.markdown("---")
    cur_cols = col_perms.get(target, col_perms.get("__default__", default_col_keys))
    new_col_sel = _mac_render_section(
        "🔒 Akses Kolom", col_items, cur_cols, default_col_keys,
        locked=set(), section_key=f"col_{target}",
    )

    # ── Akses Sub-Tab Harga ───────────────────────────────────────────
    st.markdown("---")
    cur_hs = hs_perms.get(target, hs_perms.get("__default__", default_hs_keys))
    new_hs_sel = _mac_render_section(
        "🔖 Akses Sub-Tab Harga", hs_items, cur_hs, default_hs_keys,
        locked=set(), section_key=f"hs_{target}",
    )

    # ── Tombol Simpan / Reset ─────────────────────────────────────────
    st.markdown("---")
    bc1, bc2 = st.columns([2, 1])
    with bc1:
        save_label = (
            "💾 Simpan Akses Default"
            if is_default_mode
            else f"💾 Simpan Akses untuk {target_label}"
        )
        if st.button(save_label, key=f"mac_save_{target}",
                     type="primary", use_container_width=True):
            results = []
            if is_default_mode:
                results.append(MenuAccessManager.set_default_tabs(new_selection))
                results.append(ColumnAccessManager.set_default_columns(new_col_sel))
                results.append(HargaSubTabManager.set_default_subtabs(new_hs_sel))
            else:
                results.append(MenuAccessManager.set_user_tabs(target, new_selection))
                results.append(ColumnAccessManager.set_user_columns(target, new_col_sel))
                results.append(HargaSubTabManager.set_user_subtabs(target, new_hs_sel))

            errs = [err for ok, err in results if not ok and err]
            all_ok = all(ok for ok, _ in results)
            if all_ok:
                # Force re-fetch dari DB pada render berikutnya supaya yang
                # ditampilkan = state DB sebenarnya
                for k in ("_menu_permissions_cache",
                          "_column_permissions_cache",
                          "_harga_subtab_permissions_cache"):
                    st.session_state.pop(k, None)
                if SUPABASE_PERMS_ENABLED:
                    try:
                        SupabasePermissions.invalidate(PERM_MENU)
                        SupabasePermissions.invalidate(PERM_COLUMN)
                        SupabasePermissions.invalidate(PERM_HARGA)
                    except Exception:
                        pass
                try:
                    _admin = LoginManager.get_current_user() or {}
                    log_activity(_admin.get("username", ""), "permission_change",
                                 target="__default__" if is_default_mode else target,
                                 details={
                                     "menu_tabs":     new_selection,
                                     "columns":       new_col_sel,
                                     "harga_subtabs": new_hs_sel,
                                 })
                except Exception:
                    pass
                st.success(f"✅ Akses untuk **{target_label}** disimpan.")
                st.rerun()
            else:
                st.error(f"❌ Gagal menyimpan: {', '.join(errs) or 'unknown error'}")

    with bc2:
        if not is_default_mode:
            if st.button("🗑️ Hapus Konfigurasi User", key=f"mac_remove_{target}",
                         use_container_width=True,
                         help="User akan kembali pakai pengaturan Default"):
                results = [
                    MenuAccessManager.remove_user_config(target),
                    ColumnAccessManager.remove_user_config(target),
                    HargaSubTabManager.remove_user_config(target),
                ]
                if all(ok for ok, _ in results):
                    for k in ("_menu_permissions_cache",
                              "_column_permissions_cache",
                              "_harga_subtab_permissions_cache"):
                        st.session_state.pop(k, None)
                    st.success(f"✅ Konfigurasi {target} dihapus, kembali ke Default.")
                    st.rerun()
                else:
                    errs = [e for ok, e in results if not ok and e]
                    st.error(f"❌ Gagal hapus: {', '.join(errs) or 'unknown'}")

    # ── Overview semua user ───────────────────────────────────────────
    with st.expander("📊 Overview Akses Semua User", expanded=False):
        rows = []
        all_tabs_keys = list(ALL_MENU_TABS.keys())
        for u in non_admin_users:
            uname = u.strip().lower()
            user_tabs = perms.get(uname, perms.get("__default__", all_tabs_keys))
            row = {
                "User":    u,
                "Konfig":  "Khusus" if uname in perms else "Default",
                "Jml Tab": f"{len(user_tabs)}/{len(all_tabs_keys)}",
            }
            for tk, tlabel in ALL_MENU_TABS.items():
                # Ambil emoji/teks pendek dari label
                short = tlabel.split(" ", 1)[0]
                row[short] = "✅" if tk in user_tabs else "—"
            rows.append(row)
        if rows:
            st.dataframe(pd.DataFrame(rows), hide_index=True, use_container_width=True)








# ── SIMS Image Fetcher ─────────────────────────────────────────────
try:
    from sims_fetcher import get_sims_images as _sims_fetch
    SIMS_ENABLED = True
except ImportError:
    SIMS_ENABLED = False
    def _sims_fetch(pn, force_refresh=False):
        return [], "sims_fetcher.py tidak ditemukan"

st.set_page_config(
    page_title="Part Number Finder",
    page_icon="🔍",
    layout="wide",
    initial_sidebar_state="collapsed",
    menu_items={'Get Help': None, 'Report a bug': None, 'About': None}
)

SIDEBAR_TOGGLE_JS = """
<script>
(function() {
    if (window.__sidebarToggleInjected) return;
    window.__sidebarToggleInjected = true;

    function getSidebar() {
        return document.querySelector('[data-testid="stSidebar"]');
    }

    function isCollapsed() {
        var sb = getSidebar();
        if (!sb) return false;
        return sb.getAttribute('aria-expanded') === 'false'
            || sb.classList.contains('st-emotion-cache-hidden')
            || getComputedStyle(sb).transform.includes('translateX(-')
            || getComputedStyle(sb).marginLeft.startsWith('-');
    }

    function clickOriginalToggle() {
        // Coba klik tombol toggle bawaan Streamlit
        var btn = document.querySelector('[data-testid="collapsedControl"] button')
                || document.querySelector('button[kind="header"]')
                || document.querySelector('[data-testid="stSidebarCollapsedControl"] button');
        if (btn) { btn.click(); return true; }
        // Fallback: toggle class langsung
        var sb = getSidebar();
        if (!sb) return false;
        var expanded = sb.getAttribute('aria-expanded');
        if (expanded === 'false') {
            sb.setAttribute('aria-expanded', 'true');
        } else {
            sb.setAttribute('aria-expanded', 'false');
        }
        return true;
    }

    function updateIcon() {
        var btn = document.getElementById('custom-sidebar-toggle');
        if (!btn) return;
        btn.innerHTML = isCollapsed() ? '&#9776;' : '&#10005;';
        btn.title = isCollapsed() ? 'Buka Sidebar' : 'Tutup Sidebar';
    }

    function injectButton() {
        if (document.getElementById('custom-sidebar-toggle')) return;
        var btn = document.createElement('button');
        btn.id = 'custom-sidebar-toggle';
        btn.innerHTML = '&#9776;';
        btn.title = 'Buka/Tutup Sidebar';
        btn.addEventListener('click', function() {
            clickOriginalToggle();
            setTimeout(updateIcon, 300);
        });
        document.body.appendChild(btn);
        updateIcon();
    }

    // Inject setelah DOM ready
    if (document.readyState === 'loading') {
        document.addEventListener('DOMContentLoaded', injectButton);
    } else {
        setTimeout(injectButton, 500);
    }

    // Re-inject kalau Streamlit rerun (MutationObserver)
    var observer = new MutationObserver(function() {
        injectButton();
        updateIcon();
    });
    observer.observe(document.body, { childList: true, subtree: false });
})();
</script>
"""

KEEP_ALIVE_JS = """
<script>
(function() {
    if (window.__keepAliveActive) return;
    window.__keepAliveActive = true;
    const INTERVAL_MS = 4 * 60 * 1000;  // setiap 4 menit
    function ping() {
        var xhr = new XMLHttpRequest();
        xhr.open('GET', window.location.href + '?_ka=' + Date.now(), true);
        xhr.send();
    }
    window.__keepAliveTimer = setInterval(ping, INTERVAL_MS);
    setTimeout(ping, 30 * 1000);  // ping pertama setelah 30 detik
})();
</script>
"""

def inject_keep_alive():
    # Gunakan components.html agar <script> benar-benar dieksekusi browser
    _stc.html(KEEP_ALIVE_JS + SIDEBAR_TOGGLE_JS, height=0, scrolling=False)

# ── Auto-scroll ke section "Gambar Part" setelah Prev/Next/thumbnail ──────
# Streamlit's st.rerun() reset scroll ke atas. Setelah klik Next ▶ user
# selalu harus scroll lagi ke bawah untuk lihat gambar berikutnya. Helper
# di bawah ini menandai bahwa rerun berikutnya harus scroll balik ke
# anchor `gambar-part-anchor`.
_IMAGE_SCROLL_FLAG = "_pn_scroll_to_image"

def _request_image_scroll():
    st.session_state[_IMAGE_SCROLL_FLAG] = True

def _emit_image_scroll_anchor():
    st.markdown(
        '<div id="gambar-part-anchor" '
        'style="position:relative;scroll-margin-top:80px;"></div>',
        unsafe_allow_html=True,
    )
    if st.session_state.pop(_IMAGE_SCROLL_FLAG, False):
        _stc.html(
            """
            <script>
            (function() {
              try {
                var w = window.parent || window;
                var d = w.document;
                function go(n) {
                  var el = d.getElementById('gambar-part-anchor');
                  if (el) {
                    el.scrollIntoView({behavior: 'smooth', block: 'start'});
                    return;
                  }
                  if (n < 20) setTimeout(function(){ go(n + 1); }, 80);
                }
                go(0);
              } catch (e) {}
            })();
            </script>
            """,
            height=0,
        )

TAB_PERSIST_JS = """
<script>
(function() {
    const KEY = 'pnf_active_tab';
    function attachListeners() {
        document.querySelectorAll('[data-baseweb="tab"]').forEach(function(tab, idx) {
            if (!tab._pnf_listener) {
                tab._pnf_listener = true;
                tab.addEventListener('click', function() {
                    sessionStorage.setItem(KEY, idx);
                });
            }
        });
    }
    function restoreTab() {
        var saved = sessionStorage.getItem(KEY);
        if (saved === null) return;
        var idx = parseInt(saved);
        var tabs = document.querySelectorAll('[data-baseweb="tab"]');
        if (tabs.length > idx && tabs[idx].getAttribute('aria-selected') !== 'true') {
            tabs[idx].click();
        }
    }
    var _lastTabCount = 0;
    var observer = new MutationObserver(function() {
        var tabs = document.querySelectorAll('[data-baseweb="tab"]');
        if (tabs.length !== _lastTabCount) {
            _lastTabCount = tabs.length;
            attachListeners();
            setTimeout(restoreTab, 50);
        }
    });
    observer.observe(document.body, { childList: true, subtree: true });
    setTimeout(function() { attachListeners(); restoreTab(); }, 400);
})();
</script>
"""

st.markdown("""
<link rel="preconnect" href="https://fonts.googleapis.com">
<link rel="preconnect" href="https://fonts.gstatic.com" crossorigin>
<link href="https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700;800&family=JetBrains+Mono:wght@400;500;600;700&display=swap" rel="stylesheet">
<style>
    /* ════════════════════════════════════════════════════════════════════
       MasPart Design System — putih bersih + green #028912 sebagai brand.
       Tokens disinkronkan dengan screens-shared.jsx pada handoff design.
       ════════════════════════════════════════════════════════════════════ */
    :root {
        --mp-green:        #028912;
        --mp-green-dark:   #026B0E;
        --mp-green-soft:   #E6F5E8;
        --mp-green-softer: #F1FAF2;
        --mp-green-line:   #C7E8CC;
        --mp-ink:          #0F1A12;
        --mp-ink-70:       #3D4A40;
        --mp-ink-50:       #6B7A6F;
        --mp-ink-30:       #A6B0A8;
        --mp-line:         #E5E7E3;
        --mp-line-soft:    #EFF1ED;
        --mp-bg:           #FFFFFF;
        --mp-bg-muted:     #F7F8F5;
        --mp-bg-canvas:    #FAFAF7;
        --mp-amber:        #B5750B;
        --mp-amber-soft:   #FEF6E3;
        --mp-amber-line:   #F2D88B;
        --mp-red:          #B42318;
        --mp-red-soft:     #FEEAE7;
        --mp-red-line:     #F4C8C2;
        --mp-blue:         #1856B6;
        --mp-blue-soft:    #E5EEFB;
        --mp-radius:       10px;
        --mp-radius-sm:    6px;
        --mp-shadow-card:  0 1px 2px rgba(15,26,18,.04), 0 1px 3px rgba(15,26,18,.06);
        --mp-shadow-pop:   0 6px 24px rgba(15,26,18,.10), 0 2px 6px rgba(15,26,18,.06);
        --mp-font-sans:    'Inter', -apple-system, BlinkMacSystemFont, "Segoe UI", Helvetica, Arial, sans-serif;
        --mp-font-mono:    'JetBrains Mono', ui-monospace, SFMono-Regular, Menlo, monospace;
    }

    html, body, [class*="css"] { font-family: var(--mp-font-sans); }
    body { background: var(--mp-bg); color: var(--mp-ink); }

    /* Sembunyikan chrome bawaan Streamlit */
    #MainMenu { visibility: hidden; }
    footer { visibility: hidden; }
    .stDeployButton { display: none !important; }
    header[data-testid="stHeader"] { display: none !important; height: 0 !important; }
    div[data-testid="stToolbar"] { display: none !important; }
    [data-testid="stDecoration"] { display: none !important; }
    [data-testid="stStatusWidget"] { display: none !important; }
    /* Hapus padding-top default Streamlit yang menyisakan area kosong di atas. */
    .stApp > header { display: none !important; }
    .stApp,
    [data-testid="stApp"],
    [data-testid="stAppViewContainer"] { padding-top: 0 !important; }
    .main, section.main, [data-testid="stMain"] { padding-top: 0 !important; }
    .main .block-container,
    section.main .block-container,
    [data-testid="stMain"] .block-container,
    [data-testid="stMainBlockContainer"],
    [data-testid="stAppViewContainer"] .block-container {
        padding-top: .75rem !important;
        margin-top: 0 !important;
    }
    .login-page [data-testid="stSidebar"] > div { display: none !important; }
    iframe[height="0"] { display: none !important; }
    /* Hide element-container yang isinya cuma iframe 0-height (mis. inject_keep_alive) */
    .element-container:has(iframe[height="0"]),
    [data-testid="stIFrame"]:has(iframe[height="0"]),
    [data-testid="element-container"]:has(iframe[height="0"]) {
        display: none !important;
        height: 0 !important;
        margin: 0 !important;
        padding: 0 !important;
    }
    div[data-testid="stCustomComponentV1"][height="0"] { display: none !important; }

    /* Tombol toggle sidebar custom — recolor ke brand green */
    #custom-sidebar-toggle {
        position: fixed !important;
        top: 10px !important;
        left: 10px !important;
        z-index: 999999 !important;
        width: 36px !important;
        height: 36px !important;
        background: var(--mp-green) !important;
        color: white !important;
        border: none !important;
        border-radius: var(--mp-radius-sm) !important;
        cursor: pointer !important;
        font-size: 18px !important;
        display: flex !important;
        align-items: center !important;
        justify-content: center !important;
        box-shadow: 0 2px 8px rgba(2,137,18,.25) !important;
        transition: background .2s !important;
    }
    #custom-sidebar-toggle:hover { background: var(--mp-green-dark) !important; }

    /* ── Headings ─────────────────────────────────────────────────── */
    h1, h2, h3, h4, h5 { letter-spacing: -.01em; color: var(--mp-ink); }
    h1 .accent, h2 .accent, h3 .accent { color: var(--mp-green); }

    /* Legacy classes — di-recolor ke palette baru */
    .main-header {
        font-size: 1.95rem; font-weight: 700;
        color: var(--mp-ink); letter-spacing: -.02em;
        text-align: left; margin: .25rem 0 1rem; padding-top: .4rem;
    }
    .main-header .accent { color: var(--mp-green); }
    .sub-header {
        font-size: 1.05rem; font-weight: 700;
        color: var(--mp-ink); letter-spacing: -.01em;
        margin: 1.25rem 0 .75rem;
    }
    .user-badge {
        display: inline-flex; align-items: center; gap: .45rem;
        background: var(--mp-green-soft); border: 1px solid var(--mp-green-line);
        border-radius: 999px; padding: .3rem .85rem;
        font-size: .82rem; color: var(--mp-green-dark); font-weight: 600;
    }
    .role-admin { color: var(--mp-green-dark); font-weight: 700; letter-spacing: .04em; }
    .role-user  { color: var(--mp-ink-50); font-weight: 600; letter-spacing: .04em; }
    .batch-info-box {
        background: var(--mp-green-softer);
        border-left: 4px solid var(--mp-green);
        padding: .8rem 1rem;
        border-radius: 0 var(--mp-radius-sm) var(--mp-radius-sm) 0;
        margin-bottom: 1rem; color: var(--mp-green-dark);
    }

    /* ── Primary button = green ─────────────────────────────────── */
    .stButton > button[kind="primary"],
    .stDownloadButton > button[kind="primary"],
    .stFormSubmitButton > button[kind="primary"] {
        background: var(--mp-green) !important;
        border: 1px solid var(--mp-green) !important;
        color: white !important;
        font-weight: 600 !important;
        border-radius: var(--mp-radius-sm) !important;
        box-shadow: 0 1px 2px rgba(2,137,18,.15);
        transition: background .15s, border-color .15s, box-shadow .15s;
    }
    .stButton > button[kind="primary"]:hover,
    .stDownloadButton > button[kind="primary"]:hover,
    .stFormSubmitButton > button[kind="primary"]:hover {
        background: var(--mp-green-dark) !important;
        border-color: var(--mp-green-dark) !important;
    }
    .stButton > button[kind="primary"]:focus,
    .stFormSubmitButton > button[kind="primary"]:focus {
        box-shadow: 0 0 0 3px var(--mp-green-soft) !important;
    }

    /* Secondary button = ghost (white + green border on hover) */
    .stButton > button[kind="secondary"],
    .stDownloadButton > button[kind="secondary"],
    .stFormSubmitButton > button[kind="secondary"] {
        background: var(--mp-bg) !important;
        border: 1px solid var(--mp-line) !important;
        color: var(--mp-ink) !important;
        font-weight: 600 !important;
        border-radius: var(--mp-radius-sm) !important;
        transition: background .15s, border-color .15s, color .15s;
    }
    .stButton > button[kind="secondary"]:hover,
    .stDownloadButton > button[kind="secondary"]:hover,
    .stFormSubmitButton > button[kind="secondary"]:hover {
        background: var(--mp-bg-muted) !important;
        border-color: var(--mp-green-line) !important;
        color: var(--mp-green-dark) !important;
    }

    /* ── Tabs ─────────────────────────────────────────────────────── */
    .stTabs [data-baseweb="tab-list"] {
        gap: 2px;
        border-bottom: 1px solid var(--mp-line);
        background: transparent;
    }
    .stTabs [data-baseweb="tab"] {
        padding: 10px 14px !important;
        font-weight: 500;
        font-size: 13px;
        color: var(--mp-ink-50) !important;
        background: transparent !important;
        border: none !important;
        border-bottom: 2px solid transparent !important;
        margin-bottom: -1px;
    }
    .stTabs [data-baseweb="tab"]:hover {
        color: var(--mp-ink) !important;
        background: transparent !important;
    }
    .stTabs [data-baseweb="tab"][aria-selected="true"] {
        color: var(--mp-green) !important;
        font-weight: 600 !important;
        border-bottom-color: var(--mp-green) !important;
    }
    .stTabs [data-baseweb="tab-highlight"] { background: transparent !important; }
    .stTabs [data-baseweb="tab-border"] { display: none; }

    /* ── Sidebar ──────────────────────────────────────────────────── */
    [data-testid="stSidebar"] {
        background: var(--mp-bg-muted) !important;
        border-right: 1px solid var(--mp-line);
    }
    [data-testid="stSidebar"] [data-testid="stSidebarContent"] { padding-top: 8px; }
    [data-testid="stSidebar"] h1,
    [data-testid="stSidebar"] h2,
    [data-testid="stSidebar"] h3 {
        font-size: .72rem !important; font-weight: 700 !important;
        color: var(--mp-ink-50) !important;
        letter-spacing: .08em; text-transform: uppercase;
        margin: .9rem .2rem .25rem !important;
    }
    [data-testid="stSidebar"] .stButton > button {
        background: transparent !important;
        border: 1px solid transparent !important;
        color: var(--mp-ink) !important;
        text-align: left !important;
        justify-content: flex-start !important;
        font-weight: 500 !important;
        padding: .45rem .65rem !important;
    }
    [data-testid="stSidebar"] .stButton > button:hover {
        background: var(--mp-bg) !important;
        border-color: var(--mp-line) !important;
    }
    [data-testid="stSidebar"] .stButton > button[kind="primary"] {
        background: var(--mp-green) !important;
        color: white !important;
        border-color: var(--mp-green) !important;
    }
    [data-testid="stSidebar"] hr { border-color: var(--mp-line) !important; margin: .6rem 0 !important; }
    [data-testid="stSidebar"] .stMetric { background: var(--mp-bg); border-radius: var(--mp-radius-sm); padding: .5rem .65rem; border: 1px solid var(--mp-line); }
    [data-testid="stSidebar"] [data-testid="stExpander"] { background: var(--mp-bg); border: 1px solid var(--mp-line); border-radius: var(--mp-radius-sm); margin: .35rem 0; }

    /* ── Inputs ───────────────────────────────────────────────────── */
    .stTextInput input,
    .stTextArea textarea,
    .stNumberInput input,
    .stDateInput input {
        border-radius: var(--mp-radius-sm) !important;
        border-color: var(--mp-line) !important;
        font-family: var(--mp-font-sans);
    }
    .stTextInput input:focus,
    .stTextArea textarea:focus,
    .stNumberInput input:focus,
    .stDateInput input:focus,
    .stSelectbox div[data-baseweb="select"]:focus-within,
    .stMultiSelect div[data-baseweb="select"]:focus-within {
        border-color: var(--mp-green) !important;
        box-shadow: 0 0 0 3px var(--mp-green-soft) !important;
    }
    .stSelectbox div[data-baseweb="select"] > div,
    .stMultiSelect div[data-baseweb="select"] > div {
        border-radius: var(--mp-radius-sm) !important;
        border-color: var(--mp-line) !important;
    }

    /* Radio horizontal: pill style */
    .stRadio [role="radiogroup"] {
        background: var(--mp-bg-muted);
        border: 1px solid var(--mp-line);
        border-radius: var(--mp-radius-sm);
        padding: 3px;
        display: inline-flex; gap: 2px;
    }
    .stRadio [role="radiogroup"] > label {
        margin: 0 !important;
        padding: 6px 12px !important;
        border-radius: 4px;
        cursor: pointer;
        font-size: 12.5px;
        font-weight: 500;
        color: var(--mp-ink-50);
    }
    .stRadio [role="radiogroup"] > label:has(input:checked) {
        background: var(--mp-bg);
        color: var(--mp-green) !important;
        font-weight: 600;
        box-shadow: var(--mp-shadow-card);
    }
    .stRadio [role="radiogroup"] input[type="radio"] { display: none; }
    /* Checkbox & radio accent */
    input[type="checkbox"], input[type="radio"] { accent-color: var(--mp-green); }

    /* ── File uploader ────────────────────────────────────────────── */
    [data-testid="stFileUploader"] section {
        background: var(--mp-green-softer) !important;
        border: 2px dashed var(--mp-green-line) !important;
        border-radius: var(--mp-radius) !important;
    }
    [data-testid="stFileUploader"] section:hover {
        background: var(--mp-green-soft) !important;
        border-color: var(--mp-green) !important;
    }
    [data-testid="stFileUploader"] button {
        background: var(--mp-green) !important;
        border: 1px solid var(--mp-green) !important;
        color: white !important;
        border-radius: var(--mp-radius-sm) !important;
    }

    /* ── Dataframe / table ────────────────────────────────────────── */
    [data-testid="stDataFrame"] thead tr th,
    [data-testid="stTable"] thead tr th {
        background: var(--mp-bg-muted) !important;
        color: var(--mp-ink-70) !important;
        font-weight: 600 !important;
        font-size: 11.5px !important;
        letter-spacing: .04em;
        text-transform: uppercase;
        border-bottom: 1px solid var(--mp-line) !important;
    }
    [data-testid="stDataFrame"] tbody tr:hover td { background: var(--mp-green-softer) !important; }

    /* ── Metric ───────────────────────────────────────────────────── */
    [data-testid="stMetric"] {
        background: var(--mp-bg);
        border: 1px solid var(--mp-line);
        border-radius: var(--mp-radius);
        padding: 14px 16px;
    }
    [data-testid="stMetricLabel"] {
        font-size: 11px !important; font-weight: 600 !important;
        color: var(--mp-ink-50) !important;
        letter-spacing: .04em; text-transform: uppercase;
    }
    [data-testid="stMetricValue"] {
        font-size: 24px !important; font-weight: 700 !important;
        color: var(--mp-ink); letter-spacing: -.02em;
    }
    [data-testid="stMetricDelta"] { font-size: 11px !important; color: var(--mp-ink-50) !important; }

    /* ── Expander ─────────────────────────────────────────────────── */
    [data-testid="stExpander"] {
        border: 1px solid var(--mp-line) !important;
        border-radius: var(--mp-radius) !important;
        background: var(--mp-bg);
    }
    [data-testid="stExpander"] details summary {
        font-weight: 600 !important;
        color: var(--mp-ink) !important;
    }
    [data-testid="stExpander"] details summary:hover { color: var(--mp-green-dark) !important; }

    /* ── Progress bar ─────────────────────────────────────────────── */
    .stProgress > div > div > div > div { background: var(--mp-green) !important; }
    .stProgress > div > div > div { background: var(--mp-line-soft) !important; border-radius: 999px; height: 6px; }

    /* ── Alerts ───────────────────────────────────────────────────── */
    [data-testid="stAlert"] { border-radius: var(--mp-radius-sm) !important; border: 1px solid var(--mp-line); }

    /* ── Custom utility classes untuk redesign per-screen ─────────── */
    .mp-card {
        background: var(--mp-bg);
        border: 1px solid var(--mp-line);
        border-radius: var(--mp-radius);
        padding: 18px;
        box-shadow: var(--mp-shadow-card);
    }
    .mp-card.tight { padding: 14px; }
    .mp-card.flat  { box-shadow: none; }
    .mp-chip {
        display: inline-flex; align-items: center; gap: 5px;
        padding: 3px 9px; font-size: 11.5px; font-weight: 600;
        border-radius: 999px;
        background: var(--mp-green-soft); color: var(--mp-green-dark);
        border: 1px solid var(--mp-green-line);
    }
    .mp-chip.gray  { background: var(--mp-bg-muted); color: var(--mp-ink-50); border-color: var(--mp-line); }
    .mp-chip.amber { background: var(--mp-amber-soft); color: var(--mp-amber); border-color: var(--mp-amber-line); }
    .mp-chip.red   { background: var(--mp-red-soft); color: var(--mp-red); border-color: var(--mp-red-line); }
    .mp-chip.blue  { background: var(--mp-blue-soft); color: var(--mp-blue); border-color: #BAD0F0; }
    .mp-mono { font-family: var(--mp-font-mono); letter-spacing: .01em; }

    .mp-banner {
        display: flex; gap: 10px;
        padding: 10px 14px;
        border-radius: var(--mp-radius-sm);
        font-size: 12.5px;
        align-items: flex-start;
    }
    .mp-banner.info { background: var(--mp-green-softer); border: 1px solid var(--mp-green-line); color: var(--mp-green-dark); }
    .mp-banner.warn { background: var(--mp-amber-soft); border: 1px solid var(--mp-amber-line); color: var(--mp-amber); }
    .mp-banner.err  { background: var(--mp-red-soft); border: 1px solid var(--mp-red-line); color: var(--mp-red); }

    /* Sidebar logo + user badge cards (untuk Phase B) */
    .mp-sb-brand {
        display: flex; align-items: center; gap: 10px;
        padding: 6px 8px 12px;
    }
    .mp-sb-brand .logo {
        width: 36px; height: 36px; border-radius: 9px;
        background: var(--mp-green); color: white;
        display: flex; align-items: center; justify-content: center;
        font-weight: 800; font-size: 18px; letter-spacing: -.02em;
        box-shadow: 0 2px 6px rgba(2,137,18,.25);
    }
    .mp-sb-brand .name { font-size: 15px; font-weight: 700; letter-spacing: -.01em; color: var(--mp-ink); }
    .mp-sb-brand .tag  { font-size: 10.5px; color: var(--mp-ink-50); letter-spacing: .04em; text-transform: uppercase; }

    .mp-sb-user {
        background: var(--mp-bg);
        border: 1px solid var(--mp-line);
        border-radius: var(--mp-radius-sm);
        padding: 10px 12px;
        margin-bottom: 8px;
    }
    .mp-sb-user .row { display: flex; align-items: center; gap: 9px; }
    .mp-sb-user .avatar {
        width: 30px; height: 30px; border-radius: 999px;
        background: var(--mp-green-soft); color: var(--mp-green-dark);
        display: flex; align-items: center; justify-content: center;
        font-weight: 700; font-size: 12px;
        flex: none;
    }
    .mp-sb-user .name { font-size: 12.5px; font-weight: 600; color: var(--mp-ink); white-space: nowrap; overflow: hidden; text-overflow: ellipsis; }
    .mp-sb-user .meta { font-size: 10.5px; color: var(--mp-ink-50); }
    .mp-sb-user .meta .role-pill {
        font-weight: 700; letter-spacing: .04em;
    }
    .mp-sb-user .meta .role-pill.admin { color: var(--mp-green); }
    .mp-sb-user .meta .role-pill.user  { color: var(--mp-ink-70); }

    .mp-sb-section {
        font-size: 10.5px; font-weight: 700;
        color: var(--mp-ink-50);
        letter-spacing: .08em; text-transform: uppercase;
        margin: 14px 4px 4px;
    }

    .mp-sb-stat {
        display: flex; justify-content: space-between;
        padding: 4px 8px; font-size: 11.5px; color: var(--mp-ink-50);
    }
    .mp-sb-stat b { color: var(--mp-ink); font-weight: 600; }
    .mp-sb-stat b.green { color: var(--mp-green); }
</style>
""", unsafe_allow_html=True)

SESSION_TIMEOUT_MINUTES = 720  # 12 jam
LOGIN_FOLDER    = Path("login")
DATA_FOLDER     = Path("data")
CACHE_FOLDER    = Path(".cache")
IMAGES_FOLDER   = Path("images")


# ── Login Manager ───────────────────────────────────────────────────
class LoginManager:
    def __init__(self):
        LOGIN_FOLDER.mkdir(parents=True, exist_ok=True)
        if "login_users_df" not in st.session_state:
            st.session_state.login_users_df = self._load_users()

    # ── Load Users ──────────────────────────────────────────────────────────
    @staticmethod
    def _load_users() -> pd.DataFrame:
        """
        Prioritas:
          1. Supabase (jika terkonfigurasi)
          2. Fallback ke file Excel di folder /login (perilaku lama)
        """
        if SUPABASE_ENABLED and load_users_from_supabase is not None:
            df = load_users_from_supabase()
            if not df.empty:
                return df
            # Jika Supabase kosong, tetap coba Excel sebagai fallback
        return LoginManager._load_users_from_excel()

    @staticmethod
    def _load_users_from_excel() -> pd.DataFrame:
        """Load user dari file Excel di folder /login (perilaku lama)."""
        excel_ext = (".xlsx", ".xls", ".xlsm")
        all_rows  = []
        for fpath in LOGIN_FOLDER.iterdir():
            if fpath.suffix.lower() not in excel_ext:
                continue
            try:
                xls = pd.ExcelFile(fpath, engine="openpyxl")
                for sheet in xls.sheet_names:
                    df = pd.read_excel(xls, sheet_name=sheet, dtype=str, header=None)
                    if len(df) == 0:
                        continue
                    first = df.iloc[0].astype(str).str.strip().str.lower().tolist()
                    if any(v in ("username", "user", "nama") for v in first):
                        df = df.iloc[1:].reset_index(drop=True)
                    if len(df.columns) >= 4:
                        df = df.iloc[:, 1:4]
                    elif len(df.columns) == 3:
                        pass
                    else:
                        continue
                    df.columns = ["username", "password", "role"]
                    df = df.dropna(subset=["username", "password"])
                    df["username"] = df["username"].str.strip().str.lower()
                    df["password"] = df["password"].str.strip()
                    df["role"]     = df["role"].str.strip().str.lower().fillna("user")
                    all_rows.append(df)
            except Exception:
                continue
        if all_rows:
            return pd.concat(all_rows, ignore_index=True).drop_duplicates(subset=["username"])
        return pd.DataFrame(columns=["username", "password", "role"])

    # ── Authenticate ────────────────────────────────────────────────────────
    def authenticate(self, username: str, password: str):
        """
        Autentikasi user.
        Jika Supabase aktif, verifikasi dari Supabase.
        Fallback ke DataFrame (Excel) jika Supabase tidak tersedia.
        """
        if SUPABASE_ENABLED and authenticate_from_supabase is not None:
            return authenticate_from_supabase(username, password)
        # Fallback: verifikasi dari DataFrame yang sudah di-load
        return self._authenticate_from_df(username, password)

    def _authenticate_from_df(self, username: str, password: str):
        """Autentikasi dari DataFrame (perilaku lama — Excel-based)."""
        df = st.session_state.login_users_df
        if df.empty:
            return None
        username = username.strip().lower()
        row = df[df["username"] == username]
        if row.empty:
            return None
        if hmac.compare_digest(password.strip(), row.iloc[0]["password"]):
            return {
                "username":    username,
                "role":        row.iloc[0]["role"],
                "login_time":  datetime.now(),
                "last_active": datetime.now(),
            }
        return None

    # ── Session helpers ─────────────────────────────────────────────────────
    @staticmethod
    def init_session():
        for k, v in {"is_logged_in": False, "current_user": None, "login_error": None}.items():
            if k not in st.session_state:
                st.session_state[k] = v

    @staticmethod
    def is_authenticated() -> bool:
        if not st.session_state.get("is_logged_in"):
            return False
        user = st.session_state.get("current_user")
        if user is None:
            return False
        elapsed = (datetime.now() - user["last_active"]).total_seconds() / 60
        if elapsed > SESSION_TIMEOUT_MINUTES:
            LoginManager.logout(reason="session_expired")
            st.session_state["login_error"] = "⏰ Sesi telah berakhir. Silakan login ulang."
            return False
        user["last_active"] = datetime.now()
        # Refresh online status (throttled di dalam helper).
        try:
            touch_active(user.get("username", ""))
        except Exception:
            pass
        return True

    @staticmethod
    def logout(reason: str = "manual"):
        user = st.session_state.get("current_user") or {}
        uname = user.get("username", "")
        if uname:
            try:
                log_activity(uname, "logout", details={"reason": reason})
            except Exception:
                pass
        st.session_state["is_logged_in"] = False
        st.session_state["current_user"] = None

        # Cleanup session_state besar supaya RAM langsung dibebasin saat
        # logout (bukan nunggu Streamlit GC session). Reference ke shared
        # cache (excel_files) putus → data shared TETAP hidup di
        # _load_excel_index_shared untuk user lain.
        STATIC_KEYS = (
            "excel_files", "index_data", "search_results",
            "harga_data", "harga_lookup", "stok_data",
            "stok_gudang_data", "stok_gudang_names",
            "last_index_time", "loaded_files_count", "last_file_count",
            "file_hashes", "search_type", "search_term",
            "_pn_scroll_to_image", "_img_idx_results_ts",
        )
        for k in STATIC_KEYS:
            st.session_state.pop(k, None)

        # Pattern-based keys (dynamic — di-keyed by PN / sub-feature).
        # Hindari ngutak-atik widget key (login form, dll) — cek di awal
        # supaya safe.
        DYNAMIC_PREFIXES = (
            "sims_fetched_", "sims_err_", "sims_part_info_",
            "img_idx_", "local_img_idx_",
            "_img_search_",
            "bhe_",
        )
        for k in list(st.session_state.keys()):
            if any(k.startswith(p) for p in DYNAMIC_PREFIXES):
                st.session_state.pop(k, None)

    @staticmethod
    def get_current_user():
        return st.session_state.get("current_user")


def render_login_page(login_mgr: LoginManager):
    error_msg = st.session_state.get("login_error")
    inject_keep_alive()

    # Background gradient untuk login page + center content vertikal
    st.markdown(
        """
        <style>
        [data-testid="stAppViewContainer"] {
            background: linear-gradient(180deg, #FFFFFF 0%, #F1FAF2 100%) !important;
            min-height: 100vh !important;
        }
        .stApp { background: linear-gradient(180deg, #FFFFFF 0%, #F1FAF2 100%) !important; }
        header[data-testid="stHeader"] { display: none !important; }
        div[data-testid="stToolbar"] { display: none !important; }
        [data-testid="stSidebar"] { display: none !important; }
        section.main, section[data-testid="stMain"], [data-testid="stMain"] {
            background: transparent !important;
            min-height: 100vh !important;
            display: flex !important;
            flex-direction: column !important;
            justify-content: center !important;
            align-items: center !important;
        }
        section.main > div.block-container,
        section[data-testid="stMain"] > div.block-container,
        div[data-testid="stMainBlockContainer"],
        .block-container {
            padding: 2rem 3rem !important;
            margin: 0 auto !important;
            max-width: 1500px !important;
            width: 100% !important;
            flex: 0 0 auto !important;
        }
        </style>
        """,
        unsafe_allow_html=True,
    )

    col_brand, col_card = st.columns([1.2, 1.0], gap="large")

    # ── Left: brand + tagline + stats ─────────────────────────────────
    with col_brand:
        # Hitung file Excel langsung dari filesystem supaya angka valid
        # sejak login page (file belum di-index sebelum login).
        n_files = st.session_state.get("loaded_files_count", 0)
        if not n_files:
            try:
                n_files = sum(1 for _ in DATA_FOLDER.rglob("*.xlsx")
                              if not _.name.startswith("~$"))
            except Exception:
                n_files = 0
        n_files_str = f"{n_files:,}" if n_files else "—"
        st.markdown(
            f"""
            <div>
                <div style="display:inline-flex; align-items:center; gap:12px; margin-bottom:24px;">
                    <div style="width:52px;height:52px;border-radius:12px;background:var(--mp-green);
                                color:white;display:flex;align-items:center;justify-content:center;
                                font-weight:800;font-size:26px;letter-spacing:-.02em;
                                box-shadow:0 6px 18px rgba(2,137,18,.25);">M</div>
                    <div>
                        <div style="font-size:26px;font-weight:800;letter-spacing:-.02em;line-height:1;">MasPart</div>
                        <div style="font-size:11.5px;color:var(--mp-ink-50);margin-top:4px;
                                    letter-spacing:.04em;text-transform:uppercase;font-weight:600;">
                            Part Number Finder
                        </div>
                    </div>
                </div>
                <h1 style="font-size:34px;font-weight:700;line-height:1.18;letter-spacing:-.02em;
                            margin:0 0 16px;color:var(--mp-ink);">
                    Cari sparepart <span style="color:var(--mp-green);">tanpa buka katalog manual</span>.
                </h1>
                <p style="font-size:14px;color:var(--mp-ink-70);line-height:1.6;margin:0 0 24px;max-width:440px;">
                    Database part Sinotruk · Shantui · Weichai, lengkap dengan foto SIMS,
                    harga real-time, stok opname per user, dan tools batch.
                </p>
                <div style="display:flex; gap:24px; flex-wrap:wrap;">
                    <div>
                        <div style="font-size:22px;font-weight:700;color:var(--mp-green);">{n_files_str}</div>
                        <div style="font-size:11px;color:var(--mp-ink-50);letter-spacing:.04em;
                                    text-transform:uppercase;font-weight:600;">File Excel</div>
                    </div>
                    <div>
                        <div style="font-size:22px;font-weight:700;color:var(--mp-green);">3</div>
                        <div style="font-size:11px;color:var(--mp-ink-50);letter-spacing:.04em;
                                    text-transform:uppercase;font-weight:600;">Brand</div>
                    </div>
                    <div>
                        <div style="font-size:22px;font-weight:700;color:var(--mp-green);">●</div>
                        <div style="font-size:11px;color:var(--mp-ink-50);letter-spacing:.04em;
                                    text-transform:uppercase;font-weight:600;">Supabase Live</div>
                    </div>
                </div>
            </div>
            """,
            unsafe_allow_html=True,
        )

    # ── Right: login card ─────────────────────────────────────────────
    with col_card:
        st.markdown(
            '<div style="background:white;border:1px solid var(--mp-line);border-radius:14px;'
            'padding:28px;box-shadow:0 10px 40px rgba(15,26,18,.08), 0 2px 6px rgba(15,26,18,.04);">'
            '<div style="font-size:18px;font-weight:700;letter-spacing:-.01em;">Masuk ke akun Anda</div>'
            f'<div style="font-size:12.5px;color:var(--mp-ink-50);margin-top:4px;">'
            f'Sesi aktif {SESSION_TIMEOUT_MINUTES // 60} jam setelah login.</div>'
            '</div>',
            unsafe_allow_html=True,
        )
        if error_msg:
            st.error(error_msg, icon="⚠️")
            st.session_state["login_error"] = None
        with st.form(key="login_form", clear_on_submit=False):
            username  = st.text_input("Username", placeholder="Masukkan username")
            password  = st.text_input("Password", type="password", placeholder="••••••••")
            submitted = st.form_submit_button("Login →", type="primary", use_container_width=True)
        st.markdown(
            '<div style="font-size:11.5px;color:var(--mp-ink-50);margin-top:14px;text-align:center;">'
            'Butuh akun? Hubungi admin'
            '</div>',
            unsafe_allow_html=True,
        )

    if submitted:
        if not username or not password:
            st.session_state["login_error"] = "Username dan password tidak boleh kosong."
            st.rerun()
        user_info = login_mgr.authenticate(username, password)
        if user_info:
            st.session_state["is_logged_in"] = True
            st.session_state["current_user"] = user_info
            st.session_state["login_error"]  = None
            st.rerun()
        else:
            st.session_state["login_error"] = "Username atau password salah."
            st.rerun()


# ── Parser Stok (single-total + multi-gudang) ───────────────────────
_GUDANG_HEADER_RE = re.compile(r"^\s*\d+\s*\.")          # "01.Jakarta", "25. PT BJM"
_KODE_PREFIX_RE   = re.compile(r"^\d{6}\.")              # "000001." pada Kode Barang


def _stok_to_int(v):
    """Coerce nilai qty gudang jadi int. Return 0 kalau kosong/non-numerik."""
    if v is None:
        return 0
    s = str(v).strip()
    if not s or s.lower() in ("nan", "none", "—", "-"):
        return 0
    s = s.replace(",", "").replace(".", "")
    try:
        return int(float(s))
    except Exception:
        return 0


def parse_stok_file(file_bytes):
    """
    Parse file stok. Mendukung 2 format:

      1. FORMAT LAMA (sederhana) — Kol A = Part Number, Kol D = Stok total.
      2. FORMAT MULTI-GUDANG (export Accurate "Kuantitas Barang per Gudang") —
         baris header memuat 'Kode Barang' + banyak kolom gudang ('01.Jakarta',
         '02.Pekanbaru', …) + kolom 'Total'. Kode Barang berformat
         'NNNNNN.PARTNUMBER' (prefix 6 digit dibuang).

    Return tuple:
      stok_cache   : {PN: "total"}                — kompatibel dgn kode lama
      gudang_cache : {PN: {nama_gudang: qty_int}} — hanya gudang ber-qty != 0
      gudang_names : [nama_gudang, …]             — urutan kolom gudang
    """
    try:
        raw = pd.read_excel(io.BytesIO(file_bytes), header=None, dtype=str)
    except Exception as e:
        raise RuntimeError(f"Gagal baca Excel stok: {e}")

    if raw.empty:
        return {}, {}, []

    # ── Cari baris header multi-gudang (col 0 == "Kode Barang") ──────
    header_idx = None
    for i in range(min(15, len(raw))):
        c0 = str(raw.iloc[i, 0]).strip().lower()
        if c0 in ("kode barang", "kode barang "):
            header_idx = i
            break

    # ── FORMAT LAMA (tidak ada header multi-gudang) ──────────────────
    if header_idx is None:
        df = raw
        if len(df) > 0 and any(
            str(x).lower() in ["part number", "kode", "no part"] for x in df.iloc[0]
        ):
            df = df.iloc[1:]
        ncol = df.shape[1]
        pn_i  = 0
        stk_i = 3 if ncol > 3 else (ncol - 1)
        stok_cache = {}
        for _, row in df.iterrows():
            pn = str(row.iloc[pn_i]).strip().upper()
            if not pn or pn in ("NAN", "NONE"):
                continue
            val = row.iloc[stk_i]
            stok_cache[pn] = "—" if pd.isna(val) else str(val).strip()
        return stok_cache, {}, []

    # ── FORMAT MULTI-GUDANG ──────────────────────────────────────────
    headers = [str(x).strip() if not pd.isna(x) else "" for x in raw.iloc[header_idx]]

    kode_i  = 0
    total_i = None
    gudang_cols = []   # (col_index, nama_gudang)
    for ci, h in enumerate(headers):
        hl = h.lower()
        if hl == "kode barang":
            kode_i = ci
        elif hl.startswith("total"):
            total_i = ci
        elif _GUDANG_HEADER_RE.match(h):
            gudang_cols.append((ci, h))

    gudang_names = [name for _, name in gudang_cols]

    stok_cache   = {}
    gudang_cache = {}
    for ri in range(header_idx + 1, len(raw)):
        row  = raw.iloc[ri]
        kode = str(row.iloc[kode_i]).strip()
        if not kode or kode.lower() in ("nan", "none"):
            continue
        pn = _KODE_PREFIX_RE.sub("", kode).strip().upper()
        if not pn:
            continue

        # Total: pakai kolom Total kalau ada, kalau tidak jumlahkan semua gudang
        if total_i is not None:
            total_val = _stok_to_int(row.iloc[total_i])
        else:
            total_val = sum(_stok_to_int(row.iloc[ci]) for ci, _ in gudang_cols)

        breakdown = {}
        for ci, name in gudang_cols:
            q = _stok_to_int(row.iloc[ci])
            if q != 0:
                breakdown[name] = q

        stok_cache[pn] = str(total_val)
        gudang_cache[pn] = breakdown

    return stok_cache, gudang_cache, gudang_names


# ── Search Functions ────────────────────────────────────────────────
def search_part_number(term, excel_files, stok_cache, harga_lookup=None):
    results, seen = [], set()
    term_up = term.strip().upper()
    if not term_up:
        return results

    harga_lookup = harga_lookup or {}

    for fi in excel_files:
        sn = fi["simple_name"]
        if sn in seen:
            continue
        df = fi["dataframe"]
        for indexed_pn, indices in fi.get("part_number_index", {}).items():
            if term_up in indexed_pn:
                row        = df.iloc[indices[0]]
                pn_value   = str(row["part_number"]).strip() if pd.notna(row["part_number"]) else "N/A"
                stok_value = stok_cache.get(pn_value.upper(), "—") if stok_cache else "—"
                harga_value = harga_lookup.get(pn_value.upper(), "—")
                results.append({
                    "File": sn, "Path": fi["relative_path"], "Sheet": fi["sheet"],
                    "Part Number": pn_value,
                    "Part Name": str(row["part_name"]) if pd.notna(row["part_name"]) else "N/A",
                    "Quantity": str(row["quantity"]) if pd.notna(row["quantity"]) else "N/A",
                    "Stok": stok_value, "Harga": harga_value,
                    "Excel Row": indices[0] + 2, "Full Path": fi["full_path"]
                })
                seen.add(sn)
                break
    return results


def search_part_name(term, excel_files, stok_cache, harga_lookup=None):
    """
    Cari berdasarkan Part Name.
    """
    results = []
    term_up = term.strip().upper()
    if not term_up:
        return results

    harga_lookup = harga_lookup or {}
    search_keywords = [term.strip().lower()]

    for fi in excel_files:
        df  = fi["dataframe"]
        pni = fi.get("part_name_index", {})
        matching_indices = set()

        for keyword in search_keywords:
            kw_up        = keyword.upper()
            search_words = kw_up.split()
            for word in pni.keys():
                for sw in search_words:
                    if sw in word or word in sw:
                        matching_indices.update(pni[word])
            # Fallback untuk keyword pendek (≤3 huruf)
            if not matching_indices and len(kw_up) <= 3:
                for idx, row in df.iterrows():
                    pname = str(row["part_name"]) if pd.notna(row["part_name"]) else ""
                    if kw_up in pname.upper():
                        matching_indices.add(idx)

        for idx in matching_indices:
            row   = df.iloc[idx]
            pname = str(row["part_name"]) if pd.notna(row["part_name"]) else ""
            # Harus cocok dengan keyword
            matched = any(kw.upper() in pname.upper() for kw in search_keywords)
            if matched:
                pn_value   = str(row["part_number"]).strip() if pd.notna(row["part_number"]) else "N/A"
                stok_value = stok_cache.get(pn_value.upper(), "—") if stok_cache else "—"
                harga_value = harga_lookup.get(pn_value.upper(), "—")
                results.append({
                    "File": fi["simple_name"], "Path": fi["relative_path"], "Sheet": fi["sheet"],
                    "Part Number": pn_value, "Part Name": pname if pname else "N/A",
                    "Quantity": str(row["quantity"]) if pd.notna(row["quantity"]) else "N/A",
                    "Stok": stok_value, "Harga": harga_value, "Excel Row": idx + 2, "Full Path": fi["full_path"]
                })
    return results


# ── Build Excel Functions ───────────────────────────────────────────

def build_catalog_excel(df_result: pd.DataFrame, progress_callback=None, all_part_numbers: list = None, options: dict = None) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.drawing.image import Image as XLImage
    from PIL import Image as PILImage
    import tempfile

    # Default semua aktif jika options tidak diberikan
    if options is None:
        options = {}
    inc_partname  = options.get("partname",  True)
    inc_kecocokan = options.get("kecocokan", True)
    inc_stok      = options.get("stok",      True)
    inc_qty       = options.get("qty",       True)
    inc_images    = options.get("images",    True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Catalog"

    header_fill = PatternFill("solid", fgColor="1565C0")
    header_font = Font(bold=True, color="FFFFFF", name="Arial", size=11)
    center      = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left        = Alignment(horizontal="left",   vertical="center", wrap_text=True)
    thin        = Side(style="thin", color="BDBDBD")
    border      = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Bangun header dinamis sesuai pilihan
    headers    = ["Part Number"]
    col_widths = [20]
    if inc_partname:
        headers.append("Part Name"); col_widths.append(30)
    if inc_kecocokan:
        headers.append("Kecocokan"); col_widths.append(45)
    if inc_qty:
        headers.append("Qty"); col_widths.append(12)
    if inc_stok:
        headers.append("Stok"); col_widths.append(12)
    if inc_images:
        headers.extend(["Gambar 1", "Gambar 2"]); col_widths.extend([38, 38])

    for ci, (h, w) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font = header_font; cell.fill = header_fill
        cell.alignment = center; cell.border = border
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[1].height = 22

    fill_even = PatternFill("solid", fgColor="E3F2FD")
    fill_odd  = PatternFill("solid", fgColor="FAFAFA")
    fill_nf   = PatternFill("solid", fgColor="FFEBEE")

    # Bangun lookup dari df_result berdasarkan index urutan (bukan PN sebagai key)
    # sehingga duplikat PN tetap muncul sebagai baris terpisah
    df_lookup = {}
    for _, r in df_result.iterrows():
        pn = r["_pn_group"]
        if pn not in df_lookup:
            df_lookup[pn] = {
                "Part Name": r.get("Part Name", ""),
                "kecocokan": r.get("Hasil", ""),
                "found":     r.get("Status", "") == "✅ Ditemukan",
                "Qty":       r.get("Qty", ""),
                "Stok":      r.get("Stok", ""),
            }

    # Gunakan urutan asli dari all_part_numbers — TANPA deduplikasi
    # sehingga PN yang muncul 2x di input tetap 2 baris di output
    pn_order = all_part_numbers if all_part_numbers else list(dict.fromkeys(df_result["_pn_group"].tolist()))

    # Bangun grouped sebagai list (bukan dict) agar duplikat tidak hilang
    grouped_list = []
    for pn in pn_order:
        if pn in df_lookup:
            grouped_list.append((pn, df_lookup[pn]))
        else:
            grouped_list.append((pn, {"Part Name": "", "kecocokan": "", "found": False}))

    def _make_xl_image(img_bytes, max_h=200):
        pil_img = PILImage.open(io.BytesIO(img_bytes)).convert("RGB")
        w_px, h_px = pil_img.size
        if h_px > max_h:
            ratio  = max_h / h_px
            w_px   = int(w_px * ratio)
            h_px   = max_h
            pil_img = pil_img.resize((w_px, h_px), PILImage.LANCZOS)
        tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".png")
        pil_img.save(tmp.name, format="PNG")
        tmp.close()
        xl = XLImage(tmp.name)
        xl.width  = w_px
        xl.height = h_px
        return xl, w_px, h_px, tmp.name

    tmp_files = []
    row_idx   = 2
    total_pn  = len(grouped_list)

    # Hitung indeks kolom gambar secara dinamis
    img_col_start = 1 + 1 + int(inc_partname) + int(inc_kecocokan) + int(inc_qty) + int(inc_stok)
    img_col_d_letter = get_column_letter(img_col_start)
    img_col_e_letter = get_column_letter(img_col_start + 1)

    for i, (pn, info) in enumerate(grouped_list):
        if progress_callback:
            progress_callback(i, total_pn, pn)

        kecocokan  = info["kecocokan"] if info["kecocokan"] else "—"
        is_found   = info["found"]
        part_name  = info["Part Name"]
        fill       = (fill_even if i % 2 == 0 else fill_odd) if is_found else fill_nf
        row_height = 22 if not inc_images else 80
        img_d      = None
        img_e      = None

        # ── Fallback Part Name dari SIMS (selalu berjalan jika SIMS aktif) ──
        if SIMS_ENABLED and not part_name:
            try:
                from sims_fetcher import get_sims_part_info
                sims_info, _ = get_sims_part_info(pn)
                if sims_info and sims_info.get("partName"):
                    part_name = sims_info["partName"]
            except Exception:
                pass

        # ── Fetch gambar dari SIMS (hanya jika opsi gambar aktif) ──
        if inc_images and SIMS_ENABLED:
            try:
                urls, _ = _sims_fetch(pn)
                if urls:
                    b1, _ = ExcelSearchApp.fetch_image_bytes(urls[0])
                    if b1:
                        xl, w, h, tmp_path = _make_xl_image(b1)
                        img_d = xl
                        tmp_files.append(tmp_path)
                        row_height = max(int(h * 0.75) + 10, row_height)
                        hash1 = hashlib.md5(b1).hexdigest()
                        for url2 in urls[1:]:
                            b2, _ = ExcelSearchApp.fetch_image_bytes(url2)
                            if b2 and hashlib.md5(b2).hexdigest() != hash1:
                                xl, w, h, tmp_path = _make_xl_image(b2)
                                img_e = xl
                                tmp_files.append(tmp_path)
                                row_height = max(int(h * 0.75) + 10, row_height)
                                break
            except Exception as e:
                print(f"[catalog] Gagal ambil gambar {pn}: {e}")

        ws.row_dimensions[row_idx].height = row_height

        # Isi kolom secara dinamis sesuai pilihan
        ci = 1
        for val, aln in [(pn, center)]:
            cell = ws.cell(row=row_idx, column=ci, value=val)
            cell.fill = fill; cell.border = border
            cell.alignment = aln; cell.font = Font(name="Arial", size=10)
            ci += 1
        if inc_partname:
            cell = ws.cell(row=row_idx, column=ci, value=part_name)
            cell.fill = fill; cell.border = border
            cell.alignment = left; cell.font = Font(name="Arial", size=10)
            ci += 1
        if inc_kecocokan:
            cell = ws.cell(row=row_idx, column=ci, value=kecocokan)
            cell.fill = fill; cell.border = border
            cell.alignment = left; cell.font = Font(name="Arial", size=10)
            ci += 1
        if inc_qty:
            cell = ws.cell(row=row_idx, column=ci, value=info.get("Qty", ""))
            cell.fill = fill; cell.border = border
            cell.alignment = center; cell.font = Font(name="Arial", size=10)
            ci += 1
        if inc_stok:
            cell = ws.cell(row=row_idx, column=ci, value=info.get("Stok", ""))
            cell.fill = fill; cell.border = border
            cell.alignment = center; cell.font = Font(name="Arial", size=10)
            ci += 1
        if inc_images:
            for _ in range(2):
                c = ws.cell(row=row_idx, column=ci, value="")
                c.fill = fill; c.border = border; c.alignment = center
                ci += 1
            if img_d:
                ws.add_image(img_d, f"{img_col_d_letter}{row_idx}")
            if img_e:
                ws.add_image(img_e, f"{img_col_e_letter}{row_idx}")

        row_idx += 1

    ws.freeze_panes = "A2"
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    result = buf.getvalue()

    for f in tmp_files:
        try:
            os.unlink(f)
        except Exception:
            pass

    return result


def make_template_excel() -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    wb = Workbook()
    ws = wb.active
    ws.title = "Part Number List"
    ws["A1"] = "Part Number"
    ws["A1"].font      = Font(bold=True, name="Arial", size=11, color="FFFFFF")
    ws["A1"].fill      = PatternFill("solid", fgColor="1565C0")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 20
    for i, ex in enumerate(["WG1642821034", "WG9925520270", "AZ9100443082", "WG9718820030"], start=2):
        ws.cell(row=i, column=1, value=ex).font = Font(name="Arial", size=10)
    ws.column_dimensions["A"].width = 28
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf.getvalue()


# ── Shared Excel Index (cross-session, RAM-efficient) ───────────────
# Sebelum-nya index 421 file Excel disimpan di st.session_state PER user
# (~500 MB/user) → 2–3 user concurrent langsung kena resource limit
# Streamlit Cloud (~1 GB). @st.cache_resource simpan 1 copy di proses
# yang dishare ke semua session. RAM cuma sekali untuk semua user.
@st.cache_resource(
    show_spinner="📚 Memuat index Excel (sekali saja, di-share ke semua user)..."
)
def _load_excel_index_shared(data_folder_str: str, _processor):
    """
    Walk data folder, parse semua .xlsx/.xls/.xlsm, return list entry per
    sheet. Cached antar session — re-run cuma terjadi kalau cache di-clear
    (mis. lewat tombol Refresh Data).

    Parameter `_processor` di-prefix underscore supaya Streamlit skip dari
    hashing key (callable tidak bisa di-hash konsisten). Cache key efektif
    cuma `data_folder_str`.
    """
    data_folder = Path(data_folder_str)
    excel_ext = (".xlsx", ".xls", ".xlsm")
    all_files = []
    for root, _, files in os.walk(data_folder):
        for f in files:
            if f.lower().endswith(excel_ext):
                fp = Path(root) / f
                all_files.append((fp, fp.relative_to(data_folder)))

    if not all_files:
        return [], 0, datetime.now()

    results = []
    with ThreadPoolExecutor(max_workers=4) as executor:
        futures = {executor.submit(_processor, fp, rp): fp for fp, rp in all_files}
        for future in as_completed(futures):
            try:
                res = future.result()
                if res:
                    results.extend(res)
            except Exception:
                pass

    return results, len(all_files), datetime.now()


# ── Pickle disk cache size limit (LRU eviction) ─────────────────────
# Streamlit Cloud free tier punya disk ~1 GB share dengan kode + deps +
# image cache. Tanpa cap, pickle cache bisa tumbuh tak terbatas (orphan
# kalau Excel sumber dihapus). Helper di bawah ini menjaga total size
# `.cache/*.pkl` di bawah cap dengan eviction LRU (by mtime).
_PICKLE_CACHE_MAX_MB = 500

def _enforce_pickle_cache_size_limit(max_mb: int = _PICKLE_CACHE_MAX_MB) -> None:
    try:
        files = sorted(
            CACHE_FOLDER.glob("*.pkl"), key=lambda f: f.stat().st_mtime
        )
        total = sum(f.stat().st_size for f in files)
        cap = max_mb * 1024 * 1024
        while total > cap and files:
            old = files.pop(0)
            try:
                sz = old.stat().st_size
                old.unlink()
                total -= sz
            except Exception:
                pass
    except Exception:
        pass


# ── Main App ────────────────────────────────────────────────────────
class ExcelSearchApp:
    def __init__(self):
        self.data_folder   = DATA_FOLDER
        self.cache_folder  = CACHE_FOLDER
        self.images_folder = IMAGES_FOLDER
        self.supported_ext = [".jpg", ".jpeg", ".png"]
        self.cache_folder.mkdir(exist_ok=True)
        self.images_folder.mkdir(exist_ok=True)
        self.stok_file       = DATA_FOLDER / "stok" / "stok.xlsx"
        self.stok_cache      = None
        self.stok_gudang_cache = None   # {PN: {nama_gudang: qty}} — format multi-gudang
        self.gudang_names      = []     # urutan kolom gudang
        self.harga_file      = DATA_FOLDER / "harga" / "harga.xlsx"
        self.harga_cache     = None
        self.harga_lookup    = {}  # dict PN -> "Rp xxx" — dibangun sekali saat load
        self.populasi_folder = DATA_FOLDER / "populasi"
        self._load_stok_data()
        self._load_harga_data()

        if "excel_files" not in st.session_state:
            st.session_state.index_data         = []
            st.session_state.search_results     = []
            st.session_state.file_hashes        = {}

        # SELALU overwrite excel_files dengan reference dari shared cache —
        # bukan conditional. Alasan:
        #   1. Cache_resource memang sudah men-dedup load (cache hit = return
        #      reference, biaya ~zero).
        #   2. Kalau session_state.excel_files masih nyimpan copy besar dari
        #      deploy sebelumnya (sebelum refactor ini), conditional skip
        #      bakal nahan stale copy itu di RAM. Assign unconditional bikin
        #      Python GC bebasin copy lama begitu reference terakhir hilang.
        # Semua user end up share 1 list yang sama di RAM.
        results, n_total_files, idx_time = _load_excel_index_shared(
            str(self.data_folder), self.process_single_file
        )
        st.session_state.excel_files        = results
        st.session_state.loaded_files_count = len(results)
        st.session_state.last_file_count    = n_total_files
        st.session_state.last_index_time    = idx_time

    def create_data_folder(self):
        if not self.data_folder.exists():
            self.data_folder.mkdir(parents=True)

    def get_file_hash(self, fp):
        try:
            s = fp.stat()
            return hashlib.md5(f"{fp}_{s.st_size}_{s.st_mtime}".encode()).hexdigest()
        except Exception:
            return None

    def load_file_cache(self, fp, fh):
        cf = self.cache_folder / f"{fh}.pkl"
        if cf.exists():
            try:
                with open(cf, "rb") as f:
                    return pickle.load(f)
            except Exception:
                return None
        return None

    def save_file_cache(self, fp, fh, data):
        try:
            with open(self.cache_folder / f"{fh}.pkl", "wb") as f:
                pickle.dump(data, f, protocol=pickle.HIGHEST_PROTOCOL)
            # Setelah tiap save, cek total cache size — evict yang paling tua
            # kalau lewat cap. Hindari disk full di Streamlit Cloud.
            _enforce_pickle_cache_size_limit()
        except Exception:
            pass

    @staticmethod
    def extract_simple_filename(filename):
        name = os.path.splitext(filename)[0]
        return name.split(" - ")[-1] if " - " in name else name

    def normalize_base_part_number(self, pn):
        if not pn or pd.isna(pn):
            return ""
        pn_str = str(pn).strip().upper()
        base   = pn_str.split("/", 1)[0].strip()
        return re.sub(r'[^A-Z0-9\-]', '_', base)

    def get_image_path(self, pn):
        base = self.normalize_base_part_number(pn)
        if not base:
            return None
        sub_folder = self.images_folder / base
        if sub_folder.exists() and sub_folder.is_dir():
            for ext in self.supported_ext:
                candidates = sorted(sub_folder.glob(f"*{ext}"))
                if candidates:
                    return candidates[0]
        for ext in self.supported_ext:
            p = self.images_folder / f"{base}{ext}"
            if p.exists():
                return p
        return None

    def get_all_image_paths(self, pn):
        base = self.normalize_base_part_number(pn)
        if not base:
            return []
        paths      = []
        sub_folder = self.images_folder / base
        if sub_folder.exists() and sub_folder.is_dir():
            for ext in self.supported_ext:
                paths.extend(sorted(sub_folder.glob(f"*{ext}")))
        for ext in self.supported_ext:
            p = self.images_folder / f"{base}{ext}"
            if p.exists() and p not in paths:
                paths.append(p)
        return paths

    @staticmethod
    def render_zoomable_image(img_bytes: bytes, caption: str = "", zoom_key: str = "zoom_default"):
        import base64
        zk = f"zoom_scale_{zoom_key}"
        if zk not in st.session_state:
            st.session_state[zk] = 100

        scale = st.session_state[zk]
        c1, c2, c3, c4 = st.columns([1, 1, 1, 3])
        with c1:
            if st.button("🔍＋", key=f"zi_{zoom_key}", help="Zoom In", use_container_width=True):
                st.session_state[zk] = min(scale + 25, 300)
                st.rerun()
        with c2:
            if st.button("🔍－", key=f"zo_{zoom_key}", help="Zoom Out", use_container_width=True):
                st.session_state[zk] = max(scale - 25, 25)
                st.rerun()
        with c3:
            if st.button("⟳", key=f"zr_{zoom_key}", help="Reset zoom", use_container_width=True):
                st.session_state[zk] = 100
                st.rerun()
        with c4:
            st.markdown(
                f"<div style='padding:6px 0;color:#555;font-size:.85rem;'>Zoom: <b>{st.session_state[zk]}%</b></div>",
                unsafe_allow_html=True
            )

        b64  = base64.b64encode(img_bytes).decode()
        sig  = img_bytes[:4]
        mime = "image/jpeg"
        if sig[:4] == b'\x89PNG':
            mime = "image/png"
        elif sig[:3] == b'GIF':
            mime = "image/gif"

        cur_scale    = st.session_state[zk]
        safe_caption = caption.replace("<", "&lt;").replace(">", "&gt;")
        st.markdown(f"""
<div style="overflow:auto; width:100%; text-align:center; padding:4px 0;">
  <img src="data:{mime};base64,{b64}"
       style="width:{cur_scale}%; max-width:none; transform-origin:top center;
              border-radius:8px; box-shadow:0 2px 12px rgba(0,0,0,.18); transition:width .2s ease;"
       title="{safe_caption}" />
  <div style="font-size:.78rem;color:#666;margin-top:4px;">{safe_caption}</div>
</div>""", unsafe_allow_html=True)

    @staticmethod
    @st.cache_data(ttl=1800, max_entries=200, show_spinner=False)
    def fetch_image_bytes(url: str):
        """
        Download image bytes dari URL (SIMS / Supabase / public).
        Hasil di-cache 30 menit (TTL=1800s, max 200 URL distinct).
        Tanpa cache, setiap rerun (klik tombol, ganti tab) re-download
        gambar yang sama → boros bandwidth + latency. Cache_data evict
        otomatis kalau lewat max_entries.
        """
        try:
            headers = {"User-Agent": "Mozilla/5.0"}
            if SIMS_ENABLED:
                try:
                    from sims_fetcher import _get_token, SIMS_BASE_URL
                    sims_host = SIMS_BASE_URL.replace("http://", "").replace("https://", "").split("/")[0]
                    if sims_host in url or "simscloud" in url or "cnhtcerp" in url:
                        headers["Authorization"] = _get_token()
                        headers["Referer"]       = SIMS_BASE_URL + "/"
                        headers["Origin"]        = SIMS_BASE_URL
                        headers["language"]      = "en"
                except Exception as e:
                    print(f"[debug] Gagal ambil token SIMS: {e}")

            resp = requests.get(url, timeout=15, headers=headers)
            if resp.status_code == 200:
                content_type = resp.headers.get("Content-Type", "")
                if any(t in content_type for t in ("image", "octet-stream", "jpeg", "png", "gif", "webp")):
                    return resp.content, None
                if len(resp.content) > 1000:
                    return resp.content, None
                return None, f"Konten bukan gambar (Content-Type: {content_type})"
            return None, f"HTTP {resp.status_code}"
        except requests.exceptions.ConnectionError:
            return None, "Tidak dapat terhubung ke server"
        except requests.exceptions.Timeout:
            return None, "Timeout: server tidak merespons"
        except Exception as e:
            return None, str(e)

    def _load_stok_data(self):
        if self.stok_cache is not None:
            return
        if "stok_data" in st.session_state:
            self.stok_cache        = st.session_state.stok_data
            self.stok_gudang_cache = st.session_state.get("stok_gudang_data", {})
            self.gudang_names      = st.session_state.get("stok_gudang_names", [])
            return

        # ── Coba download dari Supabase Storage ──────────────────────────
        file_bytes = None
        try:
            from admin_data_uploader import download_dataset
            file_bytes = download_dataset("stok")
            if file_bytes:
                print("[stok] ✅ stok.xlsx diunduh dari Supabase Storage.")
        except Exception as e:
            print(f"[stok] ⚠️ Gagal download dari Supabase: {e}")

        # ── Fallback ke file lokal ────────────────────────────────────────
        if not file_bytes:
            if self.stok_file.exists():
                try:
                    file_bytes = self.stok_file.read_bytes()
                    print("[stok] ℹ️ stok.xlsx dibaca dari file lokal.")
                except Exception as e:
                    print(f"[stok] ❌ Gagal baca lokal: {e}")

        if not file_bytes:
            self._set_stok_cache({}, {}, [])
            return

        # ── Parse Excel (auto-deteksi format lama / multi-gudang) ─────────
        try:
            stok_cache, gudang_cache, gudang_names = parse_stok_file(file_bytes)
            self._set_stok_cache(stok_cache, gudang_cache, gudang_names)
            if gudang_names:
                print(f"[stok] ✅ Format multi-gudang: {len(stok_cache)} PN, "
                      f"{len(gudang_names)} gudang.")
            else:
                print(f"[stok] ✅ Format total: {len(stok_cache)} PN.")
        except Exception as e:
            st.error(f"Gagal membaca stok.xlsx → {e}")
            self._set_stok_cache({}, {}, [])

    def _set_stok_cache(self, stok_cache, gudang_cache, gudang_names):
        """Simpan hasil parse stok ke instance + session_state."""
        self.stok_cache        = stok_cache
        self.stok_gudang_cache = gudang_cache
        self.gudang_names      = gudang_names
        st.session_state.stok_data         = stok_cache
        st.session_state.stok_gudang_data  = gudang_cache
        st.session_state.stok_gudang_names = gudang_names


    def _load_harga_data(self):
        if self.harga_cache is not None:
            return
        if "harga_data" in st.session_state:
            self.harga_cache  = st.session_state.harga_data
            self.harga_lookup = st.session_state.get("harga_lookup", {})
            return

        # ── Coba dari Supabase Storage dulu ───────────────────────────────
        file_bytes = None
        try:
            from admin_data_uploader import download_dataset
            file_bytes = download_dataset("harga")
            if file_bytes:
                print("[harga] ✅ harga.xlsx diunduh dari Supabase Storage.")
        except Exception as e:
            print(f"[harga] ⚠️ Gagal download dari Supabase: {e}")

        # ── Fallback ke file lokal ────────────────────────────────────────
        if not file_bytes and self.harga_file.exists():
            try:
                file_bytes = self.harga_file.read_bytes()
                print("[harga] ℹ️ harga.xlsx dibaca dari file lokal.")
            except Exception as e:
                print(f"[harga] ❌ Gagal baca lokal: {e}")

        if not file_bytes:
            self.harga_cache  = pd.DataFrame(columns=["Part Number", "Part Name", "Harga"])
            self.harga_lookup = {}
            st.session_state.harga_data   = self.harga_cache
            st.session_state.harga_lookup = self.harga_lookup
            return

        try:
            # Filter usecols supaya hanya kolom relevan (PN/Name/Harga) yang
            # di-load ke RAM. harga.xlsx user biasanya punya banyak kolom
            # tambahan (kategori, supplier, dll) yang tidak dipakai.
            def _is_useful_col(name) -> bool:
                cl = str(name).strip().lower()
                return any(kw in cl for kw in (
                    "part number", "partnumber", "no part", "kode",
                    "part name", "nama", "deskripsi",
                    "harga", "price",
                ))
            df_h = pd.read_excel(
                io.BytesIO(file_bytes), dtype=str, usecols=_is_useful_col
            )
            df_h.columns = [c.strip() for c in df_h.columns]
            col_map = {}
            for c in df_h.columns:
                cl = c.lower()
                if "part number" in cl or "partnumber" in cl or "no part" in cl or "kode" in cl:
                    col_map[c] = "Part Number"
                elif "part name" in cl or "nama" in cl or "deskripsi" in cl:
                    col_map[c] = "Part Name"
                elif "harga" in cl or "price" in cl:
                    col_map[c] = "Harga"
            df_h = df_h.rename(columns=col_map)
            for req in ("Part Number", "Part Name", "Harga"):
                if req not in df_h.columns:
                    df_h[req] = ""
            df_h["Part Number"] = df_h["Part Number"].astype(str).str.strip().str.upper()
            df_h = df_h.dropna(subset=["Part Number"])
            df_h = df_h[df_h["Part Number"] != ""]
            self.harga_cache = df_h.reset_index(drop=True)
            # Bangun lookup dict sekali di sini
            lookup = {}
            for pn_key, harga_val in zip(df_h["Part Number"], df_h["Harga"]):
                try:
                    num = float(str(harga_val).replace(",", "").strip())
                    lookup[pn_key] = f"Rp {num:,.0f}"
                except Exception:
                    lookup[pn_key] = str(harga_val) if pd.notna(harga_val) else "—"
            self.harga_lookup = lookup
            st.session_state.harga_data   = self.harga_cache
            st.session_state.harga_lookup = self.harga_lookup
        except Exception as e:
            st.error(f"Gagal membaca harga.xlsx → {e}")
            self.harga_cache  = pd.DataFrame(columns=["Part Number", "Part Name", "Harga"])
            self.harga_lookup = {}
            st.session_state.harga_data   = self.harga_cache
            st.session_state.harga_lookup = self.harga_lookup

    def render_harga_tab(self):
        st.markdown("### 💰 Daftar Harga Sparepart")

        # Cek izin akses kolom harga
        user_h = LoginManager.get_current_user()
        role_h = user_h["role"] if user_h else "user"
        if "col_harga" not in get_allowed_columns(user_h["username"], role_h):
            st.warning("🔒 Anda tidak memiliki akses untuk melihat data harga. Hubungi admin jika diperlukan.")
            return

        col_reload, _ = st.columns([1, 5])
        with col_reload:
            if st.button("🔄 Reload Harga", key="reload_harga"):
                st.session_state.pop("harga_data", None)
                self.harga_cache = None
                self._load_harga_data()
                st.rerun()

        df_h = self.harga_cache if self.harga_cache is not None else pd.DataFrame()

        if df_h.empty:
            st.warning(
                "File harga tidak ditemukan atau kosong. "
                "Pastikan file ada di `data/harga/harga.xlsx`."
            )
            return

        st.divider()

        with st.expander("🔍 Filter & Pencarian", expanded=True):
            col_s, col_sort = st.columns([3, 2])
            with col_s:
                kw_harga = st.text_input(
                    "Cari Part Number / Part Name:",
                    placeholder="Ketik untuk filter…",
                    key="harga_search_kw",
                )
            with col_sort:
                sort_by = st.selectbox(
                    "Urutkan berdasarkan:",
                    ["Part Number", "Part Name", "Harga (Terendah)", "Harga (Tertinggi)"],
                    key="harga_sort",
                )

        # df_show: filter dulu (return view/new), lalu sort. Tanpa .copy()
        # awal — df_h tidak di-mutate karena semua langkah berikut return
        # DataFrame baru (boolean filter, sort_values, dst).
        if kw_harga.strip():
            kw_up = kw_harga.strip().upper()
            mask = (
                df_h["Part Number"].str.upper().str.contains(kw_up, na=False) |
                df_h["Part Name"].astype(str).str.upper().str.contains(kw_up, na=False)
            )
            df_show = df_h[mask].reset_index(drop=True)
        else:
            df_show = df_h

        def _harga_sort_key(s):
            return pd.to_numeric(
                s.astype(str).str.replace(r"[^\d.]", "", regex=True),
                errors="coerce",
            )

        try:
            if sort_by == "Harga (Terendah)":
                df_show = df_show.sort_values(
                    "Harga", key=_harga_sort_key, ascending=True
                ).reset_index(drop=True)
            elif sort_by == "Harga (Tertinggi)":
                df_show = df_show.sort_values(
                    "Harga", key=_harga_sort_key, ascending=False
                ).reset_index(drop=True)
            elif sort_by == "Part Name":
                df_show = df_show.sort_values(
                    "Part Name", key=lambda x: x.str.upper()
                ).reset_index(drop=True)
            else:
                df_show = df_show.sort_values("Part Number").reset_index(drop=True)
        except Exception:
            pass

        c1, c2 = st.columns(2)
        c1.metric("Total Part", len(df_h))
        c2.metric("Hasil Filter", len(df_show))

        st.markdown("---")

        def fmt_harga(val):
            try:
                num = float(str(val).replace(",", "").strip())
                return f"Rp {num:,.0f}"
            except Exception:
                return str(val)

        # Construct df_display fresh — no .copy() needed.
        df_display = pd.DataFrame({
            "Part Number": df_show["Part Number"].to_numpy(),
            "Part Name":   df_show["Part Name"].to_numpy(),
            "Harga (Rp)":  df_show["Harga"].map(fmt_harga).to_numpy(),
        })

        if df_display.empty:
            st.info("Tidak ada data yang cocok dengan pencarian.")
        else:
            st.dataframe(
                df_display,
                hide_index=True,
                use_container_width=True,
                height=min(400, 56 + len(df_display) * 35),
                column_config={
                    "Part Number": st.column_config.TextColumn("Part Number", width="medium"),
                    "Part Name":   st.column_config.TextColumn("Part Name",   width="large"),
                    "Harga (Rp)":  st.column_config.TextColumn("Harga",       width="medium"),
                },
            )

            dl_buf = io.BytesIO()
            df_display.to_excel(dl_buf, index=False, engine="openpyxl")
            dl_buf.seek(0)
            st.download_button(
                label="⬇️ Download Excel",
                data=dl_buf.getvalue(),
                file_name=f"harga_sparepart_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="harga_download",
            )


    def process_single_file(self, file_path, relative_path):
        results     = []
        file_name   = file_path.name
        simple_name = self.extract_simple_filename(file_name)
        file_hash   = self.get_file_hash(file_path)
        if file_hash:
            cached = self.load_file_cache(file_path, file_hash)
            if cached:
                return cached
        try:
            xls = pd.ExcelFile(file_path, engine="openpyxl")
            for sheet_name in xls.sheet_names:
                try:
                    df = pd.read_excel(xls, sheet_name=sheet_name, usecols=[1,3,4], dtype=str)
                    df.columns = ["part_number","part_name","quantity"]

                    # Pre-normalize vectorized — 5–10× lebih cepat dari
                    # iterrows + per-row str() pada 920k baris total.
                    pn_series = (
                        df["part_number"].fillna("").astype(str)
                        .str.strip().str.upper()
                    )
                    nm_series = (
                        df["part_name"].fillna("").astype(str)
                        .str.strip().str.upper()
                    )

                    # pn_idx: dict[PN_upper, list[row_idx]]. Pakai groupby —
                    # preserve order index per group dengan sort=False.
                    pn_valid = pn_series[pn_series != ""]
                    if len(pn_valid):
                        pn_idx = (
                            pn_valid.reset_index()
                            .groupby("part_number", sort=False)["index"]
                            .apply(list)
                            .to_dict()
                        )
                    else:
                        pn_idx = {}

                    # nm_idx: per-row split tetap perlu loop (multi-key per row).
                    # Pakai .items() pada Series (~itertuples speed) — masih
                    # jauh lebih cepat dari iterrows + repeat str/strip/upper.
                    nm_idx = {}
                    for idx, nm in nm_series.items():
                        if not nm:
                            continue
                        for word in nm.split():
                            if len(word) > 2:
                                nm_idx.setdefault(word, []).append(idx)
                    results.append({
                        "full_path": str(file_path), "file_name": file_name,
                        "relative_path": str(relative_path), "simple_name": simple_name,
                        "sheet": sheet_name, "dataframe": df, "row_count": len(df),
                        "col_count": len(df.columns), "part_number_index": pn_idx,
                        "part_name_index": nm_idx,
                        "last_modified": datetime.fromtimestamp(file_path.stat().st_mtime),
                    })
                except Exception:
                    continue
        except Exception:
            pass
        if file_hash and results:
            self.save_file_cache(file_path, file_hash, results)
        return results

    def auto_load_excel_files(self):
        """
        Force-reload Excel index. Dipakai oleh tombol "🔄 Refresh Data" di
        sidebar — clear shared cache lalu re-walk + re-parse folder data.
        Path normal (init session baru) TIDAK lewat sini; langsung pakai
        `_load_excel_index_shared(...)` di __init__.
        """
        try:
            self.create_data_folder()
            # Bust cache_resource supaya next call benar-benar walk lagi
            _load_excel_index_shared.clear()
            results, n_total_files, idx_time = _load_excel_index_shared(
                str(self.data_folder), self.process_single_file
            )
            st.session_state.excel_files        = results
            st.session_state.last_index_time    = idx_time
            st.session_state.loaded_files_count = len(results)
            st.session_state.last_file_count    = n_total_files
            # Cleanup proaktif pickle cache (LRU) — handle orphan dari
            # file Excel yang sudah dihapus.
            _enforce_pickle_cache_size_limit()
        except Exception as e:
            st.error(f"Error loading Excel files: {e}")

    # ── Tab: Search Part Number ──────────────────────────────────────
    def _render_tab_search_pn(self):
        st.markdown(
            '<div style="margin: .25rem 0 .75rem;">'
            '<div style="font-size:15px;font-weight:700;letter-spacing:-.01em;color:var(--mp-ink);">'
            'Cari <span style="color:var(--mp-green);">Part Number</span></div>'
            '<div style="font-size:12px;color:var(--mp-ink-50);margin-top:2px;">'
            'Pencocokan exact pada kolom B (Part Number) \u00b7 fallback otomatis ke SIMS jika tidak ditemukan.'
            '</div></div>',
            unsafe_allow_html=True,
        )
        with st.form(key="search_pn_form", clear_on_submit=False):
            col_input, col_btn = st.columns([4, 1])
            with col_input:
                sn_input = st.text_input(
                    "Part Number",
                    placeholder="Contoh: WG1642821034/1",
                    key="sn_input",
                    label_visibility="collapsed",
                )
            with col_btn:
                go = st.form_submit_button("\U0001f50d Cari", type="primary", use_container_width=True)
            if go:
                if sn_input:
                    with st.spinner("Mencari\u2026"):
                        st.session_state.search_results = search_part_number(
                            sn_input, st.session_state.excel_files, self.stok_cache, self.harga_lookup)
                        st.session_state.search_type = "Part Number"
                        st.session_state.search_term = sn_input
                        _u = LoginManager.get_current_user() or {}
                        log_activity(_u.get("username", ""), "search_pn",
                                     target=sn_input,
                                     details={"results": len(st.session_state.search_results or [])})
                        st.rerun()
                else:
                    st.warning("Masukkan part number untuk mencari.")
        # Empty state — sebelum user search apapun
        if not st.session_state.get("search_results") and not st.session_state.get("search_term"):
            self._render_search_empty_state(kind="pn")

    # ── Tab: Search Part Name ────────────────────────────────────────
    def _render_tab_search_name(self):
        st.markdown(
            '<div style="margin: .25rem 0 .75rem;">'
            '<div style="font-size:15px;font-weight:700;letter-spacing:-.01em;color:var(--mp-ink);">'
            'Cari <span style="color:var(--mp-green);">Part Name</span></div>'
            '<div style="font-size:12px;color:var(--mp-ink-50);margin-top:2px;">'
            'Pencarian substring pada kolom D (Part Name) - case-insensitive.'
            '</div></div>',
            unsafe_allow_html=True,
        )
        with st.form(key="search_name_form", clear_on_submit=False):
            col_input, col_btn = st.columns([4, 1])
            with col_input:
                name_input = st.text_input(
                    "Part Name",
                    placeholder="Contoh: baut roda, bearing, kampas rem",
                    key="name_input",
                    label_visibility="collapsed",
                )
            with col_btn:
                go = st.form_submit_button("\U0001f50d Cari", type="primary", use_container_width=True)
            if go:
                if name_input:
                    with st.spinner("Mencari\u2026"):
                        st.session_state.search_results = search_part_name(
                            name_input, st.session_state.excel_files, self.stok_cache, self.harga_lookup)
                        st.session_state.search_type = "Part Name"
                        st.session_state.search_term = name_input
                        _u = LoginManager.get_current_user() or {}
                        log_activity(_u.get("username", ""), "search_name",
                                     target=name_input,
                                     details={"results": len(st.session_state.search_results or [])})
                        st.rerun()
                else:
                    st.warning("Masukkan nama part untuk mencari.")
        # Empty state — sebelum user search apapun
        if not st.session_state.get("search_results") and not st.session_state.get("search_term"):
            self._render_search_empty_state(kind="name")

    # ── Empty state untuk Search PN / Search Name ───────────────────
    def _render_search_empty_state(self, kind: str = "pn"):
        n_files = st.session_state.get("loaded_files_count", 0) or 0
        idx_time = st.session_state.get("last_index_time")
        idx_str  = idx_time.strftime("%H:%M:%S") if idx_time else "—"
        excel_files = st.session_state.get("excel_files", []) or []
        # Estimasi part: jumlah row dari semua dataframe (jika tersedia)
        try:
            est_parts = sum(len(fi.get("dataframe", [])) for fi in excel_files)
        except Exception:
            est_parts = 0
        try:
            brands = {fi.get("relative_path", "").split("\\")[0].split("/")[0]
                      for fi in excel_files if fi.get("relative_path")}
            n_brands = len([b for b in brands if b])
        except Exception:
            n_brands = 0

        user = LoginManager.get_current_user() or {}
        login_str = user["login_time"].strftime("%H:%M") if user.get("login_time") else "—"

        # ── 4 KPI tiles ──────────────────────────────────────────────
        st.markdown(
            f"""
<div style="display:grid;grid-template-columns:repeat(4, 1fr);gap:10px;margin:1rem 0 .8rem;">
  <div class="mp-card" style="padding:14px 16px;">
    <div style="font-size:11px;font-weight:600;color:var(--mp-ink-50);letter-spacing:.04em;text-transform:uppercase;">File Excel</div>
    <div style="font-size:24px;font-weight:700;color:var(--mp-ink);letter-spacing:-.02em;line-height:1.1;margin-top:6px;">{n_files}</div>
    <div style="font-size:11px;color:var(--mp-ink-50);margin-top:4px;">terindex &amp; siap dicari</div>
  </div>
  <div class="mp-card" style="padding:14px 16px;">
    <div style="font-size:11px;font-weight:600;color:var(--mp-ink-50);letter-spacing:.04em;text-transform:uppercase;">Estimasi Part</div>
    <div style="font-size:24px;font-weight:700;color:var(--mp-green);letter-spacing:-.02em;line-height:1.1;margin-top:6px;">{est_parts:,}</div>
    <div style="font-size:11px;color:var(--mp-ink-50);margin-top:4px;">baris di semua sheet</div>
  </div>
  <div class="mp-card" style="padding:14px 16px;">
    <div style="font-size:11px;font-weight:600;color:var(--mp-ink-50);letter-spacing:.04em;text-transform:uppercase;">Brand &amp; Unit</div>
    <div style="font-size:24px;font-weight:700;color:var(--mp-ink);letter-spacing:-.02em;line-height:1.1;margin-top:6px;">{n_brands or '—'}</div>
    <div style="font-size:11px;color:var(--mp-ink-50);margin-top:4px;">folder data terdeteksi</div>
  </div>
  <div class="mp-card" style="padding:14px 16px;">
    <div style="font-size:11px;font-weight:600;color:var(--mp-ink-50);letter-spacing:.04em;text-transform:uppercase;">Status Index</div>
    <div style="font-size:24px;font-weight:700;color:var(--mp-green);letter-spacing:-.02em;line-height:1.1;margin-top:6px;">● Live</div>
    <div style="font-size:11px;color:var(--mp-ink-50);margin-top:4px;">diperbarui {idx_str}</div>
  </div>
</div>
""",
            unsafe_allow_html=True,
        )

        # ── 2-col: Tips + Quick examples / Other features ───────────
        col_tips, col_actions = st.columns([1, 1])
        with col_tips:
            tip_title = ("Tips Pencarian Part Number" if kind == "pn"
                         else "Tips Pencarian Part Name")
            tip_items = (
                [
                    "Ketik <b>part number lengkap</b> (mis. <code>WG1642821034</code>) untuk match exact.",
                    "Cocok juga untuk format dengan slash: <code>WG1642821034/1</code>.",
                    "Jika tidak ditemukan di Excel lokal, sistem auto-fallback ke <b>SIMS</b>.",
                    "Pakai tab <b>Cari by Foto</b> kalau hanya punya gambar part.",
                ]
                if kind == "pn"
                else [
                    "Ketik <b>kata kunci</b> (mis. <code>baut roda</code>, <code>bearing</code>).",
                    "Pencarian <b>case-insensitive</b> &amp; <b>substring</b> di kolom D.",
                    "Hasil banyak? Gunakan kata kunci lebih spesifik (mis. <code>kampas rem depan</code>).",
                    "Untuk pencarian akurat, gunakan <b>Search Part Number</b>.",
                ]
            )
            tips_html = "".join(
                f'<li style="margin:8px 0;color:var(--mp-ink-70);line-height:1.5;">{t}</li>'
                for t in tip_items
            )
            st.markdown(
                f"""
<div class="mp-card" style="padding:16px 18px;">
  <div style="display:flex;align-items:center;gap:8px;margin-bottom:6px;">
    <div style="width:24px;height:24px;border-radius:6px;background:var(--mp-green-soft);
                color:var(--mp-green-dark);display:flex;align-items:center;justify-content:center;
                font-size:13px;flex:none;">💡</div>
    <div style="font-size:13.5px;font-weight:700;letter-spacing:-.01em;">{tip_title}</div>
  </div>
  <ul style="margin:.5rem 0 0;padding-left:1.1rem;font-size:12.5px;">{tips_html}</ul>
</div>
""",
                unsafe_allow_html=True,
            )

        with col_actions:
            example_items = (
                [("WG1642821034", "Clutch Master Cylinder"),
                 ("WG9925520270", "Brake Chamber T24"),
                 ("AZ9100443082", "Air Dryer Cartridge")]
                if kind == "pn"
                else [("baut roda", "Cari semua part nama 'baut roda'"),
                      ("bearing",   "Cari semua bearing di database"),
                      ("kampas rem","Cari kampas rem (depan/belakang)")]
            )
            rows_html = "".join(
                f'<div style="display:flex;justify-content:space-between;align-items:center;'
                f'padding:10px 12px;border:1px solid var(--mp-line);border-radius:8px;'
                f'background:var(--mp-bg);margin-bottom:6px;">'
                f'<span style="font-family:var(--mp-font-mono);font-weight:600;font-size:12.5px;color:var(--mp-ink);">{val}</span>'
                f'<span style="font-size:11.5px;color:var(--mp-ink-50);">{desc}</span>'
                f'</div>'
                for val, desc in example_items
            )
            title_lbl = "Contoh Part Number" if kind == "pn" else "Contoh Kata Kunci"
            st.markdown(
                f"""
<div class="mp-card" style="padding:16px 18px;">
  <div style="display:flex;align-items:center;gap:8px;margin-bottom:10px;">
    <div style="width:24px;height:24px;border-radius:6px;background:var(--mp-green-soft);
                color:var(--mp-green-dark);display:flex;align-items:center;justify-content:center;
                font-size:13px;flex:none;">✨</div>
    <div style="font-size:13.5px;font-weight:700;letter-spacing:-.01em;">{title_lbl}</div>
  </div>
  {rows_html}
  <div style="font-size:11px;color:var(--mp-ink-50);margin-top:8px;">
    Salin salah satu di atas ke kotak pencarian lalu klik <b>Cari</b>.
  </div>
</div>
""",
                unsafe_allow_html=True,
            )

        # ── Info banner di bawah ────────────────────────────────────
        login_user = user.get("username", "—").title() if user else "—"
        st.markdown(
            f'<div class="mp-banner info" style="margin-top:12px;">'
            f'<div style="width:18px;flex:none;">👋</div>'
            f'<div>Halo <b>{login_user}</b>, mulai dengan mengetik di kotak '
            f'pencarian atau gunakan tab lain di atas. Sesi login Anda dimulai pukul '
            f'<b>{login_str}</b>.</div></div>',
            unsafe_allow_html=True,
        )

    # ── Tab: Bandingkan 2 Part (Interchange Analyzer) ────────────────
    def _render_tab_compare_parts(self):
        try:
            import part_compare as _pc
        except Exception as e:
            st.error(f"Modul perbandingan tidak tersedia: {e}")
            return

        st.markdown(
            '<div style="margin:.25rem 0 .75rem;">'
            '<div style="font-size:18px;font-weight:700;letter-spacing:-.015em;color:var(--mp-ink);">'
            'Bandingkan <span style="color:var(--mp-green);">2 Part</span></div>'
            '<div style="font-size:12.5px;color:var(--mp-ink-50);margin-top:4px;">'
            'Cek interchange — kemiripan <b>BENTUK</b> (utama), <b>NAMA</b> (penguat), '
            'dan <b>WARNA</b> (info) dari foto SIMS.'
            '</div></div>',
            unsafe_allow_html=True,
        )

        if not SIMS_ENABLED:
            st.warning("⚠️ SIMS Fetcher tidak aktif — fitur ini membutuhkan akses SIMS.")
            return

        with st.form(key="compare_parts_form", clear_on_submit=False):
            c1, c2 = st.columns(2)
            with c1:
                pn1_input = st.text_input(
                    "Part Number #1:",
                    placeholder="Contoh: WG1642821034",
                    key="compare_pn1_input",
                )
            with c2:
                pn2_input = st.text_input(
                    "Part Number #2:",
                    placeholder="Contoh: WG1642821035",
                    key="compare_pn2_input",
                )
            submitted = st.form_submit_button(
                "🔬 Cek Interchange",
                type="primary",
                use_container_width=True,
            )

        if submitted:
            pn1 = (pn1_input or "").strip()
            pn2 = (pn2_input or "").strip()
            if not pn1 or not pn2:
                st.warning("Mohon isi kedua Part Number.")
                return
            if pn1.upper() == pn2.upper():
                st.warning("Part Number tidak boleh sama.")
                return

            # ── Ambil daftar URL gambar + part info dari SIMS ─────────
            with st.spinner(f"🔍 Ambil data SIMS untuk {pn1} & {pn2}..."):
                urls1, err1 = _sims_fetch(pn1)
                urls2, err2 = _sims_fetch(pn2)
                name1 = ""
                name2 = ""
                try:
                    from sims_fetcher import get_sims_part_info
                    info1, _ = get_sims_part_info(pn1)
                    info2, _ = get_sims_part_info(pn2)
                    name1 = (info1 or {}).get("partName", "") or ""
                    name2 = (info2 or {}).get("partName", "") or ""
                except Exception as e:
                    print(f"[compare] partName fetch gagal: {e}")

            if err1:
                st.error(f"❌ {pn1}: {err1}")
            if err2:
                st.error(f"❌ {pn2}: {err2}")
            if not urls1:
                st.error(f"❌ Tidak ada gambar SIMS untuk {pn1}.")
                return
            if not urls2:
                st.error(f"❌ Tidak ada gambar SIMS untuk {pn2}.")
                return

            # ── Download bytes setiap gambar (parallel) ───────────────
            with st.spinner("⬇️ Mengunduh gambar..."):
                with ThreadPoolExecutor(max_workers=6) as ex:
                    futs1 = {ex.submit(ExcelSearchApp.fetch_image_bytes, u): i for i, u in enumerate(urls1)}
                    futs2 = {ex.submit(ExcelSearchApp.fetch_image_bytes, u): i for i, u in enumerate(urls2)}
                    bytes1 = [None] * len(urls1)
                    bytes2 = [None] * len(urls2)
                    for f in as_completed(futs1):
                        i = futs1[f]
                        try:
                            b, _ = f.result()
                            bytes1[i] = b
                        except Exception:
                            bytes1[i] = None
                    for f in as_completed(futs2):
                        i = futs2[f]
                        try:
                            b, _ = f.result()
                            bytes2[i] = b
                        except Exception:
                            bytes2[i] = None

            valid1 = [b for b in bytes1 if b]
            valid2 = [b for b in bytes2 if b]
            if not valid1 or not valid2:
                st.error("❌ Gagal mengunduh konten gambar dari SIMS.")
                return

            # ── Analisis kemiripan ────────────────────────────────────
            with st.spinner("🧠 Menganalisis kemiripan (bentuk + nama + warna)..."):
                result = _pc.best_match(bytes1, bytes2, name1=name1, name2=name2)

            best  = result.get("best")
            pairs = result.get("pairs") or []
            if not best:
                st.error("❌ Tidak dapat menganalisis pasangan gambar manapun.")
                return

            st.session_state["_compare_result"] = {
                "pn1": pn1, "pn2": pn2,
                "name1": name1, "name2": name2,
                "urls1": urls1, "urls2": urls2,
                "bytes1": bytes1, "bytes2": bytes2,
                "best": best, "pairs": pairs,
            }

        # ── Render hasil (jika ada) ──────────────────────────────────
        cmp = st.session_state.get("_compare_result")
        if not cmp:
            return

        pn1, pn2   = cmp["pn1"], cmp["pn2"]
        name1      = cmp.get("name1", "") or ""
        name2      = cmp.get("name2", "") or ""
        bytes1     = cmp["bytes1"]
        bytes2     = cmp["bytes2"]
        urls1      = cmp["urls1"]
        urls2      = cmp["urls2"]
        best       = cmp["best"]
        pairs      = cmp["pairs"]

        st.divider()

        # ── Verdict utama (interchange) — hero card style sesuai mock ──
        verdict   = best["verdict"]
        vcolor    = best["color"]                       # color dari engine (hex)
        shape     = best["shape_score"]
        color_s   = best["color_score"]
        name_s    = best.get("name_score")
        overall   = (
            (shape * 0.6 + (name_s or 0) * 0.25 + color_s * 0.15)
            if name_s is not None else (shape * 0.7 + color_s * 0.3)
        ) * 100.0

        # Hero verdict — gradient soft + big icon + headline besar
        verdict_strong = shape >= 0.75 or (name_s is not None and name_s >= 0.85)
        ICON = "✓" if verdict_strong else "?"
        st.markdown(
            f"""
<div style="border:1px solid var(--mp-green-line);border-left:6px solid {vcolor};
            background:linear-gradient(90deg, var(--mp-green-soft) 0%, var(--mp-green-softer) 100%);
            border-radius:var(--mp-radius);padding:18px;margin:8px 0 16px 0;">
  <div style="display:flex;align-items:center;gap:16px;">
    <div style="width:56px;height:56px;border-radius:14px;background:{vcolor};
                color:white;display:flex;align-items:center;justify-content:center;
                font-size:30px;font-weight:800;flex:none;
                box-shadow:0 4px 14px {vcolor}55;">{ICON}</div>
    <div style="flex:1;min-width:0;">
      <div style="font-size:11px;font-weight:700;color:var(--mp-green-dark);
                  letter-spacing:.08em;text-transform:uppercase;">Hasil Analisis</div>
      <div style="font-size:24px;font-weight:800;color:{vcolor};
                  letter-spacing:-.01em;line-height:1.15;margin-top:2px;">{verdict}</div>
      <div style="font-size:12.5px;color:var(--mp-ink-70);margin-top:6px;">
        Pasangan foto terbaik: gambar #{best['i']+1} ({pn1}) vs #{best['j']+1} ({pn2}) ·
        overall <b style="color:var(--mp-green-dark);">{overall:.1f}%</b>
      </div>
    </div>
  </div>
</div>
""",
            unsafe_allow_html=True,
        )

        # ── 3 signal cards: Bentuk / Nama / Warna ────────────────────
        def _signal_color(v):
            if v is None:        return "var(--mp-ink-30)"
            if v >= 0.75:        return "var(--mp-green)"
            if v >= 0.55:        return "var(--mp-amber)"
            return "var(--mp-red)"

        def _signal_card(emoji, title, badge_label, value, desc, accent):
            value_str = f"{value*100:.1f}" if value is not None else "—"
            pct_w = max(0, min(100, (value or 0) * 100)) if value is not None else 0
            sub_unit = "<span style='font-size:18px;'>%</span>" if value is not None else ""
            return (
                f'<div class="mp-card" style="padding:14px;">'
                f'<div style="display:flex;justify-content:space-between;align-items:center;">'
                f'<span style="font-size:11px;font-weight:700;color:var(--mp-ink-50);'
                f'letter-spacing:.08em;text-transform:uppercase;">{emoji} {title}</span>'
                f'<span class="mp-chip gray" style="font-size:10px;">{badge_label}</span>'
                f'</div>'
                f'<div style="font-size:30px;font-weight:800;color:{accent};'
                f'letter-spacing:-.02em;margin-top:8px;line-height:1.1;">{value_str}{sub_unit}</div>'
                f'<div class="stProgress" style="margin-top:10px;"><div style="height:6px;'
                f'background:var(--mp-line-soft);border-radius:999px;overflow:hidden;">'
                f'<div style="width:{pct_w}%;height:100%;background:{accent};"></div></div></div>'
                f'<div style="font-size:11px;color:var(--mp-ink-50);margin-top:8px;">{desc}</div>'
                f'</div>'
            )

        s1, s2, s3 = st.columns(3)
        with s1:
            st.markdown(_signal_card("🔧", "BENTUK", "utama", shape,
                                     "pHash + dHash + SSIM + edge + aspect",
                                     _signal_color(shape)),
                        unsafe_allow_html=True)
        with s2:
            if name_s is None:
                st.markdown(_signal_card("📝", "NAMA PART", "tidak ada", None,
                                         "partName SIMS tidak tersedia",
                                         "var(--mp-ink-30)"),
                            unsafe_allow_html=True)
            else:
                st.markdown(_signal_card("📝", "NAMA PART", "penguat", name_s,
                                         "SequenceMatcher + token Jaccard",
                                         _signal_color(name_s)),
                            unsafe_allow_html=True)
        with s3:
            st.markdown(_signal_card("🎨", "WARNA", "info", color_s,
                                     "histogram RGB + mean color",
                                     _signal_color(color_s)),
                        unsafe_allow_html=True)

        # ── Disclaimer banner ────────────────────────────────────────
        st.markdown(
            '<div class="mp-banner warn" style="margin:14px 0 8px;">'
            '<div style="width:18px;flex:none;">⚠</div>'
            '<div><b>Catatan:</b> Analisis berbasis foto SIMS — hanya indikator awal '
            'interchange. Verifikasi fisik (dimensi, threading, material) dan '
            'cross-reference dokumen OEM tetap diperlukan sebelum keputusan final.</div>'
            '</div>',
            unsafe_allow_html=True,
        )

        # ── PartName side-by-side ────────────────────────────────────
        if name1 or name2:
            n1c, n2c = st.columns(2)
            with n1c:
                st.markdown(
                    f"<div style='background:#F3F4F6;padding:8px 12px;border-radius:6px;'>"
                    f"<div style='font-size:.72rem;color:#6b7280;'>Part Name {pn1}</div>"
                    f"<div style='font-size:.95rem;font-weight:600;'>{(name1 or '—')}</div></div>",
                    unsafe_allow_html=True,
                )
            with n2c:
                st.markdown(
                    f"<div style='background:#F3F4F6;padding:8px 12px;border-radius:6px;'>"
                    f"<div style='font-size:.72rem;color:#6b7280;'>Part Name {pn2}</div>"
                    f"<div style='font-size:.95rem;font-weight:600;'>{(name2 or '—')}</div></div>",
                    unsafe_allow_html=True,
                )

        # ── Foto pasangan terbaik berdampingan ───────────────────────
        col_a, col_b = st.columns(2)
        with col_a:
            st.markdown(f"**🅰️ {pn1}** — gambar #{best['i']+1}")
            try:
                ExcelSearchApp.render_zoomable_image(
                    bytes1[best["i"]],
                    caption=f"{pn1} (gbr {best['i']+1}/{len(urls1)})",
                    zoom_key=f"cmp_a_{best['i']}",
                )
            except Exception as e:
                st.error(f"Gagal render gambar A: {e}")
        with col_b:
            st.markdown(f"**🅱️ {pn2}** — gambar #{best['j']+1}")
            try:
                ExcelSearchApp.render_zoomable_image(
                    bytes2[best["j"]],
                    caption=f"{pn2} (gbr {best['j']+1}/{len(urls2)})",
                    zoom_key=f"cmp_b_{best['j']}",
                )
            except Exception as e:
                st.error(f"Gagal render gambar B: {e}")

        # ── Semua gambar — versus berdampingan ───────────────────────
        st.markdown(
            '<div style="margin:18px 0 8px;">'
            '<div style="font-size:14px;font-weight:700;color:var(--mp-ink);">'
            '🖼️ Semua Gambar — <span style="color:var(--mp-green);">Versus</span></div>'
            '<div style="font-size:12px;color:var(--mp-ink-50);margin-top:2px;">'
            f'Bandingkan seluruh foto SIMS dari kedua part secara visual '
            f'(<b>{len([b for b in bytes1 if b])}</b> foto {pn1} vs '
            f'<b>{len([b for b in bytes2 if b])}</b> foto {pn2}). '
            'Pasangan terbaik diberi badge hijau.'
            '</div></div>',
            unsafe_allow_html=True,
        )

        def _render_all_images_column(pn, all_bytes, all_urls, best_idx, side_label):
            st.markdown(
                f"<div style='font-size:13px;font-weight:700;color:var(--mp-ink);"
                f"padding:6px 10px;background:var(--mp-green-softer);"
                f"border:1px solid var(--mp-green-line);border-radius:8px;"
                f"margin-bottom:8px;'>{side_label} <code>{pn}</code> "
                f"<span style='color:var(--mp-ink-50);font-weight:500;'>"
                f"· {len([b for b in all_bytes if b])} gambar</span></div>",
                unsafe_allow_html=True,
            )
            for idx, b in enumerate(all_bytes):
                is_best = (idx == best_idx)
                badge = (
                    "<span style='background:var(--mp-green);color:white;"
                    "font-size:10px;font-weight:700;padding:2px 8px;border-radius:999px;"
                    "letter-spacing:.05em;'>★ TERBAIK</span>"
                    if is_best else ""
                )
                st.markdown(
                    f"<div style='display:flex;justify-content:space-between;"
                    f"align-items:center;margin:10px 0 4px;'>"
                    f"<span style='font-size:12px;font-weight:600;color:var(--mp-ink-70);'>"
                    f"Gambar #{idx+1}/{len(all_bytes)}</span>{badge}</div>",
                    unsafe_allow_html=True,
                )
                if not b:
                    st.markdown(
                        "<div style='padding:24px;text-align:center;"
                        "background:#FEF2F2;color:#991B1B;border:1px dashed #FCA5A5;"
                        "border-radius:8px;font-size:12px;'>"
                        "❌ Gagal mengunduh gambar</div>",
                        unsafe_allow_html=True,
                    )
                    continue
                try:
                    st.image(b, use_container_width=True)
                except Exception as e:
                    st.error(f"Gagal render gambar #{idx+1}: {e}")

        all_col_a, all_col_b = st.columns(2)
        with all_col_a:
            _render_all_images_column(pn1, bytes1, urls1, best["i"], "🅰️")
        with all_col_b:
            _render_all_images_column(pn2, bytes2, urls2, best["j"], "🅱️")

        # ── Detail sub-metrik shape & color ──────────────────────────
        with st.expander("📊 Detail Sub-Metrik (bentuk & warna)"):
            metrics = best["metrics"]
            labels  = _pc.METRIC_LABELS

            shape_keys = list(_pc.SHAPE_WEIGHTS.keys())
            color_keys = list(_pc.COLOR_WEIGHTS.keys())

            st.markdown("**🔧 Sub-metrik BENTUK**")
            df_shape = pd.DataFrame([
                {
                    "Metrik":     labels.get(k, k),
                    "Skor":       f"{metrics[k]*100:.1f}%",
                    "Bobot":      f"{_pc.SHAPE_WEIGHTS[k]*100:.0f}%",
                    "Kontribusi": f"{metrics[k]*_pc.SHAPE_WEIGHTS[k]*100:.1f}%",
                } for k in shape_keys
            ])
            st.dataframe(df_shape, hide_index=True, use_container_width=True)

            for k in shape_keys:
                v = metrics[k]
                pct = max(0.0, min(1.0, float(v))) * 100
                bar = "#16A34A" if v >= 0.7 else ("#CA8A04" if v >= 0.5 else "#DC2626")
                st.markdown(
                    f"""
<div style="margin:4px 0;">
  <div style="display:flex;justify-content:space-between;font-size:.78rem;color:#444;">
    <span>{labels.get(k, k)}</span><span><b>{pct:.1f}%</b></span>
  </div>
  <div style="background:#eee;border-radius:6px;height:8px;overflow:hidden;">
    <div style="width:{pct:.1f}%;background:{bar};height:100%;"></div>
  </div>
</div>""",
                    unsafe_allow_html=True,
                )

            st.markdown("**🎨 Sub-metrik WARNA** (info saja, bobot kecil di overall)")
            df_color = pd.DataFrame([
                {
                    "Metrik":     labels.get(k, k),
                    "Skor":       f"{metrics[k]*100:.1f}%",
                    "Bobot":      f"{_pc.COLOR_WEIGHTS[k]*100:.0f}%",
                } for k in color_keys
            ])
            st.dataframe(df_color, hide_index=True, use_container_width=True)

        # ── Info teknis ──────────────────────────────────────────────
        extras = best.get("extras", {})
        with st.expander("🔬 Info Teknis"):
            st.markdown(
                f"- **Resolusi gambar #1:** {best.get('size1')}\n"
                f"- **Resolusi gambar #2:** {best.get('size2')}\n"
                f"- **Hamming pHash:** {extras.get('hamming_phash')}/64\n"
                f"- **Hamming dHash:** {extras.get('hamming_dhash')}/64\n"
                f"- **Hamming aHash:** {extras.get('hamming_ahash')}/64\n"
                f"- **Edge density #1:** {extras.get('edge1', 0):.4f}\n"
                f"- **Edge density #2:** {extras.get('edge2', 0):.4f}\n"
                f"- **Mean RGB #1:** {[round(x,1) for x in extras.get('mean_rgb1', [])]}\n"
                f"- **Mean RGB #2:** {[round(x,1) for x in extras.get('mean_rgb2', [])]}\n"
            )

        # ── Tabel semua kombinasi pasangan ───────────────────────────
        if len(pairs) > 1:
            with st.expander(f"📋 Semua Kombinasi Pasangan ({len(pairs)})"):
                df_pairs = pd.DataFrame([
                    {
                        f"Gbr {pn1}": p["i"] + 1,
                        f"Gbr {pn2}": p["j"] + 1,
                        "Bentuk":  f"{p['shape_score']*100:.1f}%",
                        "Warna":   f"{p['color_score']*100:.1f}%",
                        "Overall": f"{p['overall']*100:.1f}%",
                        "Verdict": p["verdict"],
                    }
                    for p in sorted(pairs, key=lambda x: -x["shape_score"])
                ])
                st.dataframe(df_pairs, hide_index=True, use_container_width=True)

        if st.button("🗑️ Bersihkan Hasil", key="compare_clear_btn"):
            st.session_state.pop("_compare_result", None)
            st.rerun()

    # ── Tab: Stok Opname (per user) ──────────────────────────────────
    def _render_tab_stok_opname(self):
        if not STOK_OPNAME_ENABLED or _so is None:
            st.error("Modul `stok_opname.py` tidak ditemukan.")
            return

        user = LoginManager.get_current_user()
        if not user:
            st.warning("Anda harus login untuk menggunakan fitur ini.")
            return
        username = user["username"]

        st.markdown(f"### 📋 Stok Opname — `{username}`")

        try:
            _be = _so.backend()
        except Exception:
            _be = "file"
        if _be == "supabase":
            _be_badge = "<span style='background:#DCFCE7;color:#166534;padding:2px 8px;border-radius:6px;font-size:.72rem;'>☁️ Supabase</span>"
        else:
            _be_badge = "<span style='background:#FEF3C7;color:#92400E;padding:2px 8px;border-radius:6px;font-size:.72rem;'>💾 File lokal</span>"

        st.markdown(
            f"<div style='color:#555;font-size:.9rem;margin-bottom:8px;'>"
            f"Sesi opname tersimpan <b>per user</b> di {_be_badge}. Setiap sesi dimulai dengan "
            f"<b>upload data stok awal sendiri</b> (Excel: PN + Qty Sistem). "
            f"Setelah itu Anda mengisi qty hasil hitung fisik secara <b>manual</b> "
            f"atau <b>upload Excel</b>. Setelah difinalisasi sesi masuk <b>Riwayat Opname</b>."
            f"</div>",
            unsafe_allow_html=True,
        )

        # Part name lookup tambahan (best-effort dari cache SIMS yg sudah ada)
        part_name_lookup = {}
        try:
            from sims_fetcher import _load_part_info_json
            cache_pi = _load_part_info_json() or {}
            for pn, info in cache_pi.items():
                if isinstance(info, dict):
                    pname = info.get("partName", "")
                    if pname:
                        part_name_lookup[str(pn).strip().upper()] = pname
        except Exception:
            pass

        draft   = _so.load_draft(username)
        history = _so.load_history(username)

        # ── Banner sesi aktif / form upload data stok awal ───────────
        if draft is None:
            self._render_opname_init_upload(username)
        else:
            self._render_opname_session(username, draft, {}, part_name_lookup)

        # ── Riwayat opname ───────────────────────────────────────────
        st.divider()
        st.markdown("#### 🗂️ Riwayat Opname Saya")
        if not history:
            st.caption("Belum ada riwayat sesi yang difinalisasi.")
            return

        rows = []
        for s in history:
            sm = s.get("summary") or _so.summarize(s)
            rows.append({
                "Session ID":    s.get("session_id", "")[:8],
                "Difinalisasi":  s.get("finalized_at", "—"),
                "Total PN":      sm.get("total", 0),
                "Sudah Hitung":  sm.get("counted", 0),
                "Cocok":         sm.get("match", 0),
                "Berselisih":    sm.get("diff_count", 0),
                "Selisih Net":   sm.get("selisih_net", 0),
                "_full_id":      s.get("session_id", ""),
            })
        df_hist = pd.DataFrame(rows)
        st.dataframe(df_hist.drop(columns=["_full_id"]), hide_index=True, use_container_width=True)

        with st.expander("📥 Download / Hapus Riwayat"):
            for s in history:
                sid_full = s.get("session_id", "")
                sid      = sid_full[:8]
                fin_at   = s.get("finalized_at", "—")
                col1, col2, col3 = st.columns([3, 1, 1])
                with col1:
                    st.markdown(f"**{sid}** — {fin_at}")
                with col2:
                    try:
                        xls_bytes = _so.make_report_excel(s, part_name_lookup)
                        st.download_button(
                            "⬇️ Excel",
                            data=xls_bytes,
                            file_name=f"opname_{username}_{sid}.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            key=f"opname_dl_{sid_full}",
                            use_container_width=True,
                        )
                    except Exception as e:
                        st.error(f"Err: {e}")
                with col3:
                    if st.button("🗑️", key=f"opname_del_{sid_full}", help="Hapus dari riwayat", use_container_width=True):
                        if _so.delete_history_entry(username, sid_full):
                            st.success("Dihapus.")
                            st.rerun()
                        else:
                            st.error("Gagal hapus.")

    def _render_opname_init_upload(self, username: str):
        """Form awal: user upload Excel data stok awal, lalu buat sesi."""
        st.info("Belum ada sesi opname aktif. Mulai dengan **upload data stok** Anda.")

        st.markdown("**Langkah 1 — Download template (opsional)**")
        try:
            tpl = _so.make_initial_template_excel()
            st.download_button(
                "📄 Download Template Stok Awal",
                data=tpl,
                file_name="template_stok_awal.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="opname_dl_init_tpl",
            )
        except Exception as e:
            st.error(f"Gagal generate template: {e}")

        st.markdown("**Langkah 2 — Upload data stok awal**")
        st.caption(
            "Format: kolom **Part Number**, **Qty Sistem** (atau **Stok**), "
            "dan opsional **Part Name**. Header bebas posisi; jika tanpa header, "
            "sistem akan pakai kolom A=PN, B=Qty, C=Part Name."
        )
        uploaded = st.file_uploader(
            "📤 Upload Excel data stok awal",
            type=["xlsx", "xls", "xlsm"],
            key="opname_init_upload",
        )
        if uploaded is None:
            return

        try:
            content = uploaded.read()
        except Exception as e:
            st.error(f"Gagal baca file: {e}")
            return

        parsed, warns = _so.parse_stok_upload(content)
        for w in warns:
            st.warning(w)

        if not parsed:
            st.error("❌ Tidak ada data valid pada file. Periksa format dan coba lagi.")
            return

        # ── Preview ──────────────────────────────────────────────────
        st.success(f"✅ {len(parsed)} Part Number terbaca dari file.")
        df_prev = pd.DataFrame([
            {
                "Part Number": pn,
                "Part Name":   p.get("part_name", ""),
                "Qty Sistem":  p.get("qty_sistem"),
            }
            for pn, p in parsed.items()
        ])
        with st.expander(f"👁️ Preview Data ({len(df_prev)} baris)", expanded=True):
            st.dataframe(df_prev.head(200), hide_index=True, use_container_width=True)
            if len(df_prev) > 200:
                st.caption(f"...dan {len(df_prev) - 200} baris lainnya.")

        # Statistik singkat
        n_zero  = sum(1 for p in parsed.values() if (p.get("qty_sistem") or 0) == 0)
        n_none  = sum(1 for p in parsed.values() if p.get("qty_sistem") is None)
        n_named = sum(1 for p in parsed.values() if (p.get("part_name") or "").strip())
        c1, c2, c3, c4 = st.columns(4)
        c1.metric("Total PN", len(parsed))
        c2.metric("Qty kosong/N/A", n_none)
        c3.metric("Qty 0", n_zero)
        c4.metric("Punya Part Name", n_named)

        st.markdown("---")
        if st.button("✅ Buat Sesi Opname dari Data Ini", type="primary",
                     use_container_width=True, key="opname_create_from_upload"):
            new_sess = _so.build_new_session(parsed, username, source_filename=uploaded.name)
            ok, err = _so.save_draft(username, new_sess)
            if not ok:
                self._opname_show_save_error(err)
            else:
                st.success(f"✅ Sesi baru dibuat — {len(new_sess['items'])} PN siap diopname.")
                st.rerun()

    def _opname_show_save_error(self, err: str):
        """Tampilkan error simpan opname + hint actionable."""
        e = (err or "").lower()
        is_missing_table = (
            "pgrst205" in e
            or "could not find the table" in e
            or ("404" in e and "opname_sessions" in e)
        )
        if is_missing_table:
            st.error("❌ Tabel `opname_sessions` belum ada di Supabase.")
            st.markdown(
                "**Cara fix:** buka Supabase Dashboard → SQL Editor → paste & Run "
                "isi file `migrations/002_opname.sql` (atau klik expander di bawah)."
            )
            with st.expander("📋 SQL migrasi (klik untuk copy)", expanded=False):
                st.code(
                    """CREATE TABLE IF NOT EXISTS opname_sessions (
    id           BIGINT GENERATED BY DEFAULT AS IDENTITY PRIMARY KEY,
    session_id   UUID        NOT NULL UNIQUE,
    username     TEXT        NOT NULL,
    is_draft     BOOLEAN     NOT NULL DEFAULT TRUE,
    payload      JSONB       NOT NULL,
    started_at   TIMESTAMPTZ,
    updated_at   TIMESTAMPTZ NOT NULL DEFAULT NOW(),
    finalized_at TIMESTAMPTZ
);
CREATE INDEX IF NOT EXISTS idx_opname_user_draft
    ON opname_sessions (username) WHERE is_draft = TRUE;
CREATE INDEX IF NOT EXISTS idx_opname_user_history
    ON opname_sessions (username, finalized_at DESC) WHERE is_draft = FALSE;
CREATE UNIQUE INDEX IF NOT EXISTS uniq_opname_one_draft_per_user
    ON opname_sessions (username) WHERE is_draft = TRUE;""",
                    language="sql",
                )
        else:
            st.error(f"Gagal menyimpan sesi: {err}")

    def _render_opname_session(self, username: str, draft: dict, stok_cache: dict, part_name_lookup: dict):
        """Render UI untuk 1 sesi opname draft yang aktif."""
        items = draft.get("items", {}) or {}
        summary = _so.summarize(draft)

        # ── Header info ──────────────────────────────────────────────
        src_file = draft.get("source_filename", "")
        src_html = f" • Sumber data: <code>{src_file}</code>" if src_file else ""
        st.markdown(
            f"""
<div style="background:#EEF2FF;border-left:4px solid #4F46E5;padding:8px 12px;
            border-radius:6px;margin:6px 0;font-size:.85rem;">
🟢 <b>Sesi aktif</b> — dibuat <code>{draft.get('started_at','—')}</code>,
update terakhir <code>{draft.get('updated_at','—')}</code>{src_html}
</div>""",
            unsafe_allow_html=True,
        )

        # ── Summary cards ────────────────────────────────────────────
        m1, m2, m3, m4, m5 = st.columns(5)
        m1.metric("Total PN", summary["total"])
        m2.metric("Sudah Dihitung", summary["counted"], f"-{summary['uncounted']} blm")
        m3.metric("Cocok", summary["match"])
        m4.metric("Berselisih", summary["diff_count"])
        m5.metric("Selisih Net", summary["selisih_net"])

        # ── Sub-tabs: Manual / Upload / Selisih ──────────────────────
        st_input, st_upload, st_diff = st.tabs([
            "📝 Ketik Manual",
            "📤 Upload Excel",
            "📊 Selisih & Belum Dihitung",
        ])

        with st_input:
            self._render_opname_manual(username, draft, items, part_name_lookup)
        with st_upload:
            self._render_opname_upload(username, draft, items, part_name_lookup)
        with st_diff:
            self._render_opname_diff(items, part_name_lookup)

        # ── Action buttons (finalize / cancel) ───────────────────────
        st.divider()
        c1, c2, c3 = st.columns([1, 1, 1])
        with c1:
            if st.button("💾 Simpan Draft", use_container_width=True, key="opname_save_btn"):
                ok, err = _so.save_draft(username, draft)
                if ok:
                    st.success("✅ Draft disimpan.")
                else:
                    self._opname_show_save_error(err)
        with c2:
            if st.button("✅ Finalisasi Sesi", type="primary", use_container_width=True, key="opname_final_btn"):
                if summary["counted"] == 0:
                    st.warning("Belum ada PN yang dihitung. Isi minimal 1 qty fisik.")
                else:
                    ok, err = _so.finalize_session(username, draft)
                    if ok:
                        st.success("✅ Sesi difinalisasi & masuk Riwayat.")
                        st.rerun()
                    else:
                        self._opname_show_save_error(err)
        with c3:
            if st.button("🗑️ Batal & Hapus Draft", use_container_width=True, key="opname_cancel_btn"):
                if _so.delete_draft(username):
                    st.success("Draft dihapus.")
                    st.rerun()
                else:
                    st.error("Gagal hapus draft.")

    def _render_opname_manual(self, username: str, draft: dict, items: dict, part_name_lookup: dict):
        st.caption("Edit kolom **Qty Fisik** dan **Note** langsung di tabel. Klik **Simpan** untuk persist.")

        # Filter
        fc1, fc2 = st.columns([2, 1])
        with fc1:
            search_pn = st.text_input(
                "🔍 Filter Part Number / Part Name",
                key="opname_filter_pn",
                placeholder="Ketik untuk filter...",
            )
        with fc2:
            show_only_uncounted = st.checkbox("Hanya yang belum dihitung", key="opname_only_uncounted")

        df_full = _so.items_to_df(items, part_name_lookup)

        df_view = df_full
        if search_pn:
            q = search_pn.strip().upper()
            df_view = df_view[
                df_view["Part Number"].astype(str).str.upper().str.contains(q, na=False)
                | df_view["Part Name"].astype(str).str.upper().str.contains(q, na=False)
            ]
        if show_only_uncounted:
            df_view = df_view[df_view["Qty Fisik"].isna()]

        if df_view.empty:
            st.info("Tidak ada PN yang cocok dengan filter.")
            return

        st.caption(f"Menampilkan {len(df_view)} dari {len(df_full)} PN.")

        edited = st.data_editor(
            df_view,
            hide_index=True,
            use_container_width=True,
            num_rows="fixed",
            column_config={
                "Part Number": st.column_config.TextColumn(disabled=True, width="medium"),
                "Part Name":   st.column_config.TextColumn(disabled=True, width="large"),
                "Qty Sistem":  st.column_config.NumberColumn(disabled=True, width="small"),
                "Qty Fisik":   st.column_config.NumberColumn(width="small", min_value=0, step=1),
                "Selisih":     st.column_config.NumberColumn(disabled=True, width="small"),
                "Note":        st.column_config.TextColumn(width="medium"),
            },
            key="opname_editor",
        )

        if st.button("💾 Simpan Perubahan", type="primary", key="opname_apply_manual"):
            new_items = _so.df_to_items(edited, items)
            draft["items"] = new_items
            ok, err = _so.save_draft(username, draft)
            if ok:
                st.success("✅ Perubahan tersimpan.")
                st.rerun()
            else:
                st.error(f"Gagal simpan: {err}")

    def _render_opname_upload(self, username: str, draft: dict, items: dict, part_name_lookup: dict):
        st.caption("Download template, isi kolom **Qty Fisik** (dan **Note** opsional), lalu upload kembali.")

        # Template download
        try:
            tpl_bytes = _so.make_template_excel(draft, part_name_lookup)
            st.download_button(
                "📄 Download Template Opname",
                data=tpl_bytes,
                file_name=f"template_opname_{username}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="opname_dl_template",
            )
        except Exception as e:
            st.error(f"Gagal generate template: {e}")

        st.markdown("---")

        uploaded = st.file_uploader(
            "📤 Upload Excel hasil opname",
            type=["xlsx", "xls", "xlsm"],
            key="opname_upload_file",
        )
        if uploaded is None:
            return

        try:
            content = uploaded.read()
        except Exception as e:
            st.error(f"Gagal baca file: {e}")
            return

        parsed, warns = _so.parse_uploaded_excel(content)
        for w in warns:
            st.warning(w)

        if not parsed:
            return

        st.info(f"📥 {len(parsed)} baris terbaca dari Excel.")

        merge_mode = st.radio(
            "Mode merge:",
            [
                "🔁 Replace (overwrite qty fisik existing)",
                "➕ Hanya isi yang masih kosong",
            ],
            horizontal=False,
            key="opname_upload_mode",
        )

        if st.button("✅ Terapkan ke Sesi", type="primary", key="opname_apply_upload"):
            new_items = dict(items)
            applied   = 0
            unmatched = 0
            skipped   = 0
            for pn, payload in parsed.items():
                if pn not in new_items:
                    unmatched += 1
                    continue
                qf  = payload.get("qty_fisik")
                cur = new_items[pn].get("qty_fisik")
                if "kosong" in merge_mode and cur is not None:
                    skipped += 1
                    continue
                new_items[pn]["qty_fisik"] = qf
                note = payload.get("note", "")
                if note:
                    new_items[pn]["note"] = note
                applied += 1

            draft["items"] = new_items
            ok, err = _so.save_draft(username, draft)
            if not ok:
                st.error(f"Gagal simpan: {err}")
                return

            st.success(
                f"✅ {applied} PN diupdate"
                + (f", {skipped} dilewati (sudah terisi)" if skipped else "")
                + (f", {unmatched} PN tidak ada di sesi" if unmatched else "")
                + "."
            )
            st.rerun()

    def _render_opname_diff(self, items: dict, part_name_lookup: dict):
        df_all = _so.items_to_df(items, part_name_lookup)
        if df_all.empty:
            st.info("Belum ada data.")
            return

        df_diff = df_all[(df_all["Selisih"].notna()) & (df_all["Selisih"] != 0)]
        df_unc  = df_all[df_all["Qty Fisik"].isna()]

        st.markdown(f"**📊 PN Berselisih — {len(df_diff)} item**")
        if df_diff.empty:
            st.caption("Tidak ada selisih.")
        else:
            df_show = df_diff.copy()
            df_show = df_show.sort_values("Selisih", key=lambda s: s.abs(), ascending=False)
            st.dataframe(df_show, hide_index=True, use_container_width=True)

        st.markdown(f"**🕒 Belum Dihitung — {len(df_unc)} item**")
        if df_unc.empty:
            st.caption("Semua PN sudah dihitung.")
        else:
            st.dataframe(df_unc[["Part Number", "Part Name", "Qty Sistem"]], hide_index=True, use_container_width=True)

    # ── Batch Download Tab ───────────────────────────────────────────
    def render_batch_download_tab(self):
        st.markdown("### 📥 Batch Download")
        st.markdown("""
<div class="batch-info-box">
Upload file Excel berisi daftar Part Number (1 kolom, mulai baris 1 atau 2).<br>
Sistem akan mencari semua PN secara otomatis dan menghasilkan file katalog.
</div>
""", unsafe_allow_html=True)

        st.download_button(
            label="📄 Download Template Input",
            data=make_template_excel(),
            file_name="template_batch_input.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        st.divider()

        # ── Pilih metode input ────────────────────────────────────────
        input_mode_batch = st.radio(
            "Metode Input:",
            ["📁 Upload File Excel", "⌨️ Ketik Manual"],
            horizontal=True,
            key="batch_download_input_mode",
        )

        part_numbers_raw = []

        if input_mode_batch == "📁 Upload File Excel":
            uploaded = st.file_uploader(
                "📂 Upload file Part Number:",
                type=["xlsx", "xls", "xlsm", "csv"],
                key="batch_upload",
            )

            if uploaded is None:
                return

            try:
                if uploaded.name.endswith(".csv"):
                    df_input = pd.read_csv(uploaded, header=None, dtype=str)
                else:
                    df_input = pd.read_excel(uploaded, header=None, dtype=str)
            except Exception as e:
                st.error(f"Gagal membaca file: {e}")
                return

            col_a = df_input.iloc[:, 0].dropna().astype(str).str.strip()
            if len(col_a) > 0 and col_a.iloc[0].lower() in ("part number","part_number","partnumber","no part","kode"):
                col_a = col_a.iloc[1:]

            part_numbers_raw = col_a[col_a.str.len() > 0].tolist()

        else:  # Ketik Manual
            manual_text_batch = st.text_area(
                "Masukkan Part Number (satu per baris):",
                height=180,
                placeholder="WG1641230025\nWG9725520274\n...",
                key="batch_download_manual_text",
            )
            if not manual_text_batch.strip():
                return
            part_numbers_raw = [p.strip() for p in manual_text_batch.splitlines() if p.strip()]

        if not part_numbers_raw:
            st.warning("Tidak ada Part Number yang valid dalam file.")
            return

        # Hapus duplikat, pertahankan urutan kemunculan pertama
        seen = set()
        part_numbers = []
        duplicates   = []
        for pn in part_numbers_raw:
            pn_up = pn.upper()
            if pn_up not in seen:
                seen.add(pn_up)
                part_numbers.append(pn)
            else:
                duplicates.append(pn)

        st.info(f"📊 **{len(part_numbers)}** Part Number unik ditemukan dalam file input.")
        if duplicates:
            st.warning(f"⚠️ **{len(duplicates)}** duplikat dihapus: {', '.join(duplicates)}")

        with st.expander("👁️ Preview Part Number"):
            st.dataframe(pd.DataFrame({"Part Number": part_numbers}), hide_index=True, height=200)

        # ── Pilihan konten output ─────────────────────────────────────
        with st.expander("⚙️ Pilih Konten Output", expanded=True):
            st.caption("Centang kolom yang ingin dimasukkan ke file hasil download:")
            opt_col1, opt_col2, opt_col3 = st.columns(3)
            with opt_col1:
                batch_opt_partname   = st.checkbox("📝 Part Name",    value=True, key="batch_opt_partname")
                batch_opt_kecocokan  = st.checkbox("📁 Kecocokan File", value=True, key="batch_opt_kecocokan")
            with opt_col2:
                batch_opt_stok       = st.checkbox("📦 Stok",         value=True, key="batch_opt_stok")
                batch_opt_qty        = st.checkbox("🔢 Qty",           value=True, key="batch_opt_qty")
            with opt_col3:
                batch_opt_images     = st.checkbox("🖼️ Gambar (dari SIMS)", value=True, key="batch_opt_images")
            if not batch_opt_images:
                st.info("💡 Tanpa gambar, proses akan jauh lebih cepat.")

        if st.button("🔍 Proses Batch Search", type="primary", use_container_width=True, key="batch_process_btn"):
            if not st.session_state.excel_files:
                st.error("Tidak ada file Excel yang ter-index di folder data/.")
                st.stop()

            prog        = st.progress(0)
            status_txt  = st.empty()
            total       = len(part_numbers)
            results_all = []

            for i, pn in enumerate(part_numbers):
                status_txt.text(f"🔍 Mencari {i+1}/{total}: {pn}")
                prog.progress((i + 1) / total)
                found = search_part_number(pn, st.session_state.excel_files, self.stok_cache, self.harga_lookup)
                if found:
                    # Gabungkan semua file yang cocok ke 1 baris saja
                    hasil_list = [r["File"] for r in found]
                    sheet_list = [r["Sheet"] for r in found]
                    results_all.append({
                        "Part Number": pn,
                        "_pn_group":   pn,
                        "Hasil":       ", ".join(hasil_list),
                        "Sheet":       ", ".join(sheet_list),
                        "Part Name":   found[0]["Part Name"],
                        "Qty":         found[0]["Quantity"],
                        "Stok":        found[0]["Stok"],
                        "Status":      "✅ Ditemukan",
                    })
                else:
                    results_all.append({
                        "Part Number": pn, "_pn_group": pn,
                        "Hasil": "", "Sheet": "", "Part Name": "",
                        "Qty": "", "Stok": "", "Status": "❌ Tidak ditemukan",
                    })

            prog.empty()
            status_txt.empty()
            df_result = pd.DataFrame(results_all)

            # ── Ambil Part Name dari SIMS untuk PN yang tidak ditemukan di database lokal ──
            if SIMS_ENABLED:
                try:
                    from sims_fetcher import get_sims_part_info
                except ImportError:
                    get_sims_part_info = None

                if get_sims_part_info:
                    not_found_mask = df_result["Part Name"] == ""
                    not_found_pns  = df_result.loc[not_found_mask, "Part Number"].tolist()
                    if not_found_pns:
                        sims_prog   = st.progress(0)
                        sims_status = st.empty()
                        for si, nf_pn in enumerate(not_found_pns):
                            sims_status.text(f"🔎 Ambil Part Name dari SIMS {si+1}/{len(not_found_pns)}: {nf_pn}")
                            sims_prog.progress((si + 1) / len(not_found_pns))
                            try:
                                sims_info, _ = get_sims_part_info(nf_pn)
                                if sims_info and sims_info.get("partName"):
                                    idx = df_result.index[df_result["Part Number"] == nf_pn].tolist()
                                    for i in idx:
                                        df_result.at[i, "Part Name"] = sims_info["partName"]
                            except Exception:
                                pass
                        sims_prog.empty()
                        sims_status.empty()

            # Import get_sims_part_info jika belum tersedia di scope ini
            try:
                from sims_fetcher import get_sims_part_info
            except ImportError:
                pass

            # Simpan pilihan opsi output ke session_state
            batch_options = {
                "partname":  st.session_state.get("batch_opt_partname",  True),
                "kecocokan": st.session_state.get("batch_opt_kecocokan", True),
                "stok":      st.session_state.get("batch_opt_stok",      True),
                "qty":       st.session_state.get("batch_opt_qty",       True),
                "images":    st.session_state.get("batch_opt_images",    True),
            }
            st.session_state["batch_options"] = batch_options

            prog_cat   = st.progress(0)
            status_cat = st.empty()

            def _prog(i, tot, pn):
                prog_cat.progress((i + 1) / max(tot, 1))
                status_cat.text(f"🖼️ Fetch gambar {i+1}/{tot}: {pn}")

            try:
                cat_bytes = build_catalog_excel(
                    df_result,
                    progress_callback=_prog,
                    all_part_numbers=part_numbers,
                    options=batch_options,
                )
                st.session_state["batch_catalog_bytes"]     = cat_bytes
                st.session_state["batch_catalog_df"]        = df_result
                st.session_state["batch_catalog_timestamp"] = datetime.now().strftime("%Y%m%d_%H%M%S")
            except Exception as e:
                st.error(f"❌ Gagal membuat katalog: {e}")
            finally:
                prog_cat.empty()
                status_cat.empty()

            st.rerun()

        if "batch_catalog_df" not in st.session_state:
            return

        df_result = st.session_state["batch_catalog_df"]
        found_pn  = df_result[df_result["Status"] == "✅ Ditemukan"]["_pn_group"].nunique()
        not_found = df_result["_pn_group"].nunique() - found_pn

        c1, c2, c3 = st.columns(3)
        c1.metric("Total Part Number", df_result["_pn_group"].nunique())
        c2.metric("✅ Ditemukan", found_pn)
        c3.metric("❌ Tidak Ditemukan", not_found)

        st.markdown("#### 📋 Preview Hasil")
        user_b = LoginManager.get_current_user()
        role_b = user_b["role"] if user_b else "user"
        allowed_cols_b = get_allowed_columns(user_b["username"], role_b)

        disp_cols = ["Part Number", "Hasil", "Sheet", "Part Name", "Qty"]
        if "col_stok" in allowed_cols_b:
            disp_cols.append("Stok")
        disp_cols.append("Status")

        # Pastikan Part Name yang kosong/N/A tetap tampil dengan benar
        df_preview = df_result[disp_cols].copy()
        df_preview["Part Name"] = df_preview["Part Name"].replace({"": "—", "N/A": "—"}).fillna("—")

        st.dataframe(
            df_preview, hide_index=True,
            use_container_width=True,
            column_config={
                "Part Number": st.column_config.TextColumn(width="medium"),
                "Hasil":       st.column_config.TextColumn(width="medium"),
                "Sheet":       st.column_config.TextColumn(width="medium"),
                "Part Name":   st.column_config.TextColumn(width="large"),
                "Qty":         st.column_config.TextColumn(width="small"),
                "Stok":        st.column_config.TextColumn(width="small"),
                "Status":      st.column_config.TextColumn(width="medium"),
            }
        )

        if "batch_catalog_bytes" in st.session_state:
            ts = st.session_state.get("batch_catalog_timestamp", "result")
            st.download_button(
                label="⬇️ Download Hasil (.xlsx)",
                data=st.session_state["batch_catalog_bytes"],
                file_name=f"catalog_{ts}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                type="primary",
                use_container_width=True,
            )

    # ── Tab Harga ────────────────────────────────────────────────────
    def render_harga_tab(self):
        st.markdown("### 💰 Harga Sparepart")

        # Cek izin akses kolom harga
        user_h = LoginManager.get_current_user()
        role_h = user_h["role"] if user_h else "user"
        if "col_harga" not in get_allowed_columns(user_h["username"], role_h):
            st.warning("🔒 Anda tidak memiliki akses untuk melihat data harga. Hubungi admin jika diperlukan.")
            return

        # ── Sub-tab di dalam Tab Harga ────────────────────────────────
        # Cek izin akses sub-tab harga
        allowed_hs = get_allowed_harga_subtabs(user_h["username"], role_h)
        _hs_tabs   = []
        _hs_keys   = []
        if "subtab_list_harga" in allowed_hs:
            _hs_tabs.append("📋 List Harga")
            _hs_keys.append("subtab_list_harga")
        if "subtab_cari_harga" in allowed_hs:
            _hs_tabs.append("🔍 Cari Harga")
            _hs_keys.append("subtab_cari_harga")
        if "subtab_batch_harga" in allowed_hs:
            _hs_tabs.append("📥 Batch Cari Harga")
            _hs_keys.append("subtab_batch_harga")
        if not _hs_tabs:
            st.warning("🔒 Anda tidak memiliki akses ke sub-tab manapun di Harga. Hubungi admin.")
            return
        _rendered_tabs = st.tabs(_hs_tabs)
        sub_list  = _rendered_tabs[_hs_keys.index("subtab_list_harga")]  if "subtab_list_harga"  in _hs_keys else None
        sub_cari  = _rendered_tabs[_hs_keys.index("subtab_cari_harga")]  if "subtab_cari_harga"  in _hs_keys else None
        sub_batch = _rendered_tabs[_hs_keys.index("subtab_batch_harga")] if "subtab_batch_harga" in _hs_keys else None


        # ══════════════════════════════════════════════════════════════
        # SUB-TAB 1: LIST HARGA
        # ══════════════════════════════════════════════════════════════
        if sub_list is not None:
         with sub_list:
            st.markdown("#### 📋 Daftar Harga Sparepart")

            col_reload, _ = st.columns([1, 5])
            with col_reload:
                if st.button("🔄 Reload Harga", key="reload_harga"):
                    st.session_state.pop("harga_data", None)
                    self.harga_cache = None
                    self._load_harga_data()
                    st.rerun()

            df_h = self.harga_cache if self.harga_cache is not None else pd.DataFrame()

            if df_h.empty:
                st.warning(
                    "File harga tidak ditemukan atau kosong. "
                    "Pastikan file ada di `data/harga/harga.xlsx`."
                )
            else:
                st.divider()

                with st.expander("🔍 Filter & Pencarian", expanded=True):
                    col_s, col_sort = st.columns([3, 2])
                    with col_s:
                        kw_harga = st.text_input(
                            "Cari Part Number / Part Name:",
                            placeholder="Ketik untuk filter…",
                            key="harga_search_kw",
                        )
                    with col_sort:
                        sort_by = st.selectbox(
                            "Urutkan berdasarkan:",
                            ["Part Number", "Part Name", "Harga (Terendah)", "Harga (Tertinggi)"],
                            key="harga_sort",
                        )

                # df_show: filter → sort tanpa .copy() awal. df_h tidak
                # di-mutate karena semua langkah return DataFrame baru.
                if kw_harga.strip():
                    kw_up = kw_harga.strip().upper()
                    mask = (
                        df_h["Part Number"].str.upper().str.contains(kw_up, na=False) |
                        df_h["Part Name"].astype(str).str.upper().str.contains(kw_up, na=False)
                    )
                    df_show = df_h[mask].reset_index(drop=True)
                else:
                    df_show = df_h

                def _harga_sort_key(s):
                    return pd.to_numeric(
                        s.astype(str).str.replace(r"[^\d.]", "", regex=True),
                        errors="coerce",
                    )

                try:
                    if sort_by == "Harga (Terendah)":
                        df_show = df_show.sort_values(
                            "Harga", key=_harga_sort_key, ascending=True
                        ).reset_index(drop=True)
                    elif sort_by == "Harga (Tertinggi)":
                        df_show = df_show.sort_values(
                            "Harga", key=_harga_sort_key, ascending=False
                        ).reset_index(drop=True)
                    elif sort_by == "Part Name":
                        df_show = df_show.sort_values(
                            "Part Name", key=lambda x: x.str.upper()
                        ).reset_index(drop=True)
                    else:
                        df_show = df_show.sort_values("Part Number").reset_index(drop=True)
                except Exception:
                    pass

                c1, c2 = st.columns(2)
                c1.metric("Total Part", len(df_h))
                c2.metric("Hasil Filter", len(df_show))

                st.markdown("---")

                def fmt_harga(val):
                    try:
                        num = float(str(val).replace(",", "").strip())
                        return f"Rp {num:,.0f}"
                    except Exception:
                        return str(val)

                # Construct df_display fresh — no .copy() needed.
                df_display = pd.DataFrame({
                    "Part Number": df_show["Part Number"].to_numpy(),
                    "Part Name":   df_show["Part Name"].to_numpy(),
                    "Harga (Rp)":  df_show["Harga"].map(fmt_harga).to_numpy(),
                })

                if df_display.empty:
                    st.info("Tidak ada data yang cocok dengan pencarian.")
                else:
                    st.dataframe(
                        df_display,
                        hide_index=True,
                        use_container_width=True,
                        height=min(400, 56 + len(df_display) * 35),
                        column_config={
                            "Part Number": st.column_config.TextColumn("Part Number", width="medium"),
                            "Part Name":   st.column_config.TextColumn("Part Name",   width="large"),
                            "Harga (Rp)":  st.column_config.TextColumn("Harga",       width="medium"),
                        },
                    )

                    dl_buf = io.BytesIO()
                    df_display.to_excel(dl_buf, index=False, engine="openpyxl")
                    dl_buf.seek(0)
                    st.download_button(
                        label="⬇️ Download Excel",
                        data=dl_buf.getvalue(),
                        file_name=f"harga_sparepart_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        key="harga_download",
                    )

        # ══════════════════════════════════════════════════════════════
        # SUB-TAB 2: CARI HARGA (dari SIMS)
        # ══════════════════════════════════════════════════════════════
        if sub_cari is not None:
         with sub_cari:
            st.markdown("#### 🔍 Cari Harga Part dari SIMS")
            st.caption("Ambil Part Price langsung dari SIMS berdasarkan Part Number. Harga SIMS dalam CNY, dikonversi ke IDR otomatis.")

            # ── Helper: ambil kurs CNY → IDR ─────────────────────────
            def _get_cny_to_idr_rate() -> tuple:
                """Ambil kurs CNY→IDR dari API. Cache di session_state selama 30 menit."""
                import time as _time
                cache_key   = "_cny_idr_rate_cache"
                cache_ts_key = "_cny_idr_rate_ts"
                now = _time.time()
                cached_rate = st.session_state.get(cache_key)
                cached_ts   = st.session_state.get(cache_ts_key, 0)
                if cached_rate and (now - cached_ts) < 1800:
                    return cached_rate, None
                try:
                    resp = requests.get(
                        "https://api.exchangerate-api.com/v4/latest/CNY",
                        timeout=8
                    )
                    resp.raise_for_status()
                    data = resp.json()
                    rate = data["rates"].get("IDR")
                    if rate:
                        st.session_state[cache_key]   = float(rate)
                        st.session_state[cache_ts_key] = now
                        return float(rate), None
                    return None, "IDR tidak ada di response"
                except Exception as e:
                    # Fallback ke kurs statis jika API gagal
                    FALLBACK_RATE = 2200.0
                    st.session_state[cache_key]   = FALLBACK_RATE
                    st.session_state[cache_ts_key] = now
                    return FALLBACK_RATE, f"API gagal ({e}), menggunakan kurs fallback Rp {FALLBACK_RATE:,.0f}/CNY"

            # Tampilkan info kurs saat ini
            rate_cny_idr, rate_err = _get_cny_to_idr_rate()
            if rate_err:
                st.warning(f"⚠️ {rate_err}")
            if rate_cny_idr:
                col_rate, col_rate_refresh = st.columns([4, 1])
                with col_rate:
                    st.info(f"💱 Kurs saat ini: **1 CNY = Rp {rate_cny_idr:,.2f}**")
                with col_rate_refresh:
                    if st.button("🔄 Update Kurs", key="refresh_kurs_cny", use_container_width=True):
                        st.session_state.pop("_cny_idr_rate_cache", None)
                        st.session_state.pop("_cny_idr_rate_ts", None)
                        st.rerun()

            try:
                from sims_price_fetcher import get_sims_part_price
                price_fetcher_ok = True
            except ImportError:
                price_fetcher_ok = False

            if not price_fetcher_ok:
                st.warning("⚠️ `sims_price_fetcher.py` tidak ditemukan.")
            else:
                col_input, col_btn, col_refresh = st.columns([3, 1, 1])
                with col_input:
                    pn_input = st.text_input(
                        "Part Number",
                        placeholder="Contoh: WG1641230025",
                        label_visibility="collapsed",
                        key="sims_harga_pn_input",
                    ).strip().upper()
                with col_btn:
                    cari = st.button("🔍 Cari", type="primary", use_container_width=True, key="sims_harga_cari_btn")
                with col_refresh:
                    force = st.button("🔄 Refresh", use_container_width=True, key="sims_harga_refresh_btn",
                                      help="Abaikan cache, ambil ulang dari SIMS")

                if (cari or force) and pn_input:
                    result_key = f"sims_harga_result_{pn_input}"
                    if force:
                        st.session_state.pop(result_key, None)
                    if result_key not in st.session_state:
                        with st.spinner(f"Mengambil harga **{pn_input}** dari SIMS..."):
                            price, err = get_sims_part_price(pn_input, force_refresh=bool(force))
                        st.session_state[result_key] = {"price": price, "err": err, "pn": pn_input}

                result_key = f"sims_harga_result_{pn_input}" if pn_input else None
                res = st.session_state.get(result_key) if result_key else None

                if res:
                    if res["price"] is not None:
                        cny_price = res["price"]
                        idr_price = cny_price * rate_cny_idr if rate_cny_idr else None
                        idr_str   = f"Rp {idr_price:,.0f}" if idr_price is not None else "—"
                        st.markdown(
                            f"""
                            <div style="background:#E8F5E9;border-left:5px solid #4CAF50;
                                        padding:1rem 1.5rem;border-radius:0 10px 10px 0;margin-top:1rem;">
                                <div style="font-size:0.85rem;color:#555;">Part Number</div>
                                <div style="font-size:1.2rem;font-weight:700;color:#1B5E20;">{res['pn']}</div>
                                <div style="display:flex;gap:3rem;margin-top:12px;flex-wrap:wrap;">
                                    <div>
                                        <div style="font-size:0.8rem;color:#777;">Harga SIMS (CNY)</div>
                                        <div style="font-size:1.6rem;font-weight:800;color:#1565C0;">¥ {cny_price:,.2f}</div>
                                    </div>
                                    <div>
                                        <div style="font-size:0.8rem;color:#777;">Harga IDR</div>
                                        <div style="font-size:1.6rem;font-weight:800;color:#2E7D32;">{idr_str}</div>
                                    </div>
                                </div>
                                <div style="font-size:0.75rem;color:#999;margin-top:8px;">
                                    Kurs: 1 CNY = Rp {rate_cny_idr:,.2f}
                                </div>
                            </div>
                            """,
                            unsafe_allow_html=True,
                        )
                    else:
                        st.warning(f"⚠️ Harga tidak ditemukan untuk **{res['pn']}**.")
                        if res["err"]:
                            st.caption(f"Detail: {res['err']}")
                elif pn_input and not (cari or force):
                    st.info("Klik **Cari** untuk mengambil harga dari SIMS.")


        # ══════════════════════════════════════════════════════════════
        # SUB-TAB 3: BATCH CARI HARGA
        # ══════════════════════════════════════════════════════════════
        if sub_batch is not None:
         with sub_batch:
            # ── Kurs CNY → IDR ────────────────────────────────────────
            import time as _bhe_time
            _bhe_cache_key    = "_cny_idr_rate_cache"
            _bhe_cache_ts_key = "_cny_idr_rate_ts"
            _bhe_now = _bhe_time.time()
            _bhe_cached_rate = st.session_state.get(_bhe_cache_key)
            _bhe_cached_ts   = st.session_state.get(_bhe_cache_ts_key, 0)
            if _bhe_cached_rate and (_bhe_now - _bhe_cached_ts) < 1800:
                b_rate_batch     = _bhe_cached_rate
                b_rate_err_batch = None
            else:
                try:
                    _bhe_resp = requests.get("https://api.exchangerate-api.com/v4/latest/CNY", timeout=8)
                    _bhe_resp.raise_for_status()
                    _bhe_data = _bhe_resp.json()
                    _bhe_rate = _bhe_data["rates"].get("IDR")
                    if _bhe_rate:
                        st.session_state[_bhe_cache_key]    = float(_bhe_rate)
                        st.session_state[_bhe_cache_ts_key] = _bhe_now
                        b_rate_batch     = float(_bhe_rate)
                        b_rate_err_batch = None
                    else:
                        b_rate_batch     = 2200.0
                        b_rate_err_batch = "IDR tidak ada di response"
                except Exception as _bhe_e:
                    b_rate_batch     = 2200.0
                    b_rate_err_batch = f"API gagal ({_bhe_e}), menggunakan kurs fallback Rp 2.200/CNY"

            col_brate2, col_brate_ref2 = st.columns([4, 1])
            with col_brate2:
                if b_rate_err_batch:
                    st.warning(f"⚠️ {b_rate_err_batch}")
                if b_rate_batch:
                    st.info(f"💱 Kurs saat ini: **1 CNY = Rp {b_rate_batch:,.2f}**")
            with col_brate_ref2:
                if st.button("🔄 Update Kurs", key="bhe_kurs_refresh", use_container_width=True):
                    st.session_state.pop("_cny_idr_rate_cache", None)
                    st.session_state.pop("_cny_idr_rate_ts", None)
                    st.rerun()

            st.divider()

            try:
                from batch_harga_engine import render_batch_harga_tab
                render_batch_harga_tab(b_rate=b_rate_batch)
            except ImportError:
                st.error("❌ `batch_harga_engine.py` tidak ditemukan. Pastikan file sudah ada di direktori yang sama dengan app.py.")

    # ── SIDEBAR & DASHBOARD ──────────────────────────────────────────
    def display_dashboard(self):
        user = LoginManager.get_current_user()
        role = user["role"] if user else "user"
        inject_keep_alive()
        n_files = st.session_state.get("loaded_files_count", 0)
        idx_time = st.session_state.get("last_index_time")
        idx_human = idx_time.strftime("%H:%M:%S") if idx_time else "—"
        st.markdown(
            f"""
            <div style="display:flex;align-items:flex-end;justify-content:space-between;
                        gap:16px;flex-wrap:wrap;
                        padding:.5rem 0 1rem;border-bottom:1px solid var(--mp-line);
                        margin:-.5rem 0 1.5rem;">
                <div>
                    <h1 style="font-size:24px;font-weight:700;letter-spacing:-.015em;
                                margin:0;color:var(--mp-ink);">
                        Part Number <span style="color:var(--mp-green);">Finder</span>
                    </h1>
                    <div style="font-size:12.5px;color:var(--mp-ink-50);margin-top:4px;">
                        {n_files} file Excel terindex · Index ●&nbsp;Live · diperbarui {idx_human}
                    </div>
                </div>
                <div style="display:flex;gap:6px;font-size:11.5px;color:var(--mp-ink-50);">
                    <span class="mp-chip">● Online</span>
                </div>
            </div>
            """,
            unsafe_allow_html=True,
        )

        with st.sidebar:
            # Brand header
            st.markdown(
                '<div class="mp-sb-brand">'
                '<div class="logo">M</div>'
                '<div>'
                '<div class="name">MasPart</div>'
                '<div class="tag">Part Number Finder</div>'
                '</div>'
                '</div>',
                unsafe_allow_html=True,
            )

            # User badge card
            uname = user["username"]
            initials = "".join([p[0] for p in uname.replace(".", " ").split()[:2]]).upper() or uname[:2].upper()
            login_str = user["login_time"].strftime("%H:%M")
            role_class = "admin" if role == "admin" else "user"
            st.markdown(
                f'<div class="mp-sb-user">'
                f'<div class="row">'
                f'<div class="avatar">{initials}</div>'
                f'<div style="min-width:0;flex:1">'
                f'<div class="name">{uname.title()}</div>'
                f'<div class="meta">'
                f'<span class="role-pill {role_class}">{role.upper()}</span>'
                f' · {login_str} · timeout {SESSION_TIMEOUT_MINUTES // 60}h'
                f'</div>'
                f'</div>'
                f'</div>'
                f'</div>',
                unsafe_allow_html=True,
            )

            # ── Section: Sistem (status + refresh) ─────────────────
            st.markdown('<div class="mp-sb-section">Sistem</div>', unsafe_allow_html=True)
            n_files = st.session_state.get("loaded_files_count", 0)
            idx_time = st.session_state.get("last_index_time")
            idx_str = idx_time.strftime("%H:%M:%S") if idx_time else "—"
            st.markdown(
                f'<div class="mp-sb-stat"><span>File Excel</span><b>{n_files}</b></div>'
                f'<div class="mp-sb-stat"><span>Index</span><b class="green">● Live</b></div>'
                f'<div class="mp-sb-stat"><span>Terakhir index</span><b>{idx_str}</b></div>',
                unsafe_allow_html=True,
            )
            if st.button("🔄 Refresh Data", type="secondary", use_container_width=True, key="sb_refresh_data"):
                for cf in CACHE_FOLDER.glob("*.pkl"):
                    try: cf.unlink()
                    except Exception: pass
                for k in ("excel_files","last_index_time","last_file_count","stok_data",
                          "stok_gudang_data","stok_gudang_names",
                          "harga_data","harga_lookup"):
                    st.session_state.pop(k, None)
                self.stok_cache = None; self.stok_gudang_cache = None; self.gudang_names = []
                self.harga_cache = None; self.harga_lookup = {}
                self._load_stok_data(); self._load_harga_data()
                self.auto_load_excel_files()
                st.rerun()

            # ── Section: Admin Panel ───────────────────────────────
            if role == "admin":
                st.markdown('<div class="mp-sb-section">Admin Panel</div>', unsafe_allow_html=True)
                if st.button("👥 Reload Users", type="secondary", use_container_width=True, key="sb_reload_users"):
                    st.session_state.login_users_df = LoginManager._load_users()
                    st.toast("✅ Data user telah di-reload!")
                if st.button("🔐 Reload Menu Config", type="secondary", use_container_width=True, key="sb_reload_menu"):
                    MenuAccessManager.load_permissions(force=True)
                    st.toast("✅ Konfigurasi akses menu di-reload!")
                if st.button("🔒 Reload Kolom Config", type="secondary", use_container_width=True, key="sb_reload_col"):
                    ColumnAccessManager.load_permissions(force=True)
                    st.toast("✅ Konfigurasi akses kolom di-reload!")
                df_users = st.session_state.get("login_users_df", pd.DataFrame())
                if not df_users.empty:
                    with st.expander("📋 Daftar User"):
                        st.dataframe(
                            df_users[["username","role"]].rename(columns={"username":"Username","role":"Role"}),
                            hide_index=True,
                        )

            # ── Section: Bantuan ───────────────────────────────────
            st.markdown('<div class="mp-sb-section">Bantuan</div>', unsafe_allow_html=True)
            with st.expander("📖 Panduan Cepat"):
                st.markdown("""
1. Letakkan file Excel di folder `data/`
2. **Part Number** → kolom B | **Part Name** → kolom D
3. **Stok:** data/stok/stok.xlsx — total (Kol A=PN, Kol D=Stok) **atau** export "Kuantitas Barang per Gudang" (multi-gudang, auto-deteksi)
4. **Batch Download:** Upload Excel berisi PN di Kol A
                """)
            st.caption(f"📁 Data: `{self.data_folder.name}/`")

            # ── Logout di paling bawah ─────────────────────────────
            st.markdown('<div class="mp-sb-section">Akun</div>', unsafe_allow_html=True)
            if st.button("🚪 Logout", type="secondary", use_container_width=True, key="sb_logout"):
                LoginManager.logout()
                for k in ("excel_files","index_data","search_results",
                          "last_index_time","loaded_files_count","last_file_count"):
                    st.session_state.pop(k, None)
                st.rerun()

        # ── Trigger search dari Image Search ─────────────────────────
        # Trigger di-handle di dalam render_search_image_tab() sendiri agar
        # spinner muncul di lokasi user (bawah hasil image search), bukan
        # di atas tabs yang tidak terlihat saat scroll.

        # ── TABS ────────────────────────────────────────────────────
        st.markdown(TAB_PERSIST_JS, unsafe_allow_html=True)
        # Definisi semua tab dan method render-nya
        ALL_TAB_DEFS = [
            ("tab_search_pn",    "🔢 Search Part Number", "_render_tab_search_pn"),
            ("tab_search_name",  "📝 Search Part Name",    "_render_tab_search_name"),
            ("tab_search_image", "🖼️ Cari by Foto",        "__search_image__"),
            ("tab_compare",      "🔍 Bandingkan 2 Part",   "_render_tab_compare_parts"),
            ("tab_batch",        "📥 Batch Download",      "render_batch_download_tab"),
            ("tab_populasi",     "🚛 Populasi Unit",       "render_populasi_tab"),
            ("tab_harga",        "💰 Harga",               "render_harga_tab"),
            ("tab_opname",       "📋 Stok Opname",         "_render_tab_stok_opname"),
            ("tab_chat_ai",      "🤖 Chat AI",             "__chat_ai__"),
        ]
        if role == "admin":
            ALL_TAB_DEFS.append(("tab_menu_control", "🛡️ Menu Control",    "__menu_control__"))
            ALL_TAB_DEFS.append(("tab_data_upload",  "📊 Upload Data",     "__data_upload__"))
            ALL_TAB_DEFS.append(("tab_foto_part",    "📷 Upload Foto Part", "__foto_part__"))
            ALL_TAB_DEFS.append(("tab_image_index",  "🧠 Image Index",     "__image_index__"))
            if SUPABASE_ENABLED and render_user_management_tab is not None:
                ALL_TAB_DEFS.append(("tab_user_mgmt", "👥 User Management", "__user_mgmt__"))
            if USER_MONITORING_ENABLED:
                ALL_TAB_DEFS.append(("tab_user_monitoring", "📊 Monitoring User", "__user_monitoring__"))

        # Tentukan tab yang boleh dilihat user ini
        allowed_keys = set(get_allowed_tabs(user["username"], role))
        visible_tabs = [(k, lbl, fn) for k, lbl, fn in ALL_TAB_DEFS if k in allowed_keys]

        # Render tabs secara dinamis
        tab_objects = st.tabs([lbl for _, lbl, _ in visible_tabs])
        for tab_obj, (key, lbl, fn) in zip(tab_objects, visible_tabs):
            with tab_obj:
                if fn == "__menu_control__":
                    render_admin_menu_control_tab()
                elif fn == "__data_upload__":
                    render_data_uploader_tab()
                elif fn == "__foto_part__":
                    render_foto_part_tab()
                elif fn == "__search_image__":
                    render_search_image_tab()
                elif fn == "__image_index__":
                    render_image_index_tab()
                elif fn == "__user_mgmt__":
                    render_user_management_tab()
                elif fn == "__user_monitoring__":
                    render_user_monitoring_tab()
                elif fn == "__chat_ai__":
                    render_chat_ai_tab(
                        excel_files=st.session_state.get("excel_files"),
                        stok_cache=self.stok_cache,
                        harga_lookup=self.harga_lookup,
                    )
                else:
                    getattr(self, fn)()

        self.display_search_results()

    def _render_stok_cabang_section(self, df_res, gudang_scope):
        """
        Section di bawah hasil pencarian khusus akun cabang:
        tabel stok di gudang user + kolom 'terdekat' bila stok gudang = 0.
        Kolom Stok di tabel utama tetap total seluruh gudang.
        """
        gcache = self.stok_gudang_cache or {}
        city = gudang_label(gudang_scope)

        # PN unik (pertahankan urutan kemunculan) + Part Name
        seen = {}
        for _, r in df_res.iterrows():
            pn = str(r.get("Part Number", "")).strip().upper()
            if pn and pn not in seen:
                seen[pn] = str(r.get("Part Name", "") or "")
        if not seen:
            return

        rows = []
        for pn, pname in seen.items():
            bd = gcache.get(pn)
            if bd is None:                       # PN tak ada di data stok
                own_disp, near = "—", "—"
            else:
                own = bd.get(gudang_scope, 0)
                if own:
                    own_disp, near = str(own), "—"
                else:
                    own_disp, near = "0", "habis"
                    for g in fallback_order(gudang_scope, self.gudang_names):
                        q = bd.get(g, 0)
                        if q:
                            near = f"{gudang_label(g)} ({q})"
                            break
            rows.append({
                "Part Number":          pn,
                "Part Name":            pname,
                f"Stok {city}":         own_disp,
                "Stok Terdekat (jika 0)": near,
            })

        df_c = pd.DataFrame(rows)
        st.markdown(
            f'<div style="margin: 1rem 0 .4rem;font-size:14px;font-weight:700;'
            f'letter-spacing:-.01em;color:var(--mp-ink);">'
            f'📦 Stok Cabang Anda — {city}</div>'
            f'<div style="font-size:12px;color:var(--mp-ink-50);margin-bottom:8px;">'
            f'Kolom <b>Stok</b> di atas = total semua gudang. Tabel ini khusus '
            f'gudang <b>{city}</b>; jika 0, ditampilkan cabang terdekat yang ada stok.'
            f'</div>',
            unsafe_allow_html=True,
        )
        st.dataframe(
            df_c, hide_index=True, use_container_width=True,
            column_config={
                "Part Number":            st.column_config.TextColumn(width="medium"),
                "Part Name":              st.column_config.TextColumn(width="large"),
                f"Stok {city}":           st.column_config.TextColumn(width="small"),
                "Stok Terdekat (jika 0)": st.column_config.TextColumn(width="medium"),
            },
        )

    def _render_stok_gudang_breakdown(self, df_res):
        """Expander rincian stok per gudang untuk PN di hasil pencarian."""
        gcache = self.stok_gudang_cache or {}
        # PN unik dari hasil yang memang punya data gudang
        pns = [p for p in df_res["Part Number"].dropna().astype(str).str.upper().unique()
               if p in gcache]
        if not pns:
            return

        with st.expander("📦 Rincian stok per gudang", expanded=False):
            sel_pn = st.selectbox(
                "Pilih Part Number:", pns, key="stok_gudang_sel_pn",
            )
            breakdown = gcache.get(str(sel_pn).upper(), {}) or {}
            total = self.stok_cache.get(str(sel_pn).upper(), "—")

            if not breakdown:
                st.info(f"Stok **{sel_pn}** = 0 di semua gudang. Total: {total}")
                return

            # Tampilkan sesuai urutan kolom gudang asli
            order = {name: i for i, name in enumerate(self.gudang_names)}
            rows = sorted(breakdown.items(), key=lambda kv: order.get(kv[0], 9999))
            df_g = pd.DataFrame(
                [{"Gudang": name, "Qty": qty} for name, qty in rows]
            )
            st.markdown(
                f"<div style='font-size:13px;color:var(--mp-ink-50);margin-bottom:6px;'>"
                f"<b>{sel_pn}</b> · tersebar di {len(rows)} gudang · "
                f"Total <b>{total}</b></div>",
                unsafe_allow_html=True,
            )
            st.dataframe(
                df_g, hide_index=True, use_container_width=True,
                column_config={
                    "Gudang": st.column_config.TextColumn(width="large"),
                    "Qty":    st.column_config.NumberColumn(width="small"),
                },
            )

    def display_search_results(self):
        results = st.session_state.get("search_results", [])
        if results:
            user = LoginManager.get_current_user()
            role = user["role"] if user else "user"
            allowed_cols = get_allowed_columns(user["username"], role)
            search_term = st.session_state.get("search_term", "")

            # Cakupan gudang user: None = semua cabang (admin/mas),
            # str = nama gudang cabang user.
            gudang_scope = gudang_for_user(user["username"] if user else "", role)

            st.markdown(
                '<div style="margin: 1.25rem 0 .5rem; padding-top: 1rem; '
                'border-top: 1px solid var(--mp-line);"></div>',
                unsafe_allow_html=True,
            )
            st.markdown(
                f'<div style="display:flex;align-items:center;gap:8px;flex-wrap:wrap;'
                f'margin-bottom:10px;">'
                f'<span style="font-size:15px;font-weight:700;letter-spacing:-.01em;'
                f'color:var(--mp-ink);">📋 Hasil Pencarian</span>'
                f'<span class="mp-chip" style="font-family:var(--mp-font-mono);">'
                f'{len(results)} ditemukan</span>'
                f'<span style="font-size:12px;color:var(--mp-ink-50);">untuk</span>'
                f'<span class="mp-chip gray" style="font-family:var(--mp-font-mono);">'
                f'{search_term}</span>'
                f'</div>',
                unsafe_allow_html=True,
            )
            df_res = pd.DataFrame(results)
            # Catatan: kolom "Stok" SELALU total seluruh gudang (untuk semua
            # akun). Stok per-cabang ditampilkan di section terpisah di bawah.

            # Daftar kolom yang ingin ditampilkan, filter Stok/Harga sesuai izin
            candidate_cols = ["File", "Part Number", "Part Name", "Quantity"]
            if "col_stok" in allowed_cols:
                candidate_cols.append("Stok")
            if "col_harga" in allowed_cols:
                candidate_cols.append("Harga")
            candidate_cols += ["Sheet", "Excel Row"]
            cols = [c for c in candidate_cols if c in df_res.columns]

            col_cfg = {
                "File":        st.column_config.TextColumn(width="medium"),
                "Part Number": st.column_config.TextColumn(width="medium"),
                "Part Name":   st.column_config.TextColumn(width="large"),
                "Quantity":    st.column_config.NumberColumn(width="small"),
                "Stok":        st.column_config.TextColumn(width="small"),
                "Harga":       st.column_config.TextColumn(width="medium"),
                "Sheet":       st.column_config.TextColumn(width="medium"),
                "Excel Row":   st.column_config.NumberColumn(width="small"),
            }
            st.dataframe(df_res[cols], hide_index=True,
                         column_config={k: v for k, v in col_cfg.items() if k in cols})

            # ── Stok per cabang (section terpisah di bawah hasil) ─────────
            if "col_stok" in allowed_cols and (self.stok_gudang_cache or {}):
                if gudang_scope is None:
                    # admin / mas → rincian semua gudang
                    self._render_stok_gudang_breakdown(df_res)
                else:
                    # akun cabang → stok gudangnya sendiri + fallback terdekat
                    self._render_stok_cabang_section(df_res, gudang_scope)

            if st.session_state.get("search_type") == "Part Number":
                _emit_image_scroll_anchor()
                st.markdown(
                    '<div style="margin:1rem 0 .5rem;font-size:14px;font-weight:700;'
                    'letter-spacing:-.01em;color:var(--mp-ink);">🖼️ Gambar Part</div>',
                    unsafe_allow_html=True,
                )
                for pn in df_res["Part Number"].dropna().unique():
                    rows     = df_res[df_res["Part Number"] == pn]
                    pname_ex = rows.iloc[0]["Part Name"] if not rows.empty else "N/A"

                    sims_key     = f"sims_fetched_{pn}"
                    sims_err_key = f"sims_err_{pn}"

                    if sims_key not in st.session_state:
                        if SIMS_ENABLED:
                            with st.spinner(f"🔍 Mengambil gambar dari SIMS untuk {pn}..."):
                                fetched_urls, fetch_err = _sims_fetch(pn)
                            st.session_state[sims_key]     = fetched_urls
                            st.session_state[sims_err_key] = fetch_err
                        else:
                            st.session_state[sims_key]     = []
                            st.session_state[sims_err_key] = "SIMS tidak aktif"

                    sims_urls = st.session_state[sims_key]

                    # ── Prioritas foto: Supabase duluan, lalu SIMS ──────
                    supabase_urls = []
                    if FOTO_PART_ENABLED:
                        supabase_urls = get_supabase_photo_urls(pn)
                    img_links = supabase_urls + sims_urls

                    img_path  = self.get_image_path(pn)
                    if img_path and not img_path.exists():
                        img_path = None

                    with st.expander(f"🖼️ {pn}", expanded=True):
                        if SIMS_ENABLED:
                            col_ref, _ = st.columns([1, 4])
                            with col_ref:
                                if st.button("🔄 Refresh dari SIMS", key=f"sims_refresh_{pn}"):
                                    st.session_state.pop(sims_key, None)
                                    st.session_state.pop(sims_err_key, None)
                                    _request_image_scroll()
                                    st.rerun()
                            sims_err = st.session_state.get(sims_err_key)
                            if sims_err and not img_links and not img_path:
                                st.warning(f"⚠️ SIMS: {sims_err}")

                        if img_links:
                            idx_key = f"img_idx_{pn}"
                            if idx_key not in st.session_state:
                                st.session_state[idx_key] = 0

                            total       = len(img_links)
                            current_idx = st.session_state[idx_key]

                            if total > 1:
                                col_prev, col_info, col_next = st.columns([1, 3, 1])
                                with col_prev:
                                    if st.button("◀ Prev", key=f"prev_{pn}", disabled=(current_idx == 0)):
                                        st.session_state[idx_key] = max(0, current_idx - 1)
                                        _request_image_scroll()
                                        st.rerun()
                                with col_info:
                                    st.markdown(
                                        f"<div style='text-align:center; padding:6px 0; font-weight:600; color:#1565C0;'>"
                                        f"Gambar {current_idx + 1} / {total}</div>",
                                        unsafe_allow_html=True
                                    )
                                with col_next:
                                    if st.button("Next ▶", key=f"next_{pn}", disabled=(current_idx == total - 1)):
                                        st.session_state[idx_key] = min(total - 1, current_idx + 1)
                                        _request_image_scroll()
                                        st.rerun()

                            active_url = img_links[current_idx]
                            with st.spinner("Memuat gambar..."):
                                img_bytes, err = ExcelSearchApp.fetch_image_bytes(active_url)
                            if img_bytes:
                                try:
                                    _, col_img, _ = st.columns([1, 2, 1])
                                    with col_img:
                                        ExcelSearchApp.render_zoomable_image(
                                            img_bytes,
                                            caption=f"{pn} - {pname_ex}  (Gambar {current_idx + 1}/{total})",
                                            zoom_key=f"{pn}_{current_idx}"
                                        )
                                except Exception as e:
                                    st.error(f"⚠️ Gambar berhasil diunduh tapi gagal ditampilkan: {e}")
                                    st.caption(f"URL: {active_url}")
                            else:
                                st.warning(f"⚠️ Gagal memuat gambar: {err}")
                                st.caption(f"URL: {active_url}")

                            if total > 1:
                                st.markdown("**Pilih gambar:**")
                                thumb_cols = st.columns(min(total, 5))
                                for ti, (tc, lnk) in enumerate(zip(thumb_cols, img_links)):
                                    with tc:
                                        label = f"{'✅' if ti == current_idx else '🔲'} {ti+1}"
                                        if st.button(label, key=f"thumb_{pn}_{ti}"):
                                            st.session_state[idx_key] = ti
                                            _request_image_scroll()
                                            st.rerun()

                        elif img_path:
                            local_paths   = self.get_all_image_paths(pn)
                            if not local_paths:
                                local_paths = [img_path]
                            local_idx_key = f"local_img_idx_{pn}"
                            if local_idx_key not in st.session_state:
                                st.session_state[local_idx_key] = 0
                            local_total = len(local_paths)
                            local_cur   = min(st.session_state[local_idx_key], local_total - 1)
                            if local_total > 1:
                                col_p, col_i, col_n = st.columns([1, 3, 1])
                                with col_p:
                                    if st.button("◀ Prev", key=f"loc_prev_{pn}", disabled=(local_cur == 0)):
                                        st.session_state[local_idx_key] = max(0, local_cur - 1)
                                        _request_image_scroll()
                                        st.rerun()
                                with col_i:
                                    st.markdown(f"<div style='text-align:center;padding:6px 0;font-weight:600;color:#1565C0;'>Foto {local_cur+1} / {local_total}</div>", unsafe_allow_html=True)
                                with col_n:
                                    if st.button("Next ▶", key=f"loc_next_{pn}", disabled=(local_cur == local_total - 1)):
                                        st.session_state[local_idx_key] = min(local_total - 1, local_cur + 1)
                                        _request_image_scroll()
                                        st.rerun()
                            _, col_img, _ = st.columns([1, 2, 1])
                            with col_img:
                                img_data = local_paths[local_cur].read_bytes()
                                ExcelSearchApp.render_zoomable_image(img_data, caption=f"{pn} - {pname_ex} (Foto {local_cur+1}/{local_total})", zoom_key=f"{pn}_local_{local_cur}")
                            if local_total > 1:
                                st.markdown("**Pilih foto:**")
                                thumb_cols = st.columns(min(local_total, 5))
                                for ti, (tc, lp) in enumerate(zip(thumb_cols, local_paths)):
                                    with tc:
                                        lbl = f"{'✅' if ti == local_cur else '🔲'} {ti+1}"
                                        if st.button(lbl, key=f"loc_thumb_{pn}_{ti}"):
                                            st.session_state[local_idx_key] = ti
                                            _request_image_scroll()
                                            st.rerun()
                        else:
                            if SIMS_ENABLED and st.session_state.get(f"sims_fetched_{pn}") is not None:
                                st.caption("📷 Tidak ada gambar di SIMS untuk part ini")
                            else:
                                st.caption("Tidak ada gambar tersedia")

        elif "search_term" in st.session_state and st.session_state.get("search_results") is not None:
            search_term = st.session_state.search_term
            st.warning(f"❌ Tidak ditemukan hasil untuk '{search_term}' di database lokal")

            if st.session_state.get("search_type") == "Part Number":
                # ── Ambil Part Name dari SIMS dan tampilkan tabel ringkas ──
                sims_info_key = f"sims_part_info_{search_term}"
                if sims_info_key not in st.session_state:
                    if SIMS_ENABLED:
                        try:
                            from sims_fetcher import get_sims_part_info
                            with st.spinner(f"🔎 Mengambil info part dari SIMS..."):
                                sims_info, _ = get_sims_part_info(search_term)
                            st.session_state[sims_info_key] = sims_info
                        except Exception:
                            st.session_state[sims_info_key] = {}
                    else:
                        st.session_state[sims_info_key] = {}

                sims_info = st.session_state.get(sims_info_key, {})
                part_name_sims = sims_info.get("partName", "") if sims_info else ""

                st.markdown("---")
                st.markdown("#### 📋 Info Part")
                df_info = pd.DataFrame([{
                    "Part Number": search_term,
                    "Part Name":   part_name_sims if part_name_sims else "—",
                    "File":        "—",
                    "Sheet":       "—",
                    "Qty":         "—",
                    "Stok":        "—",
                    "Status":      "❌ Tidak ditemukan di database lokal",
                }])
                st.dataframe(
                    df_info[["Part Number", "Part Name", "File", "Sheet", "Qty", "Stok", "Status"]],
                    hide_index=True,
                    use_container_width=True,
                    column_config={
                        "Part Number": st.column_config.TextColumn(width="medium"),
                        "Part Name":   st.column_config.TextColumn(width="large"),
                        "File":        st.column_config.TextColumn(width="medium"),
                        "Sheet":       st.column_config.TextColumn(width="medium"),
                        "Qty":         st.column_config.TextColumn(width="small"),
                        "Stok":        st.column_config.TextColumn(width="small"),
                        "Status":      st.column_config.TextColumn(width="large"),
                    }
                )
                st.markdown("---")

                sims_key     = f"sims_fetched_{search_term}"
                sims_err_key = f"sims_err_{search_term}"

                if sims_key not in st.session_state:
                    if SIMS_ENABLED:
                        with st.spinner(f"🔍 Mengambil gambar dari SIMS untuk {search_term}..."):
                            fetched_urls, fetch_err = _sims_fetch(search_term)
                        st.session_state[sims_key]     = fetched_urls
                        st.session_state[sims_err_key] = fetch_err
                    else:
                        st.session_state[sims_key]     = []
                        st.session_state[sims_err_key] = "SIMS tidak aktif"

                img_links = st.session_state[sims_key]
                img_path  = self.get_image_path(search_term)
                if img_path and not img_path.exists():
                    img_path = None

                _emit_image_scroll_anchor()
                st.markdown(
                    '<div style="margin:1rem 0 .5rem;font-size:14px;font-weight:700;'
                    'letter-spacing:-.01em;color:var(--mp-ink);">🖼️ Gambar Part</div>',
                    unsafe_allow_html=True,
                )
                with st.expander(f"🖼️ {search_term}", expanded=True):
                    if SIMS_ENABLED:
                        col_ref, _ = st.columns([1, 4])
                        with col_ref:
                            if st.button("🔄 Refresh dari SIMS", key=f"nf_sims_refresh_{search_term}"):
                                st.session_state.pop(sims_key, None)
                                st.session_state.pop(sims_err_key, None)
                                _request_image_scroll()
                                st.rerun()

                    if img_links or img_path:
                        if img_links:
                            idx_key = f"img_idx_{search_term}"
                            if idx_key not in st.session_state:
                                st.session_state[idx_key] = 0

                            total       = len(img_links)
                            current_idx = st.session_state[idx_key]

                            if total > 1:
                                col_prev, col_info, col_next = st.columns([1, 3, 1])
                                with col_prev:
                                    if st.button("◀ Prev", key=f"nf_prev_{search_term}", disabled=(current_idx == 0)):
                                        st.session_state[idx_key] = max(0, current_idx - 1)
                                        _request_image_scroll()
                                        st.rerun()
                                with col_info:
                                    st.markdown(
                                        f"<div style='text-align:center; padding:6px 0; font-weight:600; color:#1565C0;'>"
                                        f"Gambar {current_idx + 1} / {total}</div>",
                                        unsafe_allow_html=True
                                    )
                                with col_next:
                                    if st.button("Next ▶", key=f"nf_next_{search_term}", disabled=(current_idx == total - 1)):
                                        st.session_state[idx_key] = min(total - 1, current_idx + 1)
                                        _request_image_scroll()
                                        st.rerun()

                            active_url = img_links[current_idx]
                            with st.spinner("Memuat gambar..."):
                                img_bytes, err = ExcelSearchApp.fetch_image_bytes(active_url)
                            if img_bytes:
                                try:
                                    _, col_img, _ = st.columns([1, 2, 1])
                                    with col_img:
                                        ExcelSearchApp.render_zoomable_image(
                                            img_bytes,
                                            caption=f"{search_term}  (Gambar {current_idx + 1}/{total})",
                                            zoom_key=f"nf_{search_term}_{current_idx}"
                                        )
                                except Exception as e:
                                    st.error(f"⚠️ Gambar berhasil diunduh tapi gagal ditampilkan: {e}")
                                    st.caption(f"URL: {active_url}")
                            else:
                                st.warning(f"⚠️ Gagal memuat gambar: {err}")
                                st.caption(f"URL: {active_url}")

                            if total > 1:
                                st.markdown("**Pilih gambar:**")
                                thumb_cols = st.columns(min(total, 5))
                                for ti, (tc, lnk) in enumerate(zip(thumb_cols, img_links)):
                                    with tc:
                                        label = f"{'✅' if ti == current_idx else '🔲'} {ti+1}"
                                        if st.button(label, key=f"nf_thumb_{search_term}_{ti}"):
                                            st.session_state[idx_key] = ti
                                            _request_image_scroll()
                                            st.rerun()

                        elif img_path:
                            local_paths_nf = self.get_all_image_paths(search_term)
                            if not local_paths_nf:
                                local_paths_nf = [img_path]
                            nf_local_idx_key = f"local_img_idx_{search_term}"
                            if nf_local_idx_key not in st.session_state:
                                st.session_state[nf_local_idx_key] = 0
                            nf_total = len(local_paths_nf)
                            nf_cur   = min(st.session_state[nf_local_idx_key], nf_total - 1)
                            if nf_total > 1:
                                col_p, col_i, col_n = st.columns([1, 3, 1])
                                with col_p:
                                    if st.button("◀ Prev", key=f"nf_loc_prev_{search_term}", disabled=(nf_cur == 0)):
                                        st.session_state[nf_local_idx_key] = max(0, nf_cur - 1)
                                        _request_image_scroll()
                                        st.rerun()
                                with col_i:
                                    st.markdown(f"<div style='text-align:center;padding:6px 0;font-weight:600;color:#1565C0;'>Foto {nf_cur+1} / {nf_total}</div>", unsafe_allow_html=True)
                                with col_n:
                                    if st.button("Next ▶", key=f"nf_loc_next_{search_term}", disabled=(nf_cur == nf_total - 1)):
                                        st.session_state[nf_local_idx_key] = min(nf_total - 1, nf_cur + 1)
                                        _request_image_scroll()
                                        st.rerun()
                            _, col_img, _ = st.columns([1, 2, 1])
                            with col_img:
                                img_data = local_paths_nf[nf_cur].read_bytes()
                                ExcelSearchApp.render_zoomable_image(img_data, caption=f"{search_term} (Foto {nf_cur+1}/{nf_total})", zoom_key=f"nf_{search_term}_local_{nf_cur}")
                            if nf_total > 1:
                                st.markdown("**Pilih foto:**")
                                thumb_cols = st.columns(min(nf_total, 5))
                                for ti, (tc, lp) in enumerate(zip(thumb_cols, local_paths_nf)):
                                    with tc:
                                        lbl = f"{'✅' if ti == nf_cur else '🔲'} {ti+1}"
                                        if st.button(lbl, key=f"nf_loc_thumb_{search_term}_{ti}"):
                                            st.session_state[nf_local_idx_key] = ti
                                            _request_image_scroll()
                                            st.rerun()
                    else:
                        sims_err = st.session_state.get(sims_err_key)
                        if sims_err:
                            st.warning(f"⚠️ SIMS: {sims_err}")
                        else:
                            st.caption("📷 Tidak ada gambar di SIMS untuk part ini")

    def _load_populasi_data(self):
        if "populasi_df" in st.session_state:
            return st.session_state.populasi_df

        excel_ext = (".xlsx", ".xls", ".xlsm")
        frames    = []

        # ── 1. Coba dari Supabase Storage (file utama: populasi.xlsx) ────
        try:
            from admin_data_uploader import download_dataset
            file_bytes = download_dataset("populasi")
            if file_bytes:
                xl = pd.ExcelFile(io.BytesIO(file_bytes), engine="openpyxl")
                for sheet in xl.sheet_names:
                    df = pd.read_excel(xl, sheet_name=sheet, dtype=str)
                    df.columns = [str(c).strip() for c in df.columns]
                    df["_source_file"]  = "populasi.xlsx"
                    df["_source_sheet"] = sheet
                    frames.append(df)
                print("[populasi] ✅ populasi.xlsx diunduh dari Supabase Storage.")
        except Exception as e:
            print(f"[populasi] ⚠️ Gagal download dari Supabase: {e}")

        # ── 2. Fallback ke folder lokal (mendukung multi-file lama) ──────
        if not frames and self.populasi_folder.exists():
            for fp in sorted(self.populasi_folder.iterdir()):
                if fp.suffix.lower() not in excel_ext:
                    continue
                try:
                    with open(fp, "rb") as f:
                        file_bytes = io.BytesIO(f.read())
                    xl = pd.ExcelFile(file_bytes, engine="openpyxl")
                    for sheet in xl.sheet_names:
                        df = pd.read_excel(xl, sheet_name=sheet, dtype=str)
                        df.columns = [str(c).strip() for c in df.columns]
                        df["_source_file"]  = fp.name
                        df["_source_sheet"] = sheet
                        frames.append(df)
                except Exception as e:
                    st.warning(f"Gagal membaca {fp.name}: {e}")

        combined = pd.concat(frames, ignore_index=True) if frames else pd.DataFrame()
        st.session_state.populasi_df = combined
        return combined

    def render_populasi_tab(self):
        st.markdown("### 🚛 Populasi Unit")

        if st.button("🔄 Refresh", key="refresh_populasi"):
            st.session_state.pop("populasi_df", None)
            st.rerun()

        df = self._load_populasi_data()

        if df.empty:
            st.warning("Tidak ada file Excel di folder data/populasi/. Pastikan file populasi sudah ditempatkan di folder tersebut.")
            return

        display_cols = [c for c in df.columns if not c.startswith("_source")]
        df_display   = df[display_cols].copy()

        with st.expander("🔍 Filter & Pencarian", expanded=True):
            search_col, filter_area = st.columns([2, 3])
            with search_col:
                keyword = st.text_input(
                    "Cari (semua kolom):", placeholder="Ketik kata kunci",
                    key="pop_keyword",
                    value=st.session_state.get("pop_keyword_val", ""),
                )
                st.session_state["pop_keyword_val"] = keyword
            with filter_area:
                fcols             = st.columns(2)
                filter_vals       = {}
                candidate_filters = ["MODEL", "JENIS", "TIPE UNIT", "LOKASI KERJA", "TAHUN", "Euro"]
                available_filters = [c for c in candidate_filters if c in df_display.columns][:4]
                for i, col in enumerate(available_filters):
                    with fcols[i % 2]:
                        options = ["Semua"] + sorted(df_display[col].dropna().unique().tolist())
                        sk      = f"pop_filter_{col}"
                        saved   = st.session_state.get(sk, "Semua")
                        if saved not in options:
                            saved = "Semua"
                        filter_vals[col] = st.selectbox(col, options, index=options.index(saved), key=sk)

        mask = pd.Series([True] * len(df_display), index=df_display.index)
        if keyword.strip():
            kw      = keyword.strip().upper()
            kw_mask = pd.Series([False] * len(df_display), index=df_display.index)
            for col in df_display.columns:
                kw_mask |= df_display[col].astype(str).str.upper().str.contains(kw, na=False)
            mask &= kw_mask
        for col, val in filter_vals.items():
            if val != "Semua":
                mask &= (df_display[col].astype(str) == val)

        df_filtered = df_display[mask].reset_index(drop=True)

        c1, c2 = st.columns(2)
        c1.metric("Total Unit", len(df_display))
        c2.metric("Hasil Filter", len(df_filtered))
        st.markdown("---")

        if df_filtered.empty:
            st.info("Tidak ada data yang cocok dengan filter.")
        else:
            df_show = df_filtered.rename(columns=lambda c: c.strip())
            st.dataframe(df_show, hide_index=True, use_container_width=True, height=500)
            dl_buf = io.BytesIO()
            df_show.to_excel(dl_buf, index=False, engine="openpyxl")
            dl_buf.seek(0)
            st.download_button(
                label="⬇️ Download Excel",
                data=dl_buf.getvalue(),
                file_name=f"populasi_unit_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="pop_download",
            )

    def run(self):
        self.display_dashboard()


def main():
    LoginManager.init_session()
    login_mgr = LoginManager()
    if not LoginManager.is_authenticated():
        render_login_page(login_mgr)
    else:
        ExcelSearchApp().run()


if __name__ == "__main__":
    main()
